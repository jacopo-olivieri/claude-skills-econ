"""U8 lint enforcement tests (failing-first per KTD-6).

Four families of new checks, each a negative/positive pair:

* (a) b4 inventory recall guarantee — required IDs, wrong-typed IDs, cluster
      size, and the deep single-substantive-ID rule.
* (b) b3b ``--shard`` mode — allocation membership, ID ranges, allowed
      statuses, blank cross-links, footer.
* (c) evidence-ladder ceiling, empty ``Evidence Checked``, blocker text on
      blocked/confirmation_needed/deferred verdicts, and code-row completeness.
* (d) b9 export parity — header parity, ``Potential Issue`` biconditional,
      and the populated-``_staging/`` requirement.
"""

import openpyxl

import regbuild as rb


def fails(res, needle):
    return res.returncode == 1 and needle in res.stdout


# =============================================================== (a) b4 recall


def test_b4_code_green_on_complete_inventory(tmp_path):
    """Positive: every candidate + every confirmed-sev>=3 row is in the plan."""
    errors = [
        rb.error_row("E-0101", status="candidate", severity="3"),
        rb.error_row("E-0102", status="confirmed", severity="4"),
        rb.error_row("E-0103", status="confirmed", severity="1"),  # not required
        rb.error_row("E-0104", status="not_error", severity=""),   # not required
    ]
    a = rb.make_b4(tmp_path, "code", canon_errors=errors)
    res = rb.lint(a, "b4-code")
    assert res.returncode == 0, res.stdout + res.stderr
    assert "LINT PASS [b4-code]" in res.stdout


def test_b4_code_red_when_a_candidate_is_omitted(tmp_path):
    """Negative: a candidate dropped from the inventory FAILS (recall gap)."""
    errors = [
        rb.error_row("E-0101", status="candidate", severity="3"),
        rb.error_row("E-0102", status="candidate", severity="2"),
    ]
    # inventory + cluster mention only E-0101; E-0102 is the dropped candidate
    inv = [("E-0101", "candidate", "static")]
    clu = [("K1", "c", "E-0101", "`audit/_code_error_recheck/k1.md`")]
    a = rb.make_b4(tmp_path, "code", canon_errors=errors,
                   inventory=inv, clusters=clu)
    res = rb.lint(a, "b4-code")
    assert fails(res, "E-0102"), res.stdout
    assert "required" in res.stdout.lower()


def test_b4_claims_red_when_unclear_row_omitted(tmp_path):
    """A severity-bearing / unclear claim row is required in the inventory."""
    claims = [
        rb.claims_row("C-0101", status="inconsistent", severity="3",
                      issue="conflict"),
        rb.claims_row("C-0102", status="unclear"),
    ]
    inv = [("C-0101", "issue", "static")]
    clu = [("K1", "c", "C-0101", "`audit/_recheck/k1.md`")]
    a = rb.make_b4(tmp_path, "claims", canon_claims=claims,
                   inventory=inv, clusters=clu)
    res = rb.lint(a, "b4-claims")
    assert fails(res, "C-0102"), res.stdout


def test_b4_code_red_on_wrong_typed_inventory_id(tmp_path):
    """A claim-shaped ID in a code recheck inventory FAILS (wrong-typed)."""
    errors = [rb.error_row("E-0101", status="candidate", severity="3")]
    inv = [("E-0101", "candidate", "static"), ("C-0142", "stray", "static")]
    clu = [("K1", "c", "E-0101; C-0142", "`audit/_code_error_recheck/k1.md`")]
    a = rb.make_b4(tmp_path, "code", canon_errors=errors,
                   inventory=inv, clusters=clu)
    res = rb.lint(a, "b4-code")
    assert fails(res, "C-0142"), res.stdout
    assert "wrong-typed" in res.stdout.lower() or "wrong type" in res.stdout.lower()


def test_b4_code_red_on_oversized_cluster(tmp_path):
    """A cluster grouping more than 8 IDs FAILS at any depth."""
    errors = [rb.error_row(f"E-01{n:02d}", status="candidate", severity="2")
              for n in range(9)]  # 9 candidates
    ids = [r[0] for r in errors]
    inv = [(i, "candidate", "static") for i in ids]
    clu = [("K1", "c", "; ".join(ids), "`audit/_code_error_recheck/k1.md`")]
    a = rb.make_b4(tmp_path, "code", canon_errors=errors,
                   inventory=inv, clusters=clu)
    res = rb.lint(a, "b4-code")
    assert res.returncode == 1
    assert "9" in res.stdout and "cluster" in res.stdout.lower()


def test_b4_code_deep_red_on_two_substantive_in_one_cluster(tmp_path):
    """At deep, a cluster with two SUBSTANTIVE (candidate) IDs FAILS."""
    errors = [
        rb.error_row("E-0101", status="candidate", severity="2"),
        rb.error_row("E-0102", status="candidate", severity="2"),
    ]
    inv = [(r[0], "candidate", "static") for r in errors]
    clu = [("K1", "c", "E-0101; E-0102", "`audit/_code_error_recheck/k1.md`")]
    a = rb.make_b4(tmp_path, "code", canon_errors=errors,
                   inventory=inv, clusters=clu, review_depth="deep")
    res = rb.lint(a, "b4-code")
    assert res.returncode == 1
    assert "substantive" in res.stdout.lower()


def test_b4_code_deep_green_on_singleton_substantive_clusters(tmp_path):
    """At deep, one substantive ID per cluster passes; clean confirmed rows
    (non-substantive) may still be grouped."""
    errors = [
        rb.error_row("E-0101", status="candidate", severity="2"),
        rb.error_row("E-0102", status="candidate", severity="2"),
        rb.error_row("E-0201", status="confirmed", severity="1"),  # sampled clean
        rb.error_row("E-0202", status="confirmed", severity="1"),  # sampled clean
    ]
    inv = [
        ("E-0101", "candidate", "static"),
        ("E-0102", "candidate", "static"),
        ("E-0201", "sampled clean", "static"),
        ("E-0202", "sampled clean", "static"),
    ]
    clu = [
        ("K1", "c1", "E-0101", "`audit/_code_error_recheck/k1.md`"),
        ("K2", "c2", "E-0102", "`audit/_code_error_recheck/k2.md`"),
        ("K3", "clean", "E-0201; E-0202", "`audit/_code_error_recheck/k3.md`"),
    ]
    a = rb.make_b4(tmp_path, "code", canon_errors=errors,
                   inventory=inv, clusters=clu, review_depth="deep")
    res = rb.lint(a, "b4-code")
    assert res.returncode == 0, res.stdout + res.stderr


# =============================================================== (b) b3b shard


def test_b3b_code_shard_green(tmp_path):
    """Positive: a well-formed code second-read shard passes --shard."""
    rows = [rb.error_row("E-2001", status="candidate", severity="2",
                         related="")]
    a, shard = rb.make_b3b_shard(tmp_path, "code", error_rows=rows)
    res = rb.lint(a, "b3b-code", shard=shard)
    assert res.returncode == 0, res.stdout + res.stderr
    assert "LINT PASS [b3b-code]" in res.stdout


def test_b3b_code_shard_red_on_confirmed_status(tmp_path):
    """Negative: a code second-read row that is not 'candidate' FAILS."""
    rows = [rb.error_row("E-2001", status="confirmed", severity="2")]
    a, shard = rb.make_b3b_shard(tmp_path, "code", error_rows=rows)
    res = rb.lint(a, "b3b-code", shard=shard)
    assert res.returncode == 1
    assert "E-2001" in res.stdout and "candidate" in res.stdout


def test_b3b_claims_shard_green(tmp_path):
    rows = [rb.claims_row("C-2001", status="inconsistent", severity="3",
                          issue="conflict")]
    a, shard = rb.make_b3b_shard(tmp_path, "claims", claims_rows=rows)
    res = rb.lint(a, "b3b-claims", shard=shard)
    assert res.returncode == 0, res.stdout + res.stderr


def test_b3b_claims_shard_red_on_confirmed_claim(tmp_path):
    """A confirmed claim in a second-read shard FAILS (only inconsistent/unclear)."""
    rows = [rb.claims_row("C-2001", status="confirmed")]
    a, shard = rb.make_b3b_shard(tmp_path, "claims", claims_rows=rows)
    res = rb.lint(a, "b3b-claims", shard=shard)
    assert res.returncode == 1
    assert "C-2001" in res.stdout


def test_b3b_claims_shard_red_without_output_table(tmp_path):
    """A claims second-read shard must carry both tables (claims, then an outputs
    table — header-only if there are no new output rows). Omitting the outputs
    table FAILS, so the strict lint and the second-read-worker skeleton's
    two-table mandate agree (mirrors the first-pass b2 shard contract)."""
    rows = [rb.claims_row("C-2001", status="inconsistent", severity="3",
                          issue="conflict")]
    a, shard = rb.make_b3b_shard(tmp_path, "claims", claims_rows=rows,
                                 omit_output_table=True)
    res = rb.lint(a, "b3b-claims", shard=shard)
    assert res.returncode == 1
    assert "outputs" in res.stdout.lower()


def test_b3b_code_shard_red_on_out_of_range_id(tmp_path):
    rows = [rb.error_row("E-5000", status="candidate", severity="2")]
    a, shard = rb.make_b3b_shard(tmp_path, "code", error_rows=rows)
    res = rb.lint(a, "b3b-code", shard=shard)
    assert res.returncode == 1
    assert "E-5000" in res.stdout and "range" in res.stdout.lower()


def test_b3b_code_shard_red_on_nonblank_crosslink(tmp_path):
    rows = [rb.error_row("E-2001", status="candidate", severity="2",
                         related="C-0142")]
    a, shard = rb.make_b3b_shard(tmp_path, "code", error_rows=rows)
    res = rb.lint(a, "b3b-code", shard=shard)
    assert res.returncode == 1
    assert "Related Claim IDs" in res.stdout


def test_b3b_code_shard_red_on_missing_footer(tmp_path):
    rows = [rb.error_row("E-2001", status="candidate", severity="2")]
    a, shard = rb.make_b3b_shard(tmp_path, "code", error_rows=rows)
    # strip the footer
    body = shard.read_text(encoding="utf-8").split("### Coverage")[0]
    shard.write_text(body, encoding="utf-8")
    res = rb.lint(a, "b3b-code", shard=shard)
    assert res.returncode == 1
    assert "footer" in res.stdout.lower() or "coverage" in res.stdout.lower()


# =============================================================== (c) evidence


def test_b5_code_green_on_static_ladder1(tmp_path):
    rows = [rb.ledger_row("E-0101", status="candidate", severity="3",
                          level="static_source_verified",
                          verdict="confirmed_error")]
    a, shard = rb.make_b5(tmp_path, "code", ledger_rows=rows, ladder_level=1)
    res = rb.lint(a, "b5-code", shard=shard)
    assert res.returncode == 0, res.stdout + res.stderr


def test_b5_code_red_on_runtime_level_at_ladder1(tmp_path):
    """A ladder-1 ledger claiming parser_or_runtime_verified (min ladder 2) FAILS."""
    rows = [rb.ledger_row("E-0101", status="candidate", severity="3",
                          level="parser_or_runtime_verified",
                          verdict="confirmed_error")]
    a, shard = rb.make_b5(tmp_path, "code", ledger_rows=rows, ladder_level=1)
    res = rb.lint(a, "b5-code", shard=shard)
    assert res.returncode == 1
    assert "parser_or_runtime_verified" in res.stdout
    assert "ladder" in res.stdout.lower()


def test_b5_code_green_on_runtime_level_at_ladder2(tmp_path):
    """The same evidence level passes when the ladder permits it."""
    rows = [rb.ledger_row("E-0101", status="candidate", severity="3",
                          level="parser_or_runtime_verified",
                          verdict="confirmed_error")]
    a, shard = rb.make_b5(tmp_path, "code", ledger_rows=rows, ladder_level=2)
    res = rb.lint(a, "b5-code", shard=shard)
    assert res.returncode == 0, res.stdout + res.stderr


def test_b5_code_red_on_empty_evidence_checked(tmp_path):
    rows = [rb.ledger_row("E-0101", status="candidate", severity="3",
                          evidence="", level="static_source_verified",
                          verdict="confirmed_error")]
    a, shard = rb.make_b5(tmp_path, "code", ledger_rows=rows, ladder_level=1)
    res = rb.lint(a, "b5-code", shard=shard)
    assert res.returncode == 1
    assert "Evidence Checked" in res.stdout and "E-0101" in res.stdout


def test_b5_code_red_on_blocked_verdict_without_blocker_text(tmp_path):
    """A blocked/deferred verdict with no text in either proposal column FAILS."""
    rows = [rb.ledger_row("E-0101", status="blocked", severity="3",
                          level="blocked_documented", verdict="blocked",
                          change="", note="")]
    a, shard = rb.make_b5(tmp_path, "code", ledger_rows=rows, ladder_level=1)
    res = rb.lint(a, "b5-code", shard=shard)
    assert res.returncode == 1
    assert "blocker" in res.stdout.lower() and "E-0101" in res.stdout


def test_b5_code_green_on_blocked_verdict_with_blocker_note(tmp_path):
    rows = [rb.ledger_row("E-0101", status="blocked", severity="3",
                          level="blocked_documented", verdict="blocked",
                          change="",
                          note="restricted data — the raw census is off-limits")]
    a, shard = rb.make_b5(tmp_path, "code", ledger_rows=rows, ladder_level=1)
    res = rb.lint(a, "b5-code", shard=shard)
    assert res.returncode == 0, res.stdout + res.stderr


# ------- code-row completeness (register-level, fires wherever error rows load)


def test_error_row_red_on_empty_error_description(tmp_path):
    """A candidate code row with empty Error Description FAILS (b3-code)."""
    a = rb.AuditDir(tmp_path)
    a.write_manifest()
    a.write("plans/code_error_review_plan.md", rb._code_b1_plan())
    row = rb.error_row("E-0101", status="candidate", severity="3", desc="")
    a.write_register("_staging/code_error_register.md", rb.ERROR_COLS, [row],
                     title="Code-error register")
    # a valid merge report so the stage reaches the row check
    a.write("_run/merge_report_code.json",
            '{"code_error_register.md": {"shard_rows": 1, "dedup_removed": 0, '
            '"added": 1, "conflicts": [], "coverage_gaps": [], '
            '"blocked_shards": []}}')
    # coverage: give the script a coverage row so b3 does not fail on coverage
    a.write("_code_errors/k1.md",
            "| Script | Outcome |\n| --- | --- |\n| `py/x.py` | findings: E-0101 |\n")
    res = rb.lint(a, "b3-code")
    assert res.returncode == 1
    assert "Error Description" in res.stdout and "E-0101" in res.stdout


def test_error_row_green_on_complete_candidate(tmp_path):
    a = rb.AuditDir(tmp_path)
    a.write_manifest()
    a.write("plans/code_error_review_plan.md", rb._code_b1_plan())
    row = rb.error_row("E-0101", status="candidate", severity="3")
    a.write_register("_staging/code_error_register.md", rb.ERROR_COLS, [row],
                     title="Code-error register")
    a.write("_run/merge_report_code.json",
            '{"code_error_register.md": {"shard_rows": 1, "dedup_removed": 0, '
            '"added": 1, "conflicts": [], "coverage_gaps": [], '
            '"blocked_shards": []}}')
    a.write("_code_errors/k1.md",
            "| Script | Outcome |\n| --- | --- |\n| `py/x.py` | findings: E-0101 |\n")
    res = rb.lint(a, "b3-code")
    assert res.returncode == 0, res.stdout + res.stderr


def test_error_row_warns_on_empty_code_location_when_blocked(tmp_path):
    """Empty Code Location WARNS (not fails) on a blocked row."""
    a = rb.AuditDir(tmp_path)
    a.write_manifest()
    a.write("plans/code_error_review_plan.md", rb._code_b1_plan())
    row = rb.error_row("E-0101", status="blocked", severity="3", location="")
    a.write_register("_staging/code_error_register.md", rb.ERROR_COLS, [row],
                     title="Code-error register")
    a.write("_run/merge_report_code.json",
            '{"code_error_register.md": {"shard_rows": 1, "dedup_removed": 0, '
            '"added": 1, "conflicts": [], "coverage_gaps": [], '
            '"blocked_shards": []}}')
    a.write("_code_errors/k1.md",
            "| Script | Outcome |\n| --- | --- |\n| `py/x.py` | blocked: restricted |\n")
    res = rb.lint(a, "b3-code")
    assert res.returncode == 0, res.stdout + res.stderr
    assert "Code Location" in res.stdout  # a WARNING line


# =============================================================== (d) b9 parity


def _b9_claims():
    return [
        rb.claims_row("C-0101", status="confirmed"),
        rb.claims_row("C-0102", status="inconsistent", severity="3",
                      issue="the paper and artifact disagree"),
    ]


def _b9_errors():
    return [rb.error_row("E-0101", status="confirmed", severity="2")]


def test_b9_green_on_faithful_export(tmp_path):
    a = rb.make_b9(tmp_path, claims_rows=_b9_claims(), error_rows=_b9_errors())
    res = rb.lint(a, "b9")
    assert res.returncode == 0, res.stdout + res.stderr
    assert "LINT PASS [b9]" in res.stdout


def test_b9_red_on_stray_potential_issue(tmp_path):
    """A Paper Claims row whose Potential Issue disagrees with severity FAILS."""
    a = rb.make_b9(tmp_path, claims_rows=_b9_claims(), error_rows=_b9_errors())
    wb_path = a.audit / "code_review.xlsx"
    wb = openpyxl.load_workbook(wb_path)
    ws = wb["Paper Claims"]
    headers = [c.value for c in ws[1]]
    pi_col = headers.index("Potential Issue") + 1
    id_col = headers.index("Claim ID") + 1
    # C-0101 has empty Severity -> Potential Issue should be FALSE; flip to TRUE
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=id_col).value == "C-0101":
            ws.cell(row=r, column=pi_col, value="TRUE")
    wb.save(wb_path)
    res = rb.lint(a, "b9")
    assert res.returncode == 1
    assert "Potential Issue" in res.stdout and "C-0101" in res.stdout


def test_b9_red_on_header_mismatch(tmp_path):
    """A renamed/added export header (parity break) FAILS."""
    a = rb.make_b9(tmp_path, claims_rows=_b9_claims(), error_rows=_b9_errors())
    wb_path = a.audit / "code_review.xlsx"
    wb = openpyxl.load_workbook(wb_path)
    ws = wb["Code Errors"]
    ws.cell(row=1, column=2, value="Renamed Column")  # break Error Type header
    wb.save(wb_path)
    res = rb.lint(a, "b9")
    assert res.returncode == 1
    assert "header" in res.stdout.lower()


def test_b9_red_on_empty_staging(tmp_path):
    """A done-b8 with empty _staging/ (so export would precede b8) FAILS."""
    a = rb.make_b9(tmp_path, claims_rows=_b9_claims(), error_rows=_b9_errors())
    # empty out the frozen staging registers
    for f in ("claims_register.md", "code_error_register.md"):
        (a.audit / "_staging" / f).write_text("", encoding="utf-8")
    res = rb.lint(a, "b9")
    assert res.returncode == 1
    assert "_staging" in res.stdout or "staging" in res.stdout.lower()
