#!/usr/bin/env python3
"""Boundary lint for research-codebase-audit registers, plans, and shards.

Mechanical enforcement of ``references/registers.md`` at every stage boundary.
Exits 0 on pass (warnings allowed), nonzero with a findings report on failure.

Usage:
    lint_registers.py --stage STAGE [--shard PATH] [--audit-dir audit]

Stages: b0, b1-claims, b1-code, b2-claims, b2-code, b3-claims, b3-code,
        b4-claims, b4-code, b5-claims, b5-code, b6a-claims, b6a-code,
        b5s-claims, b5s-code, b6b-claims, b6b-code, bC, b7, b8, b9
(b2/b5 lint one worker shard, passed with --shard; b3b-claims/b3b-code lint a
second-read shard with --shard, or the second-read merge without it).
"""

import argparse
import hashlib
import json
import re
import subprocess
import sys
from pathlib import Path

import definition_use as du
import build_detector_mapping as detector_mapping
import mechanism_schema as mechanism
from anchor_resolver import AnchorError, contains as anchor_contains, resolve_quote
from paper_sources import PaperSourceError, validate_source_set

# --------------------------------------------------------------- constants

CLAIMS_COLS = [
    "Claim ID", "Paper Context", "Paper Quote", "Used in Text", "Claim Type",
    "Claim Text", "Code/Data Source", "Output IDs", "Status", "Severity",
    "Issue Description", "Blocked Check", "Related Error IDs",
]
OUTPUT_COLS = [
    "Output ID", "Paper Object", "Paper Context", "Paper Location",
    "Output Path/Pattern", "Producing Script", "Input Dataset(s)",
    "Key Spec/Sample", "Claim IDs", "Status",
]
ERROR_COLS = [
    "Error ID", "Error Type", "Code/Data Source", "Code Location", "Status",
    "Severity", "Error Description", "Why It Matters", "Related Claim IDs",
]
LEDGER_COLS = [
    "ID", "Current Status", "Current Severity", "Evidence Checked",
    "Evidence Level", "Verdict", "Proposed Register Change",
    "Pipeline/Output Impact", "Proposed Note",
]
CODE_LEDGER_COLS = LEDGER_COLS + [
    "Proposed Status", "Proposed Severity", "Accepted Error Type",
    "Accepted Mechanism", "Outcome Witness IDs", "Duplicate Target",
    "Proposed Field Patches", "Verification Record IDs",
]
WITNESS_OUTCOME_COLS = list(mechanism.PRE_BOUNDARY_WITNESS_COLUMNS)
POST_WITNESS_COLS = list(mechanism.POST_BOUNDARY_WITNESS_COLUMNS)
MF_VERIFICATION_COLS = [
    "Channel", "Record ID", "Source ID", "Witness ID",
    "File Digest (sha256)", "Consumer", "Consumer Version", "Invocation",
    "Observed Result", "Whole-File Acceptance (yes/no)",
]
PROBE_VERIFICATION_COLS = [
    "Channel", "Record ID", "Source ID", "Witness ID",
    "Proposition Tested", "Harness / Input Domain", "Observed Result",
    "Scope Anchor",
]
RECEIPT_COLS = [
    "Channel", "Receipt ID", "Source ID", "Witness ID", "Record ID",
    "Tool", "Tool Version", "Input Digest (sha256)", "Invocation",
    "Exit Status", "Accepted (yes/no)", "Result Digest (sha256)",
]
LINEAGE_COLS = [
    "Original Error ID", "Descendant Error ID", "Channel", "Source ID",
    "Witness ID",
]
PATCHABLE_ERROR_FIELDS = {
    "Code Location", "Code/Data Source", "Error Description", "Why It Matters",
}

CLAIM_TYPES = {
    "quantitative_result", "sample_count", "treatment_definition",
    "estimation_specification", "robustness", "data_construction",
    "interpretation", "transcription", "rounding_or_precision",
}
ERROR_TYPES = {
    "syntax_or_parse_error", "missing_input_or_output", "stale_or_wrong_path",
    "undefined_variable_or_global", "merge_key_or_cardinality_error",
    "sample_filter_or_flag_error", "treatment_or_event_timing_error",
    "aggregation_or_unit_error", "output_label_or_path_mismatch",
    "version_or_dependency_error", "randomness_or_seed_error",
    "inference_or_se_specification", "weighting_error",
    "readme_or_package_mismatch", "pii_or_disclosure_risk",
}
CLAIMS_STATUS_FIRST = {"confirmed", "mapped", "unclear", "inconsistent", "blocked"}
CLAIMS_STATUS_FINAL = CLAIMS_STATUS_FIRST | {"confirmation_needed"}
OUTPUT_STATUS_FIRST = {"listed", "mapped", "confirmed", "orphan", "inconsistent", "unclear"}
OUTPUT_STATUS_FINAL = OUTPUT_STATUS_FIRST - {"listed"}
ERROR_STATUS_FIRST = {"candidate", "confirmed", "not_error", "blocked"}
ERROR_STATUS_FINAL = (ERROR_STATUS_FIRST | {"confirmation_needed"})

CLAIMS_VERDICTS = {
    "substantiated", "substantiated_but_reframe", "row_note_only",
    "not_substantiated", "confirmation_needed", "blocked",
}
ERROR_VERDICTS = {
    "confirmed_error", "not_error", "duplicate", "confirmation_needed",
    "blocked", "deferred",
}
EVIDENCE_LEVELS = {
    "static_source_verified", "artifact_verified", "data_inspected_verified",
    "parser_or_runtime_verified", "synthetic_test_verified",
    "targeted_rerun_verified", "blocked_documented",
}
# U8 (c): the evidence-level → minimum-ladder table (registers.md, section
# "Evidence levels (tied to the review ladder)"), mirrored as a dict. An
# evidence level whose minimum ladder EXCEEDS the manifest's `ladder_level`
# cannot have been produced at this run's ladder and FAILS. `targeted_rerun_
# verified` is 2 for small reruns / 3 for expensive ones; the linter mirrors the
# floor (2) it can enforce statically. `blocked_documented` is valid at any level.
EVIDENCE_MIN_LADDER = {
    "static_source_verified": 1,
    "artifact_verified": 1,
    "data_inspected_verified": 1,
    "parser_or_runtime_verified": 2,
    "synthetic_test_verified": 2,
    "targeted_rerun_verified": 2,
    "blocked_documented": 0,
}

# b3c shared-conventions artifact (advisory; consumed by the b4-code recheck grep).
CONVENTIONS_COLS = ["Convention", "Category", "Stated Definition", "Sites Already Seen"]
CONVENTION_CATEGORIES = {
    "fiscal_year_or_sample_window_boundary", "date_parse_mask",
    "missing_value_sentinel", "unit_or_scale_factor", "path_separator",
    "id_or_merge_key", "enumerated_member_list",
}
ID_RE = {"C": r"C-\d{4}", "O": r"O-\d{4}", "E": r"E-\d{4}"}
RANGE_RE = re.compile(r"([CEO]-\d{4})\s*[–—-]\s*([CEO]-\d{4})")
COORD_RE = re.compile(r"Merge-coordinator range:\s*([CEO]-\d{4})\s*[–—-]\s*([CEO]-\d{4})")
FOOTER_COLS = ["Entry ID", "Kind", "Register IDs", "Observation", "Reason"]
FOOTER_KINDS = {"candidate", "not_rowed_observation"}
FOOTER_ENTRY_RE = re.compile(r"OBS-(\d{4})")
COVERAGE_COLS = ["Script", "Outcome"]
COVERAGE_CLEAN = "clean"
COVERAGE_FINDINGS_RE = re.compile(
    r"findings:\s*((?:[CEO]-\d{4})(?:\s*[;,]\s*[CEO]-\d{4})*)$"
)
COVERAGE_BLOCKED_RE = re.compile(r"blocked:\s*(\S.*)$")
HYGIENE_SINGLETON = "@hygiene:data-and-log-lens"
SECOND_READ_PLAN_COLS = {
    "claims": [
        "Worker ID", "File/Section Scope", "Shard File", "Claim ID Range",
        "Output ID Range", "Reason", "Known Findings", "Assigned Handoff IDs",
    ],
    "code": [
        "Worker ID", "Script Scope", "Shard File", "Error ID Range", "Reason",
        "Known Findings", "Assigned Handoff IDs",
    ],
}
SECOND_READ_REASONS = {"detector", "flagged", "clean_sample", "handoff"}
SUPPLEMENTARY_PLAN_FILES = {
    "claims": "claims_supplementary_recheck_plan.md",
    "code": "code_error_supplementary_recheck_plan.md",
}
SUPPLEMENTARY_SHARD_DIRS = {
    "claims": "_recheck_supplementary",
    "code": "_code_error_recheck_supplementary",
}
SUPPLEMENTARY_SUMMARY_FILES = {
    "claims": "claims_supplementary_recheck_summary.md",
    "code": "code_error_supplementary_recheck_summary.md",
}
SUPPLEMENTARY_EMPTY = "No supplementary recheck inventory."
DISCOVERY_RANGE_RE = re.compile(
    r"^Declared supplementary discovery range:\s*([CEO]-\d{4})[–—-]([CEO]-\d{4})\s*$",
    re.M,
)
DISCOVERY_COUNTS_RE = re.compile(
    r"^Discoveries declared:\s*C=(\d+);\s*O=(\d+);\s*E=(\d+)\s*$", re.M)
OUTPUT_DISCOVERY_COLS = ["Output ID", "Claim ID", "Claim Verdict", "Output Status"]
OUTPUT_FROM_CLAIM_VERDICT = {
    "substantiated": "inconsistent",
    "substantiated_but_reframe": "inconsistent",
    "row_note_only": "mapped",
    "not_substantiated": "mapped",
    "confirmation_needed": "unclear",
    "blocked": "unclear",
}
LO_COLS = ["LO ID", "Source Shard", "Anchor", "Observation"]
LO_DISPOSITION_COLS = ["LO ID", "Prior State", "State"]
LO_EMPTY = "No late observations."
LO_DISPOSITIONS_EMPTY = "No dispositions."
BC_PLAN_COLS = [
    "BC ID", "LO ID", "Register", "Operation", "Row ID", "Payload JSON",
    "Old Value SHA256",
]
BC_RANGE_RE = re.compile(r"^Declared bC range:\s*([CEO]-\d{4})[–—-]([CEO]-\d{4})\s*$", re.M)
def has_conflict_markers(text):
    starts = re.search(r"^<{7}(\s|$)", text, re.M)
    ends = re.search(r"^>{7}(\s|$)", text, re.M)
    return bool(starts or ends)

SNAP_KEY = {
    "b3-claims": "claims_b3", "b6a-claims": "claims_b6a",
    "b6b-claims": "claims_b6b", "b3-code": "code_b3",
    "b6a-code": "code_b6a", "b6b-code": "code_b6b", "b7": "b7", "b8": "b8",
    "b3b-claims": "claims_b3b", "b3b-code": "code_b3b",
}


class Lint:
    def __init__(self):
        self.errors, self.warnings = [], []

    def fail(self, msg):
        self.errors.append(msg)

    def warn(self, msg):
        self.warnings.append(msg)

    def finish(self, stage):
        for w in self.warnings:
            print(f"WARNING [{stage}]: {w}")
        if self.errors:
            print(f"LINT FAIL [{stage}] — {len(self.errors)} finding(s):")
            for e in self.errors:
                print(f"  - {e}")
            return 1
        print(f"LINT PASS [{stage}]")
        return 0


# --------------------------------------------------------------- md parsing


def read_text(lint, path):
    if not path.is_file() or path.stat().st_size == 0:
        lint.fail(f"missing or empty file: {path}")
        return None
    raw = path.read_bytes()
    try:
        text = raw.decode("utf-8")
    except UnicodeDecodeError:
        lint.fail(f"not valid UTF-8: {path}")
        return None
    if has_conflict_markers(text):
        lint.fail(f"conflict marker in {path}")
    return text


def split_row(line):
    line = line.strip()
    if line.startswith("|"):
        line = line[1:]
    if line.endswith("|"):
        line = line[:-1]
    cells, cur, esc = [], [], False
    for ch in line:
        if esc:
            cur.append(ch)
            esc = False
        elif ch == "\\":
            cur.append(ch)
            esc = True
        elif ch == "|":
            cells.append("".join(cur).strip())
            cur = []
        else:
            cur.append(ch)
    cells.append("".join(cur).strip())
    return cells


def parse_tables(text):
    """All Markdown tables in *text* -> list of (headers, rows, start_line)."""
    lines = text.split("\n")
    tables, i = [], 0
    while i < len(lines) - 1:
        if lines[i].lstrip().startswith("|") and re.match(r"^\s*\|[\s:|-]+\|\s*$", lines[i + 1]):
            headers = split_row(lines[i])
            rows, j = [], i + 2
            while j < len(lines) and lines[j].lstrip().startswith("|"):
                rows.append(split_row(lines[j]))
                j += 1
            tables.append((headers, rows, i + 1))
            i = j
        else:
            i += 1
    return tables


def table_by_cols(lint, path, text, expected_cols, label):
    for headers, rows, line in parse_tables(text):
        if headers == expected_cols:
            bad = [k for k, r in enumerate(rows) if len(r) != len(headers)]
            for k in bad:
                lint.fail(f"{path}: {label} row {k + 1} has {len(rows[k])} cells, expected {len(headers)}")
            return [r for k, r in enumerate(rows) if k not in bad]
    lint.fail(f"{path}: no table with exact {label} columns found")
    return None


def tables_by_cols(lint, path, text, expected_cols, label, required=False):
    """Return all well-shaped rows from every table with one exact header."""
    found, matched = [], 0
    for headers, rows, _line in parse_tables(text):
        if headers != expected_cols:
            continue
        matched += 1
        for index, row in enumerate(rows, start=1):
            if len(row) != len(headers):
                lint.fail(f"{path}: {label} row {index} has {len(row)} cells, "
                          f"expected {len(headers)}")
            else:
                found.append(dict(zip(headers, row)))
    if required and matched == 0:
        lint.fail(f"{path}: no table with exact {label} columns found")
    return found


def blank_cell(value):
    return (value or "").strip().strip("`") in {"", "-", "—"}


def list_cell(value):
    if blank_cell(value):
        return []
    return [part.strip().strip("`") for part in str(value).split(";")
            if not blank_cell(part)]


def parse_field_patches(lint, path, row):
    value = row.get("Proposed Field Patches", "")
    if blank_cell(value):
        return {}
    patches = {}
    for raw in value.split(" || "):
        if " := " not in raw:
            lint.fail(f"{path}: {row['ID']} malformed Proposed Field Patches entry {raw!r}")
            continue
        column, replacement = raw.split(" := ", 1)
        if column not in PATCHABLE_ERROR_FIELDS:
            lint.fail(f"{path}: {row['ID']} field patch names non-whitelisted column {column!r}")
        elif column in patches:
            lint.fail(f"{path}: {row['ID']} field patch repeats column {column!r}")
        elif not replacement:
            lint.fail(f"{path}: {row['ID']} field patch for {column!r} is empty")
        else:
            patches[column] = replacement
    return patches


def expected_code_disposition(ledger):
    """Final (status, severity) a mapped code ledger row's verdict requires.

    Shared by the b6 merge lint and ``assemble_boundary.py --check`` so both
    gates apply the same verdict -> applied-disposition contract. Returns
    None for an invalid verdict.
    """
    verdict = (ledger.get("Verdict") or "").strip()
    proposed = (ledger.get("Proposed Severity") or "").strip()
    if verdict == "confirmed_error":
        return "confirmed", proposed
    if verdict == "not_error":
        return "not_error", ""
    if verdict == "confirmation_needed":
        capped = (str(min(int(proposed), 2))
                  if proposed in {"1", "2", "3", "4"} else proposed)
        return "confirmation_needed", capped
    if verdict in {"blocked", "deferred"}:
        return "blocked", (ledger.get("Current Severity") or "").strip()
    if verdict == "duplicate":
        target = (ledger.get("Duplicate Target") or "").strip()
        return f"duplicate_of:{target}", ""
    return None


def col(rows, cols, name):
    i = cols.index(name)
    return [r[i] for r in rows]


def ids_in(text, letter):
    return re.findall(rf"\b{ID_RE[letter]}\b", text)


def parse_range(s):
    m = RANGE_RE.search(s or "")
    if not m or m.group(1)[0] != m.group(2)[0]:
        return None
    return m.group(1)[0], int(m.group(1)[2:]), int(m.group(2)[2:])


def in_ranges(idstr, ranges):
    letter, num = idstr[0], int(idstr[2:])
    return any(l == letter and a <= num <= b for l, a, b in ranges)


def is_example_row(row):
    return any(re.fullmatch(r"[CEO]-0000", c) for c in row[:1])


def check_abs_paths(lint, path, rows, cols, path_cols):
    for name in path_cols:
        if name not in cols:
            continue
        for v in col(rows, cols, name):
            for token in re.split(r"[;,\s]+", v):
                if token.startswith("/") or token.startswith("~") or re.match(r"^[A-Za-z]:\\", token):
                    lint.fail(f"{path}: absolute path '{token}' in column '{name}' (repo-relative required)")


# --------------------------------------------------------------- vocab checks


def check_claims_rows(lint, path, rows, final=False):
    statuses = CLAIMS_STATUS_FINAL if final else CLAIMS_STATUS_FIRST
    for r in rows:
        d = dict(zip(CLAIMS_COLS, r))
        cid = d["Claim ID"]
        if not re.fullmatch(ID_RE["C"], cid):
            lint.fail(f"{path}: bad Claim ID '{cid}'")
        st = d["Status"]
        ok = st in statuses or (final and re.fullmatch(r"duplicate_of:C-\d{4}", st))
        if not ok:
            lint.fail(f"{path}: {cid} invalid status '{st}'")
        if d["Claim Type"] not in CLAIM_TYPES:
            lint.fail(f"{path}: {cid} invalid claim type '{d['Claim Type']}'")
        if not d["Paper Quote"]:
            lint.fail(f"{path}: {cid} empty Paper Quote")
        if d["Used in Text"] not in {"TRUE", "FALSE"}:
            lint.fail(f"{path}: {cid} Used in Text must be TRUE/FALSE")
        sev, issue = d["Severity"], d["Issue Description"]
        if bool(sev) != bool(issue):
            lint.fail(f"{path}: {cid} Severity/Issue Description biconditional violated")
        if (st == "blocked") != bool(d["Blocked Check"]):
            lint.fail(f"{path}: {cid} Blocked Check must be non-empty iff Status is 'blocked' (status='{st}')")
        if sev and sev not in {"1", "2", "3", "4"}:
            lint.fail(f"{path}: {cid} Severity '{sev}' not in 1-4")
        if d["Used in Text"] == "FALSE" and sev and sev != "1":
            lint.fail(f"{path}: {cid} Used in Text=FALSE cannot carry Severity > 1")


# --- U1 advisory adjudication heuristic (KTD-1: advisory, never a hard fail) --
#
# A closed claims row (`confirmed`/`blocked`) whose OWN recorded evidence — its
# Issue Description or Blocked Check — describes a paper-vs-code discrepancy is
# very likely mis-adjudicated: the escalation rule in registers.md would push it
# to `inconsistent` (or `confirmation_needed`). This lexical proxy surfaces such
# rows for human review. It is deliberately advisory (a WARNING, exit code
# unchanged): the token set both over- and under-matches, so a hard fail would
# false-block legitimate closures whose prose merely contains a contradiction
# word. The reasoning gate lives in the recheck-worker checklist; this is the
# safety net.
#
# Token set — a small, inline, documented list of contradiction phrases:
ADJUDICATION_TOKENS = (
    "whereas", "but the code", "instead", "does not match",
    "should be", "contradicts",
)
# Negation guard — phrases that, when they immediately precede a token, mark it
# as a NEUTRAL/NEGATED mention that must NOT fire (e.g. "nothing visible confirms
# or contradicts the claim"). Matched case-insensitively against the text right
# before the token's start offset.
ADJUDICATION_NEGATORS = (
    "confirms or ", "confirm or ", "confirmed or ", "nor ", "neither ",
    "no ", "not ", "nothing ",
)


def adjudication_flag(text):
    """Return the first contradiction token *actively* present in *text*, or None.

    A token is skipped when it is immediately preceded (ignoring surrounding
    whitespace) by any negator phrase — so a neutral clause such as "nothing
    visible confirms or contradicts the claim" stays silent.
    """
    if not text:
        return None
    low = text.lower()
    for tok in ADJUDICATION_TOKENS:
        start = 0
        while True:
            i = low.find(tok, start)
            if i == -1:
                break
            before = low[:i].rstrip()
            if not any(before.endswith(neg.rstrip()) for neg in ADJUDICATION_NEGATORS):
                return tok
            start = i + len(tok)
    return None


def check_adjudication_advisory(lint, path, rows, cols=None):
    """Advisory-only scan of FINAL claims registers (b6-claims, b8): flag every
    `confirmed`/`blocked` row whose own Issue Description or Blocked Check carries
    contradiction language. One WARNING per flagged row; never a hard fail.

    *cols* is the actual header order of *rows* (defaults to the canonical
    ``CLAIMS_COLS``); at b8 the staging register carries extra ``*Original``
    columns, so the caller passes those headers to keep field lookup aligned."""
    cols = cols or CLAIMS_COLS
    for r in rows:
        d = dict(zip(cols, r))
        st = d.get("Status", "")
        if st not in {"confirmed", "blocked"}:
            continue
        for field in ("Issue Description", "Blocked Check"):
            tok = adjudication_flag(d.get(field, ""))
            if tok:
                lint.warn(
                    f"adjudication: {d['Claim ID']} ({st}) — {field} contains "
                    f"contradiction language ('{tok}'); review against the "
                    f"escalation rule (registers.md)"
                )
                break


# --- U4 advisory identifier-anchoring heuristic (KTD-3: advisory, ledger-only)
#
# A recheck ledger row that closes a claim `confirmed` while the claim names a
# specific identifier (a variable, file, or parameter) that its own
# `Evidence Checked` never mentions is a likely anchoring gap: the anchoring
# rule in registers.md requires each named identifier located in the code at
# the role the claim assigns it — verifying that the operation exists and
# covers SOME variables anchors the operation, not the claim. Per KTD-3 the
# check reads the LEDGER's `Evidence Checked` column, never the claims
# register's own row (a confirmed register row has no evidence column, so
# comparing against it would fire on nearly every clean row); the claims
# register is read only to fetch each ledger ID's claim text. Like the U1
# adjudication advisory, this is a lexical proxy that both over- and
# under-matches, so it is a WARNING only — exit status never changes. Pinned
# false-positive path (see test_lint_registers.py): evidence citing a
# file:line anchor without repeating the identifier's name still warns —
# advisory noise is acceptable, a silent miss is not. Coverage limit, stated
# honestly: only rechecked rows have a ledger, so this is a tripwire over the
# recheck sample, not a guarantee over every confirmed claim.
#
# Identifier extraction — narrowly code-shaped tokens in the claim text:
#   * backtick-quoted tokens without internal whitespace (`test_score_std`);
#   * bare snake_case tokens (an interior underscore);
#   * bare dotted filenames with a code/data extension (build_panel.R).
_ANCHOR_EXTENSIONS = (
    r"do|ado|py|r|jl|m|sas|sps|csv|dta|tsv|txt|tex|md|json|ya?ml|xlsx?|rds|parquet"
)
_ANCHOR_IDENTIFIER_RE = re.compile(
    r"`([^`\s]+)`"                                    # backticked, no spaces
    r"|\b([A-Za-z][A-Za-z0-9]*(?:_[A-Za-z0-9]+)+)\b"  # snake_case
    r"|\b([\w./-]+\.(?:%s))\b" % _ANCHOR_EXTENSIONS,  # filename.ext
    re.IGNORECASE,
)
_ANCHOR_CONFIRMED_RE = re.compile(r"\bconfirmed\b", re.IGNORECASE)


def claim_named_identifiers(text):
    """Return the set of code-like identifiers *text* names (see regex above)."""
    found = set()
    for m in _ANCHOR_IDENTIFIER_RE.finditer(text or ""):
        found.add(next(g for g in m.groups() if g))
    return found


def check_anchoring_advisory(lint, audit, ledger_rows):
    """Advisory-only scan of a b5-claims recheck ledger: flag every row whose
    `Proposed Register Change` closes the claim `confirmed` while the claim's
    text names an identifier absent from the row's `Evidence Checked`. One
    WARNING per flagged row; never a hard fail. Skips silently when the
    canonical claims register is absent or unparsable (advisory robustness)."""
    reg_path = audit / "claims_register.md"
    if not reg_path.is_file():
        return
    try:
        reg_text = reg_path.read_text(encoding="utf-8")
    except (OSError, UnicodeDecodeError):
        return
    claim_text_by_id = {}
    for headers, rows, _ in parse_tables(reg_text):
        # Tolerate extra trailing columns: the post-b8 finalize step promotes
        # the rewriter's staging register, appending `*Original` columns, so
        # the real final `claims_register.md` header is CLAIMS_COLS PLUS extras.
        # Mirror load_register(..., allow_extra=True) / score_fixture's
        # _find_claims_table, and zip each row against the ACTUAL headers so the
        # claim columns are read by name, not position.
        if set(CLAIMS_COLS) <= set(headers):
            for r in rows:
                if len(r) != len(headers):
                    continue
                d = dict(zip(headers, r))
                claim_text_by_id[d["Claim ID"]] = d["Claim Text"]
            break
    if not claim_text_by_id:
        return
    for r in ledger_rows:
        d = dict(zip(LEDGER_COLS, r))
        rid = d["ID"]
        # only a close to `confirmed` is in scope (the word-boundary regex
        # deliberately does not match `confirmation_needed`).
        if not _ANCHOR_CONFIRMED_RE.search(d["Proposed Register Change"] or ""):
            continue
        evidence = (d["Evidence Checked"] or "").lower()
        missing = sorted(
            tok for tok in claim_named_identifiers(claim_text_by_id.get(rid, ""))
            if tok.lower() not in evidence
        )
        if missing:
            lint.warn(
                f"anchoring: {rid} closes confirmed but the claim names "
                f"identifier(s) absent from 'Evidence Checked': "
                f"{', '.join('`%s`' % t for t in missing)}; verify each is "
                f"anchored at its claimed role (anchoring rule, registers.md)"
            )


# --- U5 advisory filename-parameter reconciliation (KTD-4: advisory, blocked
#     rows only, same-syntactic-shape tokens only) --------------------------
#
# A `blocked` claims row whose claim text states a numeric parameter while a
# filename the row itself cites encodes a DIFFERENT value of the same
# syntactic shape has already transcribed a visible contradiction — the
# blocked-row escalation rule in registers.md says such a row cannot rest at
# `blocked`. This lint is the finalize-stage tripwire behind the reading-side
# reconciliation sweep (section-worker.md); it attaches at b8 (the U1
# adjudication advisory's finalize sibling; U4's anchoring advisory sits at
# b5). Per KTD-4 the crude any-numeric-mismatch version is explicitly
# rejected as too noisy — filenames carry incidental years, versions, and
# resolutions, and claim text is dense with estimates and sample sizes — so
# the check is restricted to blocked rows and compares only tokens of the
# SAME syntactic shape:
#   * ratio composites — "one-in-ten" / "1 in 10" in prose vs `1in10` /
#     `1_in_10` / `1-in-10` in a filename stem; spelled-out number words in
#     the claim are normalized to digits first (vocabulary mirrors the P-14
#     signatures in scripts/score_fixture.py);
#   * keyed parameter composites — an alpha key attached to a number
#     (`rp100`, `10km`, `q99`) compared only against a filename token
#     sharing the SAME alpha key. A key present on only one side is never a
#     mismatch, so an incidental year (no key at all) or version (`v3` with
#     no `v`-keyed claim token) stays silent.
# Advisory only: one WARNING per flagged row carrying the literal token
# ``filename-parameter``; exit status never changes.

# Spelled-out number words normalized before comparison (mirrors the
# score_fixture.py P-14 vocabulary: "one-in-ten", "one in twenty", ...).
_FP_NUMBER_WORDS = {
    "one": 1, "two": 2, "three": 3, "four": 4, "five": 5, "six": 6,
    "seven": 7, "eight": 8, "nine": 9, "ten": 10, "eleven": 11,
    "twelve": 12, "thirteen": 13, "fourteen": 14, "fifteen": 15,
    "sixteen": 16, "seventeen": 17, "eighteen": 18, "nineteen": 19,
    "twenty": 20, "thirty": 30, "forty": 40, "fifty": 50, "sixty": 60,
    "seventy": 70, "eighty": 80, "ninety": 90, "hundred": 100,
    "thousand": 1000,
}
_FP_NUMBER_WORD_RE = re.compile(
    r"\b(%s)\b" % "|".join(_FP_NUMBER_WORDS), re.IGNORECASE)
# filenames cited in a register cell (same extension vocabulary as the U4
# anchoring advisory above).
_FP_FILENAME_RE = re.compile(
    r"\b([\w./-]+\.(?:%s))\b" % _ANCHOR_EXTENSIONS, re.IGNORECASE)
# ratio composite in prose (after number-word normalization): "1-in-10",
# "1 in 10"; in a filename stem: "1in10", "1_in_10", "1-in-10".
_FP_PROSE_RATIO_RE = re.compile(r"\b(\d+)[\s-]+in[\s-]+(\d+)\b", re.IGNORECASE)
_FP_FNAME_RATIO_RE = re.compile(r"(\d+)[-_]?in[-_]?(\d+)", re.IGNORECASE)
# keyed composite: alpha key attached (no whitespace) to a number, either
# order — `rp100`, `rp-100`, `10km`, `q99`. Attachment (or a single
# hyphen/underscore) is required so free-standing counts ("4,832
# households") never form a token. Lookarounds rather than \b: an
# underscore is a \w character, so \b would never fire inside a filename
# stem like `grid_25km` — here `-`/`_` count as token separators.
_FP_KEYED_RES = (
    re.compile(r"(?<![A-Za-z0-9])([a-z]{1,10})[-_]?(\d+)(?![A-Za-z0-9])",
               re.IGNORECASE),   # key-first
    re.compile(r"(?<![A-Za-z0-9])(\d+)[-_]?([a-z]{1,10})(?![A-Za-z0-9])",
               re.IGNORECASE),   # number-first
)


def _fp_normalize_number_words(text):
    return _FP_NUMBER_WORD_RE.sub(
        lambda m: str(_FP_NUMBER_WORDS[m.group(1).lower()]), text)


def _fp_tokens(text, ratio_re):
    """(ratios, keyed) parameter tokens of *text*: ratios as a set of
    (numerator, denominator) int pairs; keyed composites as {key: {values}}.
    Ratio matches are masked before keyed extraction so `1in20` never also
    yields an `in`-keyed composite."""
    ratios = set()

    def mask(m):
        ratios.add((int(m.group(1)), int(m.group(2))))
        return " " * len(m.group(0))

    masked = ratio_re.sub(mask, text)
    keyed = {}
    for rx in _FP_KEYED_RES:
        for m in rx.finditer(masked):
            a, b = m.group(1), m.group(2)
            key, num = (a, b) if a[0].isalpha() else (b, a)
            keyed.setdefault(key.lower(), set()).add(int(num))
    return ratios, keyed


def _fp_claim_params(claim_text):
    """Parameter tokens the CLAIM states: cited filenames are masked out
    first (their embedded tokens belong to the filename side), then
    spelled-out number words are normalized to digits."""
    prose = _FP_FILENAME_RE.sub(lambda m: " " * len(m.group(0)), claim_text or "")
    return _fp_tokens(_fp_normalize_number_words(prose), _FP_PROSE_RATIO_RE)


def _fp_filename_params(*cells):
    """Parameter tokens embedded in every filename cited across *cells*,
    extracted from each filename's stem (basename, extension stripped)."""
    ratios, keyed = set(), {}
    for cell in cells:
        for m in _FP_FILENAME_RE.finditer(cell or ""):
            stem = m.group(1).rsplit("/", 1)[-1].rsplit(".", 1)[0]
            r, k = _fp_tokens(stem, _FP_FNAME_RATIO_RE)
            ratios |= r
            for key, vals in k.items():
                keyed.setdefault(key, set()).update(vals)
    return ratios, keyed


def check_filename_parameter_advisory(lint, path, rows, cols=None):
    """Advisory-only scan of a FINAL claims register (b8): flag every
    `blocked` row whose claim-stated parameter disagrees with a same-shape
    token embedded in a filename the row cites (Code/Data Source, Blocked
    Check, or the claim text itself). One WARNING per flagged row; never a
    hard fail. *cols* is the actual header order of *rows* (defaults to the
    canonical ``CLAIMS_COLS``; at b8 the staging register carries extra
    ``*Original`` columns)."""
    cols = cols or CLAIMS_COLS
    for r in rows:
        d = dict(zip(cols, r))
        if d.get("Status") != "blocked":
            continue
        c_ratios, c_keyed = _fp_claim_params(d.get("Claim Text", ""))
        f_ratios, f_keyed = _fp_filename_params(
            d.get("Code/Data Source", ""), d.get("Blocked Check", ""),
            d.get("Claim Text", ""))
        mismatches = []
        if c_ratios and f_ratios and not (c_ratios & f_ratios):
            mismatches.append(
                "ratio %s vs filename %s" % (
                    "/".join(sorted("%d-in-%d" % p for p in c_ratios)),
                    "/".join(sorted("%d-in-%d" % p for p in f_ratios))))
        for key in sorted(set(c_keyed) & set(f_keyed)):
            if not (c_keyed[key] & f_keyed[key]):
                mismatches.append(
                    "'%s' %s vs filename %s" % (
                        key,
                        "/".join(str(v) for v in sorted(c_keyed[key])),
                        "/".join(str(v) for v in sorted(f_keyed[key]))))
        if mismatches:
            lint.warn(
                f"filename-parameter: {d.get('Claim ID', '?')} (blocked) — "
                f"claim-stated parameter disagrees with a same-shape token "
                f"in a cited filename ({'; '.join(mismatches)}); reconcile "
                f"per the blocked-row escalation rule (registers.md)"
            )


def check_output_rows(lint, path, rows, final=False):
    statuses = OUTPUT_STATUS_FINAL if final else OUTPUT_STATUS_FIRST
    for r in rows:
        d = dict(zip(OUTPUT_COLS, r))
        oid = d["Output ID"]
        if not re.fullmatch(ID_RE["O"], oid):
            lint.fail(f"{path}: bad Output ID '{oid}'")
        st = d["Status"]
        ok = st in statuses or (final and re.fullmatch(r"duplicate_of:O-\d{4}", st))
        if not ok:
            lint.fail(f"{path}: {oid} invalid status '{st}'")


def check_error_rows(lint, path, rows, final=False):
    statuses = ERROR_STATUS_FINAL if final else ERROR_STATUS_FIRST
    for r in rows:
        d = dict(zip(ERROR_COLS, r))
        eid = d["Error ID"]
        if not re.fullmatch(ID_RE["E"], eid):
            lint.fail(f"{path}: bad Error ID '{eid}'")
        st = d["Status"]
        is_dup = bool(re.fullmatch(r"duplicate_of:E-\d{4}", st))
        if not (st in statuses or (final and is_dup)):
            lint.fail(f"{path}: {eid} invalid status '{st}'")
        if d["Error Type"] not in ERROR_TYPES:
            lint.fail(f"{path}: {eid} invalid error type '{d['Error Type']}'")
        sev = d["Severity"]
        needs_sev = st in {"candidate", "confirmed", "confirmation_needed", "blocked"}
        if needs_sev != bool(sev):
            lint.fail(f"{path}: {eid} Severity must be filled iff status is candidate/confirmed/confirmation_needed/blocked (status='{st}')")
        if sev and sev not in {"1", "2", "3", "4"}:
            lint.fail(f"{path}: {eid} Severity '{sev}' not in 1-4")
        # U8 (c): completeness of the author-facing / anchoring columns on an
        # ACTIVE code-error row (candidate or confirmed). An empty description or
        # source is a dead row that reaches the author with no content; an empty
        # Code Location loses the anchor. On `blocked`, the anchor may be genuinely
        # absent (restricted material), so an empty Code Location only WARNS there.
        if st in {"candidate", "confirmed"}:
            for c in ("Code/Data Source", "Error Description", "Why It Matters", "Code Location"):
                if not d.get(c, "").strip():
                    lint.fail(f"{path}: {eid} ({st}) has an empty '{c}'")
        elif st == "blocked" and not d.get("Code Location", "").strip():
            lint.warn(f"{path}: {eid} (blocked) has an empty 'Code Location' (anchor missing)")


def check_unique(lint, path, ids, label):
    seen = set()
    for i in ids:
        if i in seen:
            lint.fail(f"{path}: duplicate {label} '{i}'")
        seen.add(i)


def check_bidirectional(lint, a_rows, a_cols, a_id, a_link, b_rows, b_cols, b_id, b_link, label):
    def links(rows, cols, idc, linkc):
        out = {}
        for r in rows:
            d = dict(zip(cols, r))
            out[d[idc]] = set(re.findall(r"[CEO]-\d{4}", d[linkc]))
        return out

    fwd, back = links(a_rows, a_cols, a_id, a_link), links(b_rows, b_cols, b_id, b_link)
    for aid, targets in fwd.items():
        for t in targets:
            if t not in back:
                lint.fail(f"{label}: {aid} links {t}, which does not exist")
            elif aid not in back[t]:
                lint.fail(f"{label}: one-way link {aid} -> {t} (reverse missing)")
    for bid, targets in back.items():
        for t in targets:
            if t not in fwd:
                lint.fail(f"{label}: {bid} links {t}, which does not exist")
            elif bid not in fwd[t]:
                lint.fail(f"{label}: one-way link {bid} -> {t} (reverse missing)")


# --------------------------------------------------------------- plan parsing


def parse_plan(lint, path, key_col):
    """Parse an allocation table keyed by *key_col*; return list of row dicts + coord ranges."""
    text = read_text(lint, path)
    if text is None:
        return None, []
    for headers, rows, _ in parse_tables(text):
        if key_col in headers and "Shard File" in headers:
            alloc = []
            for k, r in enumerate(rows):
                if len(r) != len(headers):
                    lint.fail(f"{path}: allocation row {k + 1} has {len(r)} cells, expected {len(headers)}")
                else:
                    alloc.append(dict(zip(headers, r)))
            coord = [parse_range(f"{a}–{b}") for a, b in COORD_RE.findall(text)]
            return alloc, [c for c in coord if c]
    lint.fail(f"{path}: allocation table with '{key_col}' and 'Shard File' not found")
    return None, []


def check_manifest_worker_shards(lint, manifest, stage_key, plan, key_col,
                                 allocations=None):
    """A worker stage with planned work must record manifest shard evidence.

    The boundary lint runs before the conductor changes ``running`` to
    ``done``, so stage status cannot guard this check.  ``allocations`` lets
    b3b reuse the plan it already parsed; b6 supplies no allocation so the
    recheck cluster table is read here.
    """
    if allocations is None:
        allocations, _ = parse_plan(lint, plan, key_col)
    if allocations is None:
        return
    worker_count = len(allocations)
    if not worker_count:
        return

    stages = manifest.get("stages", {}) if isinstance(manifest, dict) else {}
    entry = stages.get(stage_key) if isinstance(stages, dict) else None
    if not isinstance(entry, dict):
        lint.fail(
            f"manifest stage '{stage_key}' is missing, but {plan} lists "
            f"{worker_count} planned worker(s)"
        )
    elif not entry.get("shards"):
        lint.fail(
            f"manifest stage '{stage_key}' has no shards, but "
            f"{plan} lists {worker_count} planned worker(s)"
        )


def alloc_ranges(lint, path, alloc, range_cols):
    ranges = []
    for row in alloc:
        for rc in range_cols:
            pr = parse_range(row.get(rc, ""))
            if pr is None:
                lint.fail(f"{path}: unparseable {rc} '{row.get(rc, '')}' for {row.get('Worker ID') or row.get('Chunk ID')}")
            else:
                ranges.append(pr)
    return ranges


def check_disjoint(lint, path, ranges):
    by_letter = {}
    for l, a, b in ranges:
        by_letter.setdefault(l, []).append((a, b))
    for l, spans in by_letter.items():
        spans.sort()
        for (a1, b1), (a2, b2) in zip(spans, spans[1:]):
            if a2 <= b1:
                lint.fail(f"{path}: overlapping {l}- ID ranges {a1}-{b1} and {a2}-{b2}")


def scope_tokens(cell):
    return {t.strip().strip("`") for t in re.split(r"[;,]", cell or "") if t.strip()}


def normalized_audit_path(value):
    """Normalize a plan/shard path to its package-relative ``audit/...`` wire."""
    raw = str(value or "").strip().strip("`").replace("\\", "/")
    if not raw:
        return ""
    marker = "/audit/"
    if marker in raw:
        raw = "audit/" + raw.split(marker, 1)[1]
    elif raw.startswith("audit/"):
        pass
    elif raw.startswith("/"):
        return raw
    else:
        raw = "audit/" + raw.lstrip("./")
    return raw


def audit_path(audit, value):
    normalized = normalized_audit_path(value)
    if normalized.startswith("audit/"):
        return audit.parent / normalized
    return Path(value)


def check_identifier_exhaustion(lint, path, ranges):
    """Refuse a declaration that consumes the last four-digit register ID."""
    labels = {"C": "claims", "O": "output", "E": "code-error"}
    for letter, _start, end in ranges:
        if end >= 9999:
            lint.fail(
                f"{path}: {labels[letter]} register identifier space exhausted "
                f"at {letter}-9999 (four-digit IDs cannot wrap)"
            )


def second_read_allocations(lint, path, stream):
    """Parse and validate the documented U6a second-read allocation contract."""
    text = read_text(lint, path)
    if text is None:
        return None, []
    expected = SECOND_READ_PLAN_COLS[stream]
    matches = [(rows, line) for headers, rows, line in parse_tables(text)
               if headers == expected]
    if len(matches) != 1:
        lint.fail(
            f"{path}: expected exactly one second-read allocation table with columns "
            + " | ".join(expected)
        )
        return None, []
    rows, _line = matches[0]
    allocations = []
    for index, row in enumerate(rows, start=1):
        if len(row) != len(expected):
            lint.fail(
                f"{path}: allocation row {index} has {len(row)} cells, "
                f"expected {len(expected)}"
            )
            continue
        entry = dict(zip(expected, row))
        if entry["Reason"] not in SECOND_READ_REASONS:
            lint.fail(
                f"{path}: {entry['Worker ID']} has invalid Reason "
                f"{entry['Reason']!r}"
            )
        allocations.append(entry)
    coord = [parse_range(f"{a}–{b}") for a, b in COORD_RE.findall(text)]
    if stream == "claims":
        audit = Path(path).parent.parent
        ledger_path = audit / "_run" / "snapshots" / "claims_b3" / "handoff_ledger.json"
        if ledger_path.is_file():
            try:
                ledger = json.loads(ledger_path.read_text(encoding="utf-8"))
                forwarded = {entry["id"] for entry in ledger.get("H", [])
                             if entry.get("state") == "forwarded"}
            except (OSError, json.JSONDecodeError) as exc:
                lint.fail(f"{ledger_path}: invalid handoff ledger ({exc})")
                forwarded = set()
            assigned = []
            for entry in allocations:
                raw = entry["Assigned Handoff IDs"]
                ids = [] if blank_cell(raw) else [item.strip() for item in raw.split(",")]
                if any(not re.fullmatch(r"H-\d{4}", item) for item in ids):
                    lint.fail(f"{path}: {entry['Worker ID']} has malformed Assigned Handoff IDs")
                for handoff_id in ids:
                    if handoff_id not in forwarded:
                        lint.fail(f"{path}: {entry['Worker ID']} names non-forwarded {handoff_id}")
                assigned.extend(ids)
            duplicates = sorted({item for item in assigned if assigned.count(item) > 1})
            if duplicates:
                lint.fail(f"{path}: forwarded H IDs assigned more than once: {duplicates}")
            if set(assigned) != forwarded:
                lint.fail(f"{path}: Assigned Handoff IDs do not exactly cover forwarded ledger IDs; "
                          f"missing={sorted(forwarded - set(assigned))}, "
                          f"extra={sorted(set(assigned) - forwarded)}")
        else:
            for entry in allocations:
                if not blank_cell(entry["Assigned Handoff IDs"]):
                    lint.fail(f"{path}: Assigned Handoff IDs must remain empty until U7 "
                              "ledger activation (claims_b3 ledger snapshot missing)")
    else:
        for entry in allocations:
            if not blank_cell(entry["Assigned Handoff IDs"]):
                lint.fail(
                    f"{path}: code-stream Assigned Handoff IDs must remain empty"
                )
    return allocations, [item for item in coord if item]


def coverage_outcome(lint, path, row_number, value):
    """Return (kind, IDs/reason) for one exact coverage outcome."""
    outcome = str(value or "").strip()
    if outcome == COVERAGE_CLEAN:
        return "clean", []
    match = COVERAGE_FINDINGS_RE.fullmatch(outcome)
    if match:
        return "findings", re.findall(r"[CEO]-\d{4}", match.group(1))
    match = COVERAGE_BLOCKED_RE.fullmatch(outcome)
    if match:
        return "blocked", match.group(1).strip()
    lint.fail(
        f"{path}: coverage row {row_number} outcome {outcome!r} must be exactly "
        "'clean', 'findings: <IDs>', or 'blocked: <reason>'"
    )
    return "invalid", []


def typed_shard_footer(lint, path, text, stream, recheck=False):
    """Validate the shared typed footer and return (entries, coverage rows).

    Recheck workers cannot mint register rows.  In that context ``candidate``
    therefore means a defect-character observation for the consuming merge and
    its Register IDs cell is deliberately empty.
    """
    table_matches = [(rows, line) for headers, rows, line in parse_tables(text)
                     if headers == FOOTER_COLS]
    if len(table_matches) != 1:
        lint.fail(
            f"{path}: expected exactly one typed footer table with columns "
            + " | ".join(FOOTER_COLS)
        )
        entries = []
    else:
        entries = []
        rows, _line = table_matches[0]
        for index, row in enumerate(rows, start=1):
            if len(row) != len(FOOTER_COLS):
                lint.fail(
                    f"{path}: typed footer row {index} has {len(row)} cells, "
                    f"expected {len(FOOTER_COLS)}"
                )
                continue
            entry = dict(zip(FOOTER_COLS, row))
            expected_id = f"OBS-{index:04d}"
            if entry["Entry ID"] != expected_id:
                lint.fail(
                    f"{path}: typed footer entry {index} must be {expected_id}, "
                    f"got {entry['Entry ID']!r}"
                )
            if entry["Kind"] not in FOOTER_KINDS:
                lint.fail(
                    f"{path}: {entry['Entry ID']} has invalid Kind "
                    f"{entry['Kind']!r}"
                )
            if blank_cell(entry["Observation"]) or "\n" in entry["Observation"]:
                lint.fail(f"{path}: {entry['Entry ID']} requires a one-line Observation")
            ids = re.findall(r"[CEO]-\d{4}", entry["Register IDs"])
            if entry["Kind"] == "candidate":
                if recheck and (ids or not blank_cell(entry["Register IDs"])):
                    lint.fail(
                        f"{path}: {entry['Entry ID']} recheck candidate cannot mint Register IDs"
                    )
                elif not recheck and not ids:
                    lint.fail(f"{path}: {entry['Entry ID']} candidate requires Register IDs")
                if not blank_cell(entry["Reason"]):
                    lint.fail(f"{path}: {entry['Entry ID']} candidate Reason must be empty")
            elif entry["Kind"] == "not_rowed_observation":
                if ids or not blank_cell(entry["Register IDs"]):
                    lint.fail(
                        f"{path}: {entry['Entry ID']} not_rowed_observation cannot name Register IDs"
                    )
                if blank_cell(entry["Reason"]) or "\n" in entry["Reason"]:
                    lint.fail(
                        f"{path}: {entry['Entry ID']} not_rowed_observation requires a one-line Reason"
                    )
            entry["_ids"] = ids
            entries.append(entry)

    coverage = []
    if stream == "code" and not recheck:
        matches = [(rows, line) for headers, rows, line in parse_tables(text)
                   if headers == COVERAGE_COLS]
        if len(matches) != 1:
            lint.fail(
                f"{path}: expected exactly one code coverage table with columns "
                + " | ".join(COVERAGE_COLS)
            )
        else:
            rows, _line = matches[0]
            seen = set()
            for index, row in enumerate(rows, start=1):
                if len(row) != len(COVERAGE_COLS):
                    lint.fail(
                        f"{path}: coverage row {index} has {len(row)} cells, "
                        f"expected {len(COVERAGE_COLS)}"
                    )
                    continue
                data = dict(zip(COVERAGE_COLS, row))
                script = data["Script"].strip().strip("`")
                if (not script or script.startswith("/") or ".." in Path(script).parts
                        or (script.startswith("@") and script != HYGIENE_SINGLETON)):
                    lint.fail(f"{path}: invalid coverage key {script!r}")
                if script in seen:
                    lint.fail(f"{path}: duplicate coverage key {script!r}")
                seen.add(script)
                kind, detail = coverage_outcome(lint, path, index, data["Outcome"])
                coverage.append({"script": script, "outcome": data["Outcome"],
                                 "kind": kind, "detail": detail})
    return entries, coverage


def shard_register_ids(text, stream):
    columns = ((CLAIMS_COLS, "Claim ID"), (OUTPUT_COLS, "Output ID")) \
        if stream == "claims" else ((ERROR_COLS, "Error ID"),)
    found = set()
    for headers, rows, _line in parse_tables(text):
        for expected, id_col in columns:
            if headers == expected:
                index = headers.index(id_col)
                found.update(row[index] for row in rows
                             if len(row) == len(headers) and not is_example_row(row))
    return found


def validate_footer_candidates(lint, path, text, stream, entries, coverage):
    row_ids = shard_register_ids(text, stream)
    footer_ids = set()
    for entry in entries:
        if entry["Kind"] != "candidate":
            continue
        footer_ids.update(entry["_ids"])
        for register_id in entry["_ids"]:
            if register_id not in row_ids:
                lint.fail(
                    f"{path}: {entry['Entry ID']} cites {register_id}, which is not a row in the shard"
                )
    if footer_ids != row_ids:
        missing = sorted(row_ids - footer_ids)
        extra = sorted(footer_ids - row_ids)
        lint.fail(
            f"{path}: typed-footer candidate IDs do not equal shard row IDs; "
            f"missing={missing}, extra={extra}"
        )
    if stream != "code":
        return
    cited = set()
    clean_scripts = {row["script"] for row in coverage if row["kind"] == "clean"}
    for row in coverage:
        if row["kind"] == "findings":
            cited.update(row["detail"])
            for register_id in row["detail"]:
                if register_id not in row_ids:
                    lint.fail(
                        f"{path}: coverage cites {register_id}, which is not a row in the shard"
                    )
    if cited != row_ids:
        lint.fail(
            f"{path}: findings coverage IDs do not equal code-error row IDs; "
            f"missing={sorted(row_ids - cited)}, extra={sorted(cited - row_ids)}"
        )
    for headers, rows, _line in parse_tables(text):
        if headers != ERROR_COLS:
            continue
        for row in rows:
            if len(row) != len(headers) or is_example_row(row):
                continue
            data = dict(zip(headers, row))
            sources = scope_tokens(data["Code/Data Source"])
            for source in sorted(sources & clean_scripts):
                lint.fail(
                    f"{path}: {source!r} carries finding {data['Error ID']} and cannot be clean"
                )


def parse_footer_dispositions(lint, report_path, report):
    """Parse the documented footer-disposition line grammar."""
    raw = report.get("footer_dispositions", [])
    if not isinstance(raw, list):
        lint.fail(f"{report_path}: footer_dispositions must be a list")
        return []
    parsed = []
    pattern = re.compile(
        r"^(audit/[^#]+)#(OBS-\d{4})\s+\|\s+"
        r"(candidate:((?:[CEO]-\d{4})(?:\s*[;,]\s*[CEO]-\d{4})*)|"
        r"late_observation:(LO-[CE]-\d{4})|dismissed:(\S.*))$"
    )
    for index, line in enumerate(raw, start=1):
        if not isinstance(line, str) or not (match := pattern.fullmatch(line.strip())):
            lint.fail(
                f"{report_path}: footer disposition {index} must serialize as "
                "'<audit shard>#<OBS-ID> | candidate:<IDs>', "
                "'... | late_observation:<LO-ID>', or '... | dismissed:<reason>'"
            )
            continue
        ids = re.findall(r"[CEO]-\d{4}", match.group(4) or "")
        action = "candidate" if match.group(4) else (
            "late_observation" if match.group(5) else "dismissed")
        parsed.append({"key": (match.group(1), match.group(2)),
                       "action": action, "ids": ids,
                       "lo_id": match.group(5) or ""})
    return parsed


def reconcile_footer_dispositions(lint, audit, allocations, stream,
                                  report_path, report, staging_ids, manifest,
                                  worker_stage):
    """Prove the exact-one typed-footer entry/disposition join."""
    stage = manifest.get("stages", {}).get(worker_stage, {}) \
        if isinstance(manifest, dict) else {}
    shard_states = stage.get("shards", {}) if isinstance(stage, dict) else {}
    expected = {}
    for allocation in allocations or []:
        wire = normalized_audit_path(allocation.get("Shard File"))
        state = shard_states.get(wire, shard_states.get(allocation.get("Shard File"), {}))
        if isinstance(state, dict) and state.get("status") == "blocked":
            continue
        path = audit_path(audit, wire)
        text = read_text(lint, path)
        if text is None:
            continue
        entries, coverage = typed_shard_footer(lint, path, text, stream)
        validate_footer_candidates(lint, path, text, stream, entries, coverage)
        for entry in entries:
            key = (wire, entry["Entry ID"])
            if key in expected:
                lint.fail(f"{path}: duplicate typed-footer join key {wire}#{entry['Entry ID']}")
            expected[key] = entry
    actual = {}
    for disposition in parse_footer_dispositions(lint, report_path, report):
        key = disposition["key"]
        if key in actual:
            lint.fail(f"{report_path}: duplicate footer disposition for {key[0]}#{key[1]}")
        actual[key] = disposition
    for key in sorted(set(expected) - set(actual)):
        lint.fail(f"{report_path}: typed footer entry {key[0]}#{key[1]} is undispositioned")
    for key in sorted(set(actual) - set(expected)):
        lint.fail(f"{report_path}: disposition has no typed footer entry {key[0]}#{key[1]}")
    for key in sorted(set(expected) & set(actual)):
        entry, disposition = expected[key], actual[key]
        if entry["Kind"] == "candidate" and disposition["action"] != "candidate":
            lint.fail(
                f"{report_path}: {key[0]}#{key[1]} candidate entry cannot be dismissed"
            )
            continue
        if disposition["action"] == "candidate":
            if set(disposition["ids"]) - staging_ids:
                lint.fail(
                    f"{report_path}: {key[0]}#{key[1]} disposition cites absent staging IDs "
                    f"{sorted(set(disposition['ids']) - staging_ids)}"
                )
            if entry["Kind"] == "candidate" and set(disposition["ids"]) != set(entry["_ids"]):
                lint.fail(
                    f"{report_path}: {key[0]}#{key[1]} candidate disposition IDs "
                    "do not match the shard entry"
                )


# --------------------------------------------------------------- stage checks


def stage_b0(lint, audit, manifest):
    for f, cols in [
        ("claims_register.md", CLAIMS_COLS),
        ("output_register.md", OUTPUT_COLS),
        ("code_error_register.md", ERROR_COLS),
    ]:
        text = read_text(lint, audit / f)
        if text is None:
            continue
        rows = table_by_cols(lint, audit / f, text, cols, f)
        if rows:
            real = [r for r in rows if not is_example_row(r)]
            if real:
                lint.fail(f"{audit / f}: register must be empty at b0 (found {len(real)} rows)")
    if read_text(lint, audit / "audit_readme.md") is None:
        pass
    cm = read_text(lint, audit / "CODEMAP.md")
    if cm is not None:
        for letter in "SDB":
            found = re.findall(rf"\b{letter}-\d{{4}}\b", cm)
            check_unique(lint, audit / "CODEMAP.md", found, f"{letter}- ID")
        if not re.search(r"PRECONDITIONS:\s*\d\s*/\s*5", cm):
            lint.fail(f"{audit / 'CODEMAP.md'}: missing 'PRECONDITIONS: <n>/5' score line")
    if manifest and manifest.get("paper_source_set") is not None:
        try:
            validate_source_set(audit.parent, manifest)
        except PaperSourceError as exc:
            lint.fail(f"manifest: {exc}")
    elif manifest:
        src, blanked = manifest.get("paper_source_path"), manifest.get("paper_audit_path")
        if src and Path(src).suffix == ".tex":
            if not blanked or blanked == src:
                lint.fail("manifest: .tex paper requires a distinct paper_audit_path (blanked copy)")
            elif not Path(blanked).is_file():
                lint.fail(f"blanked paper missing: {blanked}")
            elif not Path(src).is_file():
                lint.fail(f"paper source missing: {src}")
            else:
                n_src = Path(src).read_text(encoding="utf-8").count("\n")
                n_bl = Path(blanked).read_text(encoding="utf-8").count("\n")
                if n_src != n_bl:
                    lint.fail(f"blanked paper line count {n_bl} != source {n_src}")


def _stage_b1_claim_handoffs(lint, audit, manifest, plan):
    """U7 exact plan schema, paper partition, and H/adjudication ranges."""
    from claim_handoffs import (
        ADJUDICATION_RANGE_RE, CLAIMS_PLAN_COLS, load_claims_allocations,
        parse_h_range, validate_partition,
    )
    try:
        allocations, text = load_claims_allocations(plan)
    except (OSError, ValueError) as exc:
        lint.fail(str(exc))
        return None, []
    expected_override = manifest.get("allocation_override")
    if isinstance(expected_override, dict) and allocations != expected_override.get("allocation"):
        lint.fail(f"{plan}: executed allocation does not equal manifest allocation_override")
    check_unique(lint, plan, [row["Worker ID"] for row in allocations], "Worker ID")
    check_unique(lint, plan, [row["Shard File"] for row in allocations], "Shard File")
    ranges = alloc_ranges(lint, plan, allocations, ["Claim ID Range", "Output ID Range"])
    h_ranges = []
    for row in allocations:
        parsed = parse_h_range(row["H ID Range"])
        if parsed is None:
            lint.fail(f"{plan}: unparseable H ID Range {row['H ID Range']!r} for {row['Worker ID']}")
        else:
            h_ranges.append(parsed)
    check_disjoint(lint, plan, ranges)
    check_disjoint(lint, plan, h_ranges)
    check_identifier_exhaustion(lint, plan, ranges)
    for _letter, _start, end in h_ranges:
        if end >= 9999:
            lint.fail(f"{plan}: handoff identifier space exhausted at H-9999")
    coord = [parse_range(f"{a}–{b}") for a, b in COORD_RE.findall(text)]
    coord = [item for item in coord if item]
    matches = ADJUDICATION_RANGE_RE.findall(text)
    adjudication = [parse_range(f"{start}–{end}") for start, end in matches]
    adjudication = [item for item in adjudication if item]
    if len(adjudication) != 1 or adjudication[0][2] - adjudication[0][1] + 1 != 50:
        lint.fail(f"{plan}: expected exactly one 50-ID adjudication C-mint range")
    check_disjoint(lint, plan, ranges + coord + adjudication)
    if len([item for item in coord if item[0] == "C"]) != 1 \
            or len([item for item in coord if item[0] == "O"]) != 1:
        lint.fail(f"{plan}: expected exactly one merge-coordinator range each for C- and O-")
    try:
        validate_partition(allocations, manifest["paper_source_set"], audit.parent)
    except (KeyError, ValueError, OSError) as exc:
        lint.fail(f"{plan}: {exc}")
    return allocations, coord


def stage_b1(lint, audit, stream, manifest=None):
    manifest = manifest or {}
    if stream == "claims":
        plan = audit / "plans" / "claims_review_plan.md"
        if manifest.get("paper_source_set") is not None:
            _stage_b1_claim_handoffs(lint, audit, manifest, plan)
            return
        alloc, coord = parse_plan(lint, plan, "Worker ID")
        if alloc is None:
            return
        check_unique(lint, plan, [a["Worker ID"] for a in alloc], "Worker ID")
        check_unique(lint, plan, [a["Shard File"] for a in alloc], "Shard File")
        ranges = alloc_ranges(lint, plan, alloc, ["Claim ID Range", "Output ID Range"])
        check_disjoint(lint, plan, ranges + coord)
        check_identifier_exhaustion(lint, plan, ranges + coord)
        if len([c for c in coord if c[0] == "C"]) != 1 or len([c for c in coord if c[0] == "O"]) != 1:
            lint.fail(f"{plan}: expected exactly one merge-coordinator range each for C- and O-")
    else:
        plan = audit / "plans" / "code_error_review_plan.md"
        alloc, coord = parse_plan(lint, plan, "Chunk ID")
        if alloc is None:
            return
        check_unique(lint, plan, [a["Chunk ID"] for a in alloc], "Chunk ID")
        check_unique(lint, plan, [a["Shard File"] for a in alloc], "Shard File")
        ranges = alloc_ranges(lint, plan, alloc, ["Error ID Range"])
        check_disjoint(lint, plan, ranges + coord)
        check_identifier_exhaustion(lint, plan, ranges + coord)
        if len([c for c in coord if c[0] == "E"]) != 1:
            lint.fail(f"{plan}: expected exactly one merge-coordinator range for E-")
        # every inventory script in exactly one chunk (token match, not substring)
        text = plan.read_text(encoding="utf-8")
        inv = None
        for headers, rows, _ in parse_tables(text):
            if "Script" in headers and "Chunk" in headers:
                inv = []
                for k, r in enumerate(rows):
                    if len(r) != len(headers):
                        lint.fail(f"{plan}: inventory row {k + 1} has {len(r)} cells, expected {len(headers)}")
                    else:
                        inv.append(dict(zip(headers, r)))
                break
        if inv is None:
            lint.fail(f"{plan}: script inventory table (with 'Script' and 'Chunk' columns) not found")
        else:
            for row in inv:
                script = row["Script"].strip("`")
                n = sum(1 for a in alloc if script in scope_tokens(a["Script Scope"]))
                if n != 1:
                    lint.fail(f"{plan}: inventory script '{script}' appears in {n} chunks (expected exactly 1)")
        hygiene = None
        for headers, rows, _ in parse_tables(text):
            if headers == ["Hygiene File", "Chunk"]:
                hygiene = []
                for index, row in enumerate(rows, start=1):
                    if len(row) != len(headers):
                        lint.fail(
                            f"{plan}: hygiene inventory row {index} has {len(row)} cells, "
                            f"expected {len(headers)}"
                        )
                    else:
                        hygiene.append(dict(zip(headers, row)))
                break
        if not hygiene:
            lint.fail(
                f"{plan}: hygiene file inventory table with exact columns "
                "'Hygiene File | Chunk' is required and cannot be empty"
            )
        else:
            allocation_by_chunk = {a["Chunk ID"]: a for a in alloc}
            seen_hygiene = set()
            for row in hygiene:
                path = row["Hygiene File"].strip().strip("`")
                if path in seen_hygiene:
                    lint.fail(f"{plan}: duplicate hygiene file {path!r}")
                seen_hygiene.add(path)
                if row["Chunk"] not in allocation_by_chunk:
                    lint.fail(
                        f"{plan}: hygiene file {path!r} names unknown chunk {row['Chunk']!r}"
                    )


def shard_footer(lint, path, text):
    # U6a replaces free-form coordinator notes with a typed entry table.
    # Coverage remains stream-specific: code uses the exact table grammar;
    # claims keeps its exhaustive prose checklist.
    low = text.lower()
    if "coverage" not in low:
        lint.fail(f"{path}: missing coverage note in shard footer")


def find_alloc_for_shard(alloc, shard):
    for a in alloc or []:
        if a["Shard File"].strip("`") == str(shard) or str(shard).endswith(a["Shard File"].strip("`")):
            return a
    return None


def _exact_or_empty_table(lint, path, text, columns, empty_sentence, label):
    matches = [(rows, line) for headers, rows, line in parse_tables(text)
               if headers == columns]
    has_empty = bool(re.search(
        rf"(?m)^\s*{re.escape(empty_sentence)}\s*$", text
    ))
    if len(matches) > 1 or (matches and has_empty):
        lint.fail(f"{path}: {label} must use exactly one table or exact empty form")
        return []
    if not matches:
        if not has_empty:
            lint.fail(f"{path}: missing {label}; use exact empty form '{empty_sentence}'")
        return []
    rows, _line = matches[0]
    output = []
    for index, row in enumerate(rows, start=1):
        if len(row) != len(columns):
            lint.fail(f"{path}: {label} row {index} has {len(row)} cells, expected {len(columns)}")
        else:
            output.append(dict(zip(columns, row)))
    return output


def _claims_rows_by_id(claims_rows):
    """Canon membership map for coverage citations. Paper Context stays the
    prose locator registers.md mandates; the machine anchor travels on the
    coverage/resolution entry's Covering Range / Covering Quote cells."""
    by_id = {}
    for row in claims_rows:
        claim = dict(zip(CLAIMS_COLS, row)) if not isinstance(row, dict) else row
        by_id[claim["Claim ID"]] = claim
    return by_id


def _validate_coverage_row(lint, path, row, obligation_anchor, claims_by_id,
                           source_set, package_root):
    from claim_handoffs import validate_disposition
    outcome = row["Outcome"]
    if outcome == "covered":
        claim_id = row["C-ID / Reason"]
        claim = claims_by_id.get(claim_id)
        if claim is None:
            lint.fail(f"{path}: covered obligation cites absent claim {claim_id}")
            return
        if row["Covering Quote"] != claim["Paper Quote"]:
            lint.fail(f"{path}: covering quote for {claim_id} is not verbatim Paper Quote")
            return
        try:
            carried = resolve_quote(
                source_set, row["Covering Range"], row["Covering Quote"], package_root
            )
        except AnchorError as exc:
            lint.fail(f"{path}: {exc}")
            return
        if not anchor_contains(carried, obligation_anchor):
            lint.fail(f"{path}: covering row {claim_id} does not contain the obligation assertion")
        if not blank_cell(row["Evidence"]):
            lint.fail(f"{path}: covered obligation Evidence must be blank")
    elif outcome == "disposition":
        try:
            validate_disposition(row["C-ID / Reason"], row["Evidence"])
        except ValueError as exc:
            lint.fail(f"{path}: {exc}")
        if not (blank_cell(row["Covering Range"]) and blank_cell(row["Covering Quote"])):
            lint.fail(f"{path}: disposition covering range and quote must be blank")
    else:
        lint.fail(f"{path}: invalid terminal outcome {outcome!r}")


def _validate_u7_b2_shard(lint, audit, shard, text, allocation, claims_rows, manifest):
    from claim_handoffs import (
        HANDOFF_COLS, X_COVERAGE_COLS, parse_h_range,
    )
    handoffs = _exact_or_empty_table(
        lint, shard, text, HANDOFF_COLS, "No handoffs.", "Handoffs"
    )
    coverage = _exact_or_empty_table(
        lint, shard, text, X_COVERAGE_COLS, "No assigned cross-references.",
        "Cross-reference coverage",
    )
    source_set = manifest["paper_source_set"]
    h_range = parse_h_range(allocation.get("H ID Range"))
    seen = set()
    for row in handoffs:
        handoff_id = row["H ID"]
        if not re.fullmatch(r"H-\d{4}", handoff_id):
            lint.fail(f"{shard}: invalid handoff ID {handoff_id!r}")
        elif handoff_id in seen:
            lint.fail(f"{shard}: duplicate handoff ID {handoff_id}")
        elif h_range is None or not (h_range[1] <= int(handoff_id[2:]) <= h_range[2]):
            lint.fail(f"{shard}: {handoff_id} outside the shard's allocated H range")
        seen.add(handoff_id)
        if blank_cell(row["Asserted Substance"]) or "\n" in row["Asserted Substance"]:
            lint.fail(f"{shard}: {handoff_id} requires one-sentence Asserted Substance")
        try:
            resolve_quote(source_set, row["Anchor"], row["Quote"], audit.parent)
        except AnchorError as exc:
            lint.fail(f"{shard}: {handoff_id}: {exc}")
    inventory_path = audit / "_run" / "crossref_inventory.json"
    assignments_path = audit / "_run" / "crossref_assignments.json"
    try:
        inventory = json.loads(inventory_path.read_text(encoding="utf-8"))
        assignments = json.loads(assignments_path.read_text(encoding="utf-8"))["assignments"]
    except (OSError, json.JSONDecodeError, KeyError) as exc:
        lint.fail(f"{shard}: cannot load crossref artifacts: {exc}")
        return
    by_id = {entry["id"]: entry for entry in inventory.get("entries", [])}
    expected = {x_id for x_id, worker in assignments.items()
                if worker == allocation["Worker ID"]}
    actual = [row["X ID"] for row in coverage]
    if len(actual) != len(set(actual)):
        lint.fail(f"{shard}: duplicate X coverage ID")
    if set(actual) != expected:
        lint.fail(f"{shard}: X coverage IDs do not equal assignment; "
                  f"missing={sorted(expected - set(actual))}, extra={sorted(set(actual) - expected)}")
    claims_by_id = _claims_rows_by_id(claims_rows)
    for row in coverage:
        obligation = by_id.get(row["X ID"])
        if obligation is None:
            continue
        _validate_coverage_row(
            lint, shard, row, obligation["anchor"], claims_by_id,
            source_set, audit.parent,
        )


def _validate_u7_b3b_resolutions(lint, audit, shard, text, allocation,
                                 claims_rows, manifest):
    from claim_handoffs import HANDOFF_COLS, HANDOFF_RESOLUTION_COLS
    ledger_path = audit / "_run" / "snapshots" / "claims_b3" / "handoff_ledger.json"
    try:
        ledger = json.loads(ledger_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        lint.fail(f"{ledger_path}: invalid handoff ledger ({exc})")
        return set()
    filed = {entry["id"]: entry for entry in ledger.get("H", [])}
    raw = allocation["Assigned Handoff IDs"]
    assigned = set() if blank_cell(raw) else {item.strip() for item in raw.split(",")}
    rows = _exact_or_empty_table(
        lint, shard, text, HANDOFF_RESOLUTION_COLS, "No assigned handoffs.",
        "Handoffs resolution",
    )
    actual = [row["H ID"] for row in rows]
    if len(actual) != len(set(actual)):
        lint.fail(f"{shard}: duplicate Handoff resolution ID")
    if set(actual) != assigned:
        lint.fail(f"{shard}: handoff resolution IDs do not equal Assigned Handoff IDs; "
                  f"missing={sorted(assigned - set(actual))}, extra={sorted(set(actual) - assigned)}")
    source_set = manifest["paper_source_set"]
    claims_by_id = _claims_rows_by_id(claims_rows)
    cited = set()
    for row in rows:
        original = filed.get(row["H ID"])
        if original is None:
            lint.fail(f"{shard}: resolution names unknown handoff {row['H ID']}")
            continue
        copied = {
            "H ID": original["id"], "Anchor": original["anchor"],
            "Quote": original["quote"],
            "Asserted Substance": original["asserted_substance"],
            "Referenced Objects": original["referenced_objects"],
        }
        for column in HANDOFF_COLS:
            if row[column] != copied[column]:
                lint.fail(f"{shard}: {row['H ID']} changes filed {column}")
        synthetic = {
            "Outcome": "covered" if row["Resolution"] == "resolved" else row["Resolution"],
            "C-ID / Reason": row["C-ID / Reason"], "Evidence": row["Evidence"],
            "Covering Range": row["Covering Range"], "Covering Quote": row["Covering Quote"],
        }
        if row["Resolution"] not in {"resolved", "disposition"}:
            lint.fail(f"{shard}: invalid handoff Resolution {row['Resolution']!r}")
            continue
        _validate_coverage_row(
            lint, shard, synthetic, original["resolved_anchor"], claims_by_id,
            source_set, audit.parent,
        )
        if row["Resolution"] == "resolved" and re.fullmatch(r"C-\d{4}", row["C-ID / Reason"]):
            cited.add(row["C-ID / Reason"])
    return cited


def stage_b2(lint, audit, stream, shard):
    if shard is None:
        lint.fail("b2 requires --shard")
        return
    text = read_text(lint, shard)
    if text is None:
        return
    if stream == "claims":
        alloc, _ = parse_plan(lint, audit / "plans" / "claims_review_plan.md", "Worker ID")
        a = find_alloc_for_shard(alloc, shard)
        if a is None:
            lint.fail(f"{shard}: not found in the plan's allocation table")
            return
        ranges = [parse_range(a["Claim ID Range"]), parse_range(a["Output ID Range"])]
        ranges = [r for r in ranges if r]
        c_rows = table_by_cols(lint, shard, text, CLAIMS_COLS, "claims")
        o_rows = table_by_cols(lint, shard, text, OUTPUT_COLS, "outputs")
        if c_rows is None or o_rows is None:
            return
        check_claims_rows(lint, shard, c_rows)
        check_output_rows(lint, shard, o_rows)
        cids = col(c_rows, CLAIMS_COLS, "Claim ID")
        oids = col(o_rows, OUTPUT_COLS, "Output ID")
        check_unique(lint, shard, cids + oids, "ID")
        for i in cids + oids:
            if not in_ranges(i, ranges):
                lint.fail(f"{shard}: {i} outside assigned ranges")
        for v in col(c_rows, CLAIMS_COLS, "Related Error IDs"):
            if v:
                lint.fail(f"{shard}: Related Error IDs must be empty at b2")
        oset = set(oids)
        for r in c_rows:
            d = dict(zip(CLAIMS_COLS, r))
            for o in re.findall(r"O-\d{4}", d["Output IDs"]):
                if o not in oset:
                    lint.fail(f"{shard}: {d['Claim ID']} references {o} not present in shard")
        check_bidirectional(
            lint, c_rows, CLAIMS_COLS, "Claim ID", "Output IDs",
            o_rows, OUTPUT_COLS, "Output ID", "Claim IDs", f"{shard} C<->O",
        )
        check_abs_paths(lint, shard, c_rows, CLAIMS_COLS, ["Code/Data Source"])
        check_abs_paths(lint, shard, o_rows, OUTPUT_COLS, ["Output Path/Pattern", "Producing Script"])
        manifest_path = audit / "_run" / "manifest.json"
        if manifest_path.is_file():
            try:
                manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            except json.JSONDecodeError:
                manifest = {}
            if manifest.get("paper_source_set") is not None:
                _validate_u7_b2_shard(lint, audit, shard, text, a, c_rows, manifest)
    else:
        alloc, _ = parse_plan(lint, audit / "plans" / "code_error_review_plan.md", "Chunk ID")
        a = find_alloc_for_shard(alloc, shard)
        if a is None:
            lint.fail(f"{shard}: not found in the plan's allocation table")
            return
        ranges = [r for r in [parse_range(a["Error ID Range"])] if r]
        e_rows = table_by_cols(lint, shard, text, ERROR_COLS, "code errors")
        if e_rows is None:
            return
        check_error_rows(lint, shard, e_rows)
        eids = col(e_rows, ERROR_COLS, "Error ID")
        check_unique(lint, shard, eids, "Error ID")
        for i in eids:
            if not in_ranges(i, ranges):
                lint.fail(f"{shard}: {i} outside assigned range")
        for v in col(e_rows, ERROR_COLS, "Related Claim IDs"):
            if v:
                lint.fail(f"{shard}: Related Claim IDs must be empty at b2")
        check_abs_paths(lint, shard, e_rows, ERROR_COLS, ["Code/Data Source", "Code Location"])
    shard_footer(lint, shard, text)
    entries, coverage = typed_shard_footer(lint, shard, text, stream)
    validate_footer_candidates(lint, shard, text, stream, entries, coverage)


def load_register(lint, path, cols, allow_extra=False):
    text = read_text(lint, path)
    if text is None:
        return None
    if allow_extra:
        for headers, rows, _ in parse_tables(text):
            if set(cols) <= set(headers):
                return headers, rows
        lint.fail(f"{path}: no table containing required columns")
        return None
    rows = table_by_cols(lint, path, text, cols, path.name)
    if rows is None:
        return None
    return cols, [r for r in rows if not is_example_row(r)]


def code_review_inventory(lint, plan, allocations):
    """Return inventory path -> owning declared shard, including hygiene."""
    text = read_text(lint, plan)
    if text is None:
        return {}
    shard_by_chunk = {
        row["Chunk ID"]: normalized_audit_path(row["Shard File"])
        for row in allocations or []
    }
    inventory = {}
    script_table = hygiene_table = None
    for headers, rows, _line in parse_tables(text):
        if "Script" in headers and "Chunk" in headers and script_table is None:
            script_table = (headers, rows)
        if headers == ["Hygiene File", "Chunk"]:
            hygiene_table = (headers, rows)
    for label, table, path_col in (
            ("script", script_table, "Script"),
            ("hygiene", hygiene_table, "Hygiene File")):
        if table is None:
            lint.fail(f"{plan}: missing {label} inventory table")
            continue
        headers, rows = table
        for index, row in enumerate(rows, start=1):
            if len(row) != len(headers):
                lint.fail(f"{plan}: malformed {label} inventory row {index}")
                continue
            data = dict(zip(headers, row))
            path = data[path_col].strip().strip("`")
            owner = shard_by_chunk.get(data["Chunk"])
            if not owner:
                lint.fail(
                    f"{plan}: {label} inventory path {path!r} names unknown chunk "
                    f"{data['Chunk']!r}"
                )
                continue
            if path in inventory:
                lint.fail(f"{plan}: inventory path {path!r} is declared more than once")
            inventory[path] = owner
    hygiene_owners = sorted({owner for path, owner in inventory.items()
                             if hygiene_table and path in {
                                 row[0].strip().strip("`") for row in hygiene_table[1]
                                 if len(row) == 2
                             }})
    if len(hygiene_owners) != 1:
        lint.fail(
            f"{plan}: hygiene files must belong to exactly one shard, got {hygiene_owners}"
        )
    else:
        inventory[HYGIENE_SINGLETON] = hygiene_owners[0]
    return inventory


def manifest_shard_state(manifest, stage, wire):
    entry = manifest.get("stages", {}).get(stage, {}) if isinstance(manifest, dict) else {}
    shards = entry.get("shards", {}) if isinstance(entry, dict) else {}
    for raw, value in (shards.items() if isinstance(shards, dict) else ()):
        if normalized_audit_path(raw) == wire and isinstance(value, dict):
            return value.get("status")
    return None


def reconcile_code_coverage(lint, audit, plan, allocations, manifest):
    """Prove every code/hygiene inventory key has one earned outcome."""
    inventory = code_review_inventory(lint, plan, allocations)
    observed = {}
    for allocation in allocations or []:
        wire = normalized_audit_path(allocation["Shard File"])
        if manifest_shard_state(manifest, "code_b2", wire) == "blocked":
            continue
        path = audit_path(audit, wire)
        text = read_text(lint, path)
        if text is None:
            continue
        entries, coverage = typed_shard_footer(lint, path, text, "code")
        validate_footer_candidates(lint, path, text, "code", entries, coverage)
        for row in coverage:
            row = {**row, "shard": wire}
            observed.setdefault(row["script"], []).append(row)
    unreviewed = []
    for path, owner in sorted(inventory.items()):
        rows = observed.get(path, [])
        blocked_owner = manifest_shard_state(manifest, "code_b2", owner) == "blocked"
        if blocked_owner:
            if rows:
                lint.fail(
                    f"coverage: blocked owner {owner} for {path!r} conflicts with "
                    f"{len(rows)} coverage row(s)"
                )
            unreviewed.append(path)
            continue
        if len(rows) != 1:
            lint.fail(
                f"coverage: inventory path {path!r} has {len(rows)} coverage rows "
                "across all shards (expected exactly 1)"
            )
    for path, rows in sorted(observed.items()):
        if path not in inventory:
            lint.fail(f"coverage: shard row names non-inventory path {path!r}")
        if len(rows) > 1:
            outcomes = [row["outcome"] for row in rows]
            lint.fail(f"coverage: conflicting outcomes for {path!r}: {outcomes}")
    earned = {
        path: (rows[0]["outcome"] if len(rows) == 1 else None)
        for path, rows in sorted(observed.items()) if path in inventory
    }
    return unreviewed, earned


# U6a phase-C erratum (checklist §3): the b3/b3b certification obligations
# read immutable evidence only — promoted canonical registers, frozen stage
# snapshots, and shard files — never ``audit/_staging``. Promotion atomically
# renames staging over canon, so at certification time canon IS the promoted
# stage output; once a later stage freezes its pre-merge snapshot, that copy
# is the exact stage-era state for every re-verification after canon evolves.
POST_STAGE_SNAPSHOTS = {
    ("b3", "claims"): ("claims_b3b",),
    ("b3", "code"): ("code_b3d",),
    ("b3b", "claims"): ("claims_adjudication", "claims_b6a"),
    ("b3b", "code"): ("code_b6a",),
}


def promoted_register(lint, audit, boundary, stream, fname, cols):
    """Load the frozen post-stage register evidence for one boundary."""
    for name in POST_STAGE_SNAPSHOTS[(boundary, stream)]:
        path = audit / "_run" / "snapshots" / name / fname
        if path.is_file():
            return path, load_register(lint, path, cols)
    path = audit / fname
    return path, load_register(lint, path, cols)


def _validate_handoff_report_block(lint, audit, report_path, report, stage):
    ledger_path = audit / "_run" / "snapshots" / stage / "handoff_ledger.json"
    if not ledger_path.is_file():
        return
    try:
        payload = ledger_path.read_bytes()
        ledger = json.loads(payload)
    except (OSError, json.JSONDecodeError) as exc:
        lint.fail(f"{ledger_path}: invalid immutable handoff ledger ({exc})")
        return
    expected = {
        "H": len(ledger.get("H", [])), "X": len(ledger.get("X", [])),
        "sha256": hashlib.sha256(payload).hexdigest(),
    }
    if report.get("handoff_ledger") != expected:
        lint.fail(f"{report_path}: handoff_ledger block disagrees with {ledger_path}")


def stage_b3(lint, audit, stream, manifest):
    worker_stage = f"{stream}_b2"
    # Direct synthetic row tests from before U6 do not model the worker stage.
    # Every initialized production manifest does, so the read-layer contract is
    # enforced at all real b3 certification boundaries without weakening those
    # older, intentionally narrow test fixtures.
    enforce_read_layer = worker_stage in manifest.get("stages", {})
    if stream == "claims":
        plan = audit / "plans" / "claims_review_plan.md"
        alloc, _ = parse_plan(lint, plan, "Worker ID")
        ranges = alloc_ranges(lint, plan, alloc or [], ["Claim ID Range", "Output ID Range"]) if alloc else []
        c_path, c = promoted_register(
            lint, audit, "b3", stream, "claims_register.md", CLAIMS_COLS)
        o_path, o = promoted_register(
            lint, audit, "b3", stream, "output_register.md", OUTPUT_COLS)
        report_path = audit / "_run" / "merge_report_claims.json"
        if c is None or o is None:
            return
        _, c_rows = c
        _, o_rows = o
        check_claims_rows(lint, c_path, c_rows)
        check_output_rows(lint, o_path, o_rows)
        cids = col(c_rows, CLAIMS_COLS, "Claim ID")
        oids = col(o_rows, OUTPUT_COLS, "Output ID")
        check_unique(lint, c_path.parent, cids + oids, "ID")
        for i in cids + oids:
            if ranges and not in_ranges(i, ranges):
                lint.fail(f"merged register: {i} outside union of planned worker ranges")
        for v in col(c_rows, CLAIMS_COLS, "Related Error IDs"):
            if v:
                lint.fail("merged claims: Related Error IDs must still be blank at b3")
        check_bidirectional(
            lint, c_rows, CLAIMS_COLS, "Claim ID", "Output IDs",
            o_rows, OUTPUT_COLS, "Output ID", "Claim IDs", "b3 C<->O",
        )
        counts = {"claims_register.md": len(c_rows), "output_register.md": len(o_rows)}
        staging_ids = set(cids + oids)
        unreviewed = None
    else:
        plan = audit / "plans" / "code_error_review_plan.md"
        alloc, _ = parse_plan(lint, plan, "Chunk ID")
        ranges = alloc_ranges(lint, plan, alloc or [], ["Error ID Range"]) if alloc else []
        e_path, e = promoted_register(
            lint, audit, "b3", stream, "code_error_register.md", ERROR_COLS)
        report_path = audit / "_run" / "merge_report_code.json"
        if e is None:
            return
        _, e_rows = e
        check_error_rows(lint, e_path, e_rows)
        eids = col(e_rows, ERROR_COLS, "Error ID")
        check_unique(lint, e_path.parent, eids, "Error ID")
        for i in eids:
            if ranges and not in_ranges(i, ranges):
                lint.fail(f"merged register: {i} outside union of planned chunk ranges")
        for v in col(e_rows, ERROR_COLS, "Related Claim IDs"):
            if v:
                lint.fail("merged errors: Related Claim IDs must still be blank at b3")
        if enforce_read_layer:
            unreviewed, coverage_outcomes = reconcile_code_coverage(
                lint, audit, plan, alloc, manifest)
        else:
            unreviewed, coverage_outcomes = [], {}
        counts = {"code_error_register.md": len(e_rows)}
        staging_ids = set(eids)
    rep_text = read_text(lint, report_path)
    if rep_text is None:
        return
    try:
        rep = json.loads(rep_text)
    except json.JSONDecodeError as exc:
        lint.fail(f"{report_path}: invalid JSON ({exc})")
        return
    if stream == "claims":
        _validate_handoff_report_block(
            lint, audit, report_path, rep, "claims_b3"
        )
    if enforce_read_layer:
        reconcile_footer_dispositions(
            lint, audit, alloc, stream, report_path, rep, staging_ids, manifest,
            worker_stage,
        )
    if stream == "code" and enforce_read_layer:
        reported = rep.get("unreviewed_files")
        if not isinstance(reported, list):
            lint.fail(f"{report_path}: must carry list-valued 'unreviewed_files'")
        elif sorted(reported) != sorted(unreviewed):
            lint.fail(
                f"{report_path}: unreviewed_files {sorted(reported)} != "
                f"manifest-backed blocked inventory {sorted(unreviewed)}"
            )
        recorded_outcomes = rep.get("coverage_outcomes")
        if not isinstance(recorded_outcomes, dict):
            lint.fail(f"{report_path}: must carry object-valued 'coverage_outcomes'")
        elif recorded_outcomes != coverage_outcomes:
            lint.fail(
                f"{report_path}: coverage_outcomes disagree with shard evidence; "
                f"recorded={recorded_outcomes}, observed={coverage_outcomes}"
            )
    for reg, n in counts.items():
        entry = rep.get(reg)
        if not isinstance(entry, dict):
            lint.fail(f"{report_path}: missing per-register entry '{reg}'")
            continue
        try:
            sr, dd, ad = int(entry["shard_rows"]), int(entry["dedup_removed"]), int(entry["added"])
        except (KeyError, ValueError, TypeError):
            lint.fail(f"{report_path}: '{reg}' must carry integer shard_rows/dedup_removed/added")
            continue
        if sr - dd != ad:
            lint.fail(f"{report_path}: '{reg}' identity violated: {sr} - {dd} != {ad}")
        if ad != n:
            lint.fail(f"{report_path}: '{reg}' added={ad} but merged register has {n} rows")
        for key in ("conflicts", "coverage_gaps", "blocked_shards"):
            if not isinstance(entry.get(key), list):
                lint.fail(f"{report_path}: '{reg}' must carry list-valued '{key}'")


def stage_b3b(lint, audit, stream, manifest):
    """Second-read recall sweep merge: new rows only, all unverified candidates, no b3 row
    deleted or mutated. Compares the promoted post-b3b evidence against the
    pre-sweep (b3b) snapshot; never reads ``audit/_staging`` (phase-C erratum)."""
    snap = audit / "_run" / "snapshots" / SNAP_KEY[f"b3b-{stream}"]
    if stream == "claims":
        plan = audit / "plans" / "claims_second_read_plan.md"
        alloc, coord = second_read_allocations(lint, plan, stream)
        ranges = alloc_ranges(lint, plan, alloc or [], ["Claim ID Range", "Output ID Range"]) if alloc else []
        files = [("claims_register.md", CLAIMS_COLS, "Claim ID"), ("output_register.md", OUTPUT_COLS, "Output ID")]
        report_path = audit / "_run" / "merge_report_claims_b3b.json"
        b1_plan = audit / "plans" / "claims_review_plan.md"
        b1_alloc, b1_coord = parse_plan(lint, b1_plan, "Worker ID")
        b1_ranges = alloc_ranges(lint, b1_plan, b1_alloc or [], ["Claim ID Range", "Output ID Range"]) if b1_alloc else []
    else:
        plan = audit / "plans" / "code_error_second_read_plan.md"
        alloc, coord = second_read_allocations(lint, plan, stream)
        ranges = alloc_ranges(lint, plan, alloc or [], ["Error ID Range"]) if alloc else []
        files = [("code_error_register.md", ERROR_COLS, "Error ID")]
        report_path = audit / "_run" / "merge_report_code_b3b.json"
        b1_plan = audit / "plans" / "code_error_review_plan.md"
        b1_alloc, b1_coord = parse_plan(lint, b1_plan, "Chunk ID")
        b1_ranges = alloc_ranges(lint, b1_plan, b1_alloc or [], ["Error ID Range"]) if b1_alloc else []
        detector_ranges = []
        mapping_path = audit / "_run" / "detector_mapping.md"
        if mapping_path.exists():
            try:
                declared, _display, _rows = detector_mapping.load_mapping(mapping_path)
                detector_ranges.append(declared)
            except detector_mapping.MappingError as exc:
                lint.fail(f"{mapping_path}: {exc}")
    check_manifest_worker_shards(
        lint, manifest, f"{stream}_b3b", plan, "Worker ID", alloc,
    )
    # machinery (a): b3b ranges disjoint from b1 ranges, the merge-coordinator range, and each other
    check_disjoint(lint, plan, ranges + coord + b1_ranges + b1_coord
                   + (detector_ranges if stream == "code" else []))
    check_identifier_exhaustion(
        lint, plan, ranges + coord + b1_ranges + b1_coord
        + (detector_ranges if stream == "code" else []),
    )
    handoff_resolution_cids = set()
    if stream == "claims" and (
            audit / "_run" / "snapshots" / "claims_b3" / "handoff_ledger.json").is_file():
        from claim_handoffs import HANDOFF_RESOLUTION_COLS
        for allocation in alloc or []:
            if manifest_shard_state(
                    manifest, "claims_b3b",
                    normalized_audit_path(allocation["Shard File"])) == "blocked":
                continue
            shard_path = audit_path(audit, allocation["Shard File"])
            shard_text = read_text(lint, shard_path)
            if shard_text is None:
                continue
            for row in tables_by_cols(
                    lint, shard_path, shard_text, HANDOFF_RESOLUTION_COLS,
                    "handoff resolutions"):
                if row["Resolution"] == "resolved":
                    handoff_resolution_cids.add(row["C-ID / Reason"])
    if stream == "code" and "code_b3d" in (
            manifest.get("stages", {}) if isinstance(manifest, dict) else {}):
        # The post-b3d ordering of the b3b baseline is enforced structurally:
        # the plan builder refuses to run before code_b3d is certified done,
        # and --check reads only the frozen code_b3b snapshot (test-pinned per
        # the checklist; the phase-C erratum removed the runtime mtime check).
        command = [
            sys.executable, str(Path(__file__).with_name("build_second_read_plan.py")),
            str(audit.parent), "--audit-dir", str(audit), "--check",
        ]
        checked = subprocess.run(command, capture_output=True, text=True)
        if checked.returncode:
            lint.fail(
                "code second-read plan does not match recomputed trigger/sample: "
                + (checked.stderr or checked.stdout).strip()
            )

    total_new = 0
    for f, cols, idc in files:
        st_path, st = promoted_register(lint, audit, "b3b", stream, f, cols)
        sn = load_register(lint, snap / f, cols)
        if st is None or sn is None:
            continue
        _, st_rows = st
        _, sn_rows = sn
        if f == "claims_register.md":
            check_claims_rows(lint, st_path, st_rows)
        elif f == "output_register.md":
            check_output_rows(lint, st_path, st_rows)
        else:
            check_error_rows(lint, st_path, st_rows)
        st_ids = col(st_rows, cols, idc)
        check_unique(lint, st_path, st_ids, idc)
        st_by = {dict(zip(cols, r))[idc]: dict(zip(cols, r)) for r in st_rows}
        sn_by = {dict(zip(cols, r))[idc]: dict(zip(cols, r)) for r in sn_rows}
        st_set, sn_set = set(st_by), set(sn_by)
        for i in sorted(sn_set - st_set):
            lint.fail(f"{st_path}: b3 row {i} deleted at second-read merge (rows are never deleted)")
        for i in sorted(sn_set & st_set):
            for c in cols:
                if st_by[i][c] != sn_by[i][c]:
                    lint.fail(f"{st_path}: b3 row {i} column '{c}' changed at second-read merge (the sweep only adds rows)")
        new_ids = st_set - sn_set
        total_new += len(new_ids)
        for i in sorted(new_ids):
            if ranges and not in_ranges(i, ranges):
                lint.fail(f"{st_path}: new second-read row {i} outside the b3b-allocated ranges")
            status = st_by[i]["Status"]
            if f == "code_error_register.md" and status != "candidate":
                lint.fail(f"{st_path}: new second-read row {i} status '{status}' (must be 'candidate')")
            elif (f == "claims_register.md"
                  and status not in {"inconsistent", "unclear"}
                  and i not in handoff_resolution_cids):
                lint.fail(f"{st_path}: new second-read claim {i} status '{status}' (must be 'inconsistent' or 'unclear')")
            elif f == "output_register.md" and status not in {"mapped", "orphan", "unclear", "inconsistent"}:
                lint.fail(f"{st_path}: new second-read output {i} status '{status}' (must be mapped/orphan/unclear/inconsistent)")
        link_col = {"claims_register.md": "Related Error IDs", "code_error_register.md": "Related Claim IDs"}.get(f)
        if link_col:
            for v in col(st_rows, cols, link_col):
                if v:
                    lint.fail(f"{st_path}: {link_col} must still be blank at b3b (cross-link is a later stage)")
    if stream == "claims":
        _c_path, c = promoted_register(
            lint, audit, "b3b", stream, "claims_register.md", CLAIMS_COLS)
        _o_path, o = promoted_register(
            lint, audit, "b3b", stream, "output_register.md", OUTPUT_COLS)
        if c and o:
            check_bidirectional(
                lint, c[1], CLAIMS_COLS, "Claim ID", "Output IDs",
                o[1], OUTPUT_COLS, "Output ID", "Claim IDs", "b3b C<->O",
            )
    rep_text = read_text(lint, report_path)
    if rep_text is None:
        return
    try:
        rep = json.loads(rep_text)
    except json.JSONDecodeError as exc:
        lint.fail(f"{report_path}: invalid JSON ({exc})")
        return
    if stream == "claims":
        _validate_handoff_report_block(
            lint, audit, report_path, rep, "claims_b3b"
        )
    staging_ids = set()
    for f, cols, idc in files:
        _path, loaded = promoted_register(lint, audit, "b3b", stream, f, cols)
        if loaded:
            staging_ids.update(col(loaded[1], cols, idc))
    reconcile_footer_dispositions(
        lint, audit, alloc, stream, report_path, rep, staging_ids, manifest,
        f"{stream}_b3b",
    )
    added_total = 0
    for f, cols, idc in files:
        entry = rep.get(f)
        if not isinstance(entry, dict):
            lint.fail(f"{report_path}: missing per-register entry '{f}'")
            continue
        try:
            sr, dd, ad = int(entry["shard_rows"]), int(entry["dedup_removed"]), int(entry["added"])
        except (KeyError, ValueError, TypeError):
            lint.fail(f"{report_path}: '{f}' must carry integer shard_rows/dedup_removed/added")
            continue
        if sr - dd != ad:
            lint.fail(f"{report_path}: '{f}' identity violated: {sr} - {dd} != {ad}")
        added_total += ad
    if added_total != total_new:
        lint.fail(f"{report_path}: added total {added_total} != {total_new} new row(s) in the merged register")


# --- U8 (b): b3b second-read shard boundary check ----------------------------
#
# Mirrors the b2 first-pass shard check but against the second-read allocation
# plan (`claims_second_read_plan.md` / `code_error_second_read_plan.md`, keyed by
# `Worker ID`). A b3b shard is a first-pass-shaped shard (canonical columns, a
# footer) whose new rows are all UNVERIFIED: code rows are `candidate`; claims
# rows are `inconsistent`/`unclear`; output rows use the permitted output set
# (`mapped`/`orphan`/`unclear`/`inconsistent`, never `listed`/`confirmed`). These
# match the second-read-worker skeleton and the b3b merge check in stage_b3b.

B3B_CLAIM_STATUSES = {"inconsistent", "unclear"}
B3B_OUTPUT_STATUSES = {"mapped", "orphan", "unclear", "inconsistent"}


def second_read_plan_path(audit, stream):
    return audit / "plans" / (
        "claims_second_read_plan.md" if stream == "claims"
        else "code_error_second_read_plan.md")


def stage_b3b_shard(lint, audit, stream, shard):
    plan = second_read_plan_path(audit, stream)
    alloc, _ = second_read_allocations(lint, plan, stream)
    text = read_text(lint, shard)
    if text is None:
        return
    a = find_alloc_for_shard(alloc, shard)
    if a is None:
        lint.fail(f"{shard}: not found in the second-read plan's allocation table")
        return
    if stream == "claims":
        ranges = [r for r in (parse_range(a.get("Claim ID Range", "")),
                              parse_range(a.get("Output ID Range", ""))) if r]
        c_rows = table_by_cols(lint, shard, text, CLAIMS_COLS, "claims")
        o_rows = table_by_cols(lint, shard, text, OUTPUT_COLS, "outputs")
        if c_rows is None or o_rows is None:
            return
        check_claims_rows(lint, shard, c_rows)
        check_output_rows(lint, shard, o_rows)
        manifest = {}
        manifest_path = audit / "_run" / "manifest.json"
        if manifest_path.is_file():
            try:
                manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            except json.JSONDecodeError:
                pass
        handoff_claim_ids = set()
        if (audit / "_run" / "snapshots" / "claims_b3" / "handoff_ledger.json").is_file():
            handoff_claim_ids = _validate_u7_b3b_resolutions(
                lint, audit, shard, text, a, c_rows, manifest
            )
        cids = col(c_rows, CLAIMS_COLS, "Claim ID")
        oids = col(o_rows, OUTPUT_COLS, "Output ID")
        check_unique(lint, shard, cids + oids, "ID")
        for i in cids + oids:
            if not in_ranges(i, ranges):
                lint.fail(f"{shard}: {i} outside the second-read-allocated ranges")
        for r in c_rows:
            d = dict(zip(CLAIMS_COLS, r))
            if d["Status"] not in B3B_CLAIM_STATUSES and d["Claim ID"] not in handoff_claim_ids:
                lint.fail(f"{shard}: second-read claim {d['Claim ID']} status "
                          f"'{d['Status']}' (must be 'inconsistent'/'unclear' unless "
                          "cited by a handoff resolution)")
        for r in o_rows:
            d = dict(zip(OUTPUT_COLS, r))
            if d["Status"] not in B3B_OUTPUT_STATUSES:
                lint.fail(f"{shard}: second-read output {d['Output ID']} status "
                          f"'{d['Status']}' (must be mapped/orphan/unclear/inconsistent)")
        for v in col(c_rows, CLAIMS_COLS, "Related Error IDs"):
            if v:
                lint.fail(f"{shard}: Related Error IDs must be blank at b3b (cross-link is a later stage)")
        check_abs_paths(lint, shard, c_rows, CLAIMS_COLS, ["Code/Data Source"])
        check_abs_paths(lint, shard, o_rows, OUTPUT_COLS, ["Output Path/Pattern", "Producing Script"])
    else:
        ranges = [r for r in (parse_range(a.get("Error ID Range", "")),) if r]
        e_rows = table_by_cols(lint, shard, text, ERROR_COLS, "code errors")
        if e_rows is None:
            return
        check_error_rows(lint, shard, e_rows)
        eids = col(e_rows, ERROR_COLS, "Error ID")
        check_unique(lint, shard, eids, "Error ID")
        for i in eids:
            if not in_ranges(i, ranges):
                lint.fail(f"{shard}: {i} outside the second-read-allocated range")
        for r in e_rows:
            d = dict(zip(ERROR_COLS, r))
            if d["Status"] != "candidate":
                lint.fail(f"{shard}: second-read code row {d['Error ID']} status "
                          f"'{d['Status']}' (must be 'candidate')")
        for v in col(e_rows, ERROR_COLS, "Related Claim IDs"):
            if v:
                lint.fail(f"{shard}: Related Claim IDs must be blank at b3b (cross-link is a later stage)")
        check_abs_paths(lint, shard, e_rows, ERROR_COLS, ["Code/Data Source", "Code Location"])
    shard_footer(lint, shard, text)
    entries, coverage = typed_shard_footer(lint, shard, text, stream)
    validate_footer_candidates(lint, shard, text, stream, entries, coverage)


def recheck_plan_path(audit, stream, supplementary=False):
    name = (SUPPLEMENTARY_PLAN_FILES[stream] if supplementary else
            ("claims_recheck_plan.md" if stream == "claims"
             else "code_error_recheck_plan.md"))
    return audit / "plans" / name


def parse_recheck_plan(lint, audit, stream, supplementary=False):
    plan = recheck_plan_path(audit, stream, supplementary)
    text = read_text(lint, plan)
    if text is None:
        return None
    inventory, clusters = [], []
    for headers, rows, _ in parse_tables(text):
        if "Reason" in headers and any(h == "ID" for h in headers):
            inventory = [dict(zip(headers, r)) for r in rows if len(r) == len(headers)]
        if "Cluster ID" in headers and "Assigned IDs" in headers and "Shard File" in headers:
            clusters = [dict(zip(headers, r)) for r in rows if len(r) == len(headers)]
    if supplementary and not inventory:
        if SUPPLEMENTARY_EMPTY not in text:
            lint.fail(f"{plan}: empty inventory requires exact '{SUPPLEMENTARY_EMPTY}'")
        if clusters:
            lint.fail(f"{plan}: empty supplementary inventory cannot declare clusters")
    else:
        if not inventory:
            lint.fail(f"{plan}: inventory table (columns incl. 'ID' and 'Reason') not found")
        if not clusters:
            lint.fail(f"{plan}: cluster table (Cluster ID / Assigned IDs / Shard File) not found")
    if "audit_readme.md" not in text:
        lint.fail(f"{plan}: missing pointer to the verdict/evidence vocabulary in audit_readme.md")
    return plan, inventory, clusters


def parse_detector_mappings(lint, audit):
    path = audit / "_run" / "detector_mapping.md"
    try:
        _declared, _display, rows = detector_mapping.load_mapping(path)
        return detector_mapping.actionable_rows(rows)
    except detector_mapping.MappingError as exc:
        lint.fail(f"{path}: {exc}")
        return []


def _names_source(text, source_id):
    return bool(re.search(
        rf"(?<![A-Za-z0-9_-]){re.escape(source_id)}(?![A-Za-z0-9_-])",
        text or "",
    ))


def check_detector_mapping_b4(lint, audit, plan, inventory):
    mappings = parse_detector_mappings(lint, audit)
    inv_by_id = {}
    for row in inventory:
        inv_by_id.setdefault(row.get("ID", ""), []).append(row)
    register = load_register(lint, audit / "code_error_register.md", ERROR_COLS)
    canonical = {}
    if register:
        canonical = {dict(zip(ERROR_COLS, row))["Error ID"]:
                     dict(zip(ERROR_COLS, row)) for row in register[1]}
    by_error = {}
    for row in mappings:
        by_error.setdefault(row["Error ID"], []).append(row)
    for eid, rows in by_error.items():
        inv_rows = inv_by_id.get(eid, [])
        if len(inv_rows) != 1:
            lint.fail(f"{plan}: mapped Error ID {eid} is absent "
                      "from the b4 inventory")
            continue
        evidence = inv_rows[0].get("Likely Evidence", "")
        for source_id in sorted({row["Source ID"] for row in rows}):
            if not _names_source(evidence, source_id):
                lint.fail(f"{plan}: inventory row {eid} Likely Evidence does not name "
                          f"mapped source ID {source_id}")
        if any(row["Mapping Kind"] == "new_candidate" for row in rows):
            canonical_row = canonical.get(eid)
            if canonical_row is None:
                continue  # canon_ids reports this independently
            if canonical_row.get("Status") != "candidate":
                lint.fail(f"{plan}: new_candidate mapping to {eid} must be a "
                          "candidate canonical row")
            if (any(row["Channel"] == "DU" for row in rows)
                    and canonical_row.get("Error Type") != "sample_filter_or_flag_error"):
                lint.fail(f"{plan}: DU new_candidate mapping to {eid} must be typed "
                          "sample_filter_or_flag_error")


def _all_code_recheck_ledger_rows(lint, audit, supplementary=False):
    rows = []
    root = audit / (SUPPLEMENTARY_SHARD_DIRS["code"] if supplementary
                    else "_code_error_recheck")
    if not root.is_dir():
        return rows
    for path in sorted(root.rglob("*.md")):
        text = read_text(lint, path)
        if text is None:
            continue
        for headers, table_rows, _ in parse_tables(text):
            if headers == CODE_LEDGER_COLS:
                rows.extend((path, dict(zip(headers, row)))
                            for row in table_rows if len(row) == len(headers))
    return rows


def check_detector_mapping_b6(lint, audit, final_path=None):
    plan = recheck_plan_path(audit, "code")
    mappings = parse_detector_mappings(lint, audit)
    if not mappings:
        return
    mapped = {}
    mapping_by_key = {}
    for row in mappings:
        key = (row["Channel"], row["Source ID"], row["Witness ID"])
        mapped.setdefault(row["Error ID"], []).append(row)
        mapping_by_key[key] = row
    ledger_rows = _all_code_recheck_ledger_rows(lint, audit)
    ledgers_by_id = {}
    for path, row in ledger_rows:
        ledgers_by_id.setdefault(row.get("ID", ""), []).append((path, row))
    frozen_post_b6a = audit / "_run/snapshots/code_b6b/code_error_register.md"
    final_path = final_path or (
        frozen_post_b6a if frozen_post_b6a.is_file() else audit / "code_error_register.md")
    final = load_register(lint, final_path, ERROR_COLS)
    final_by_id = {}
    if final:
        final_by_id = {dict(zip(ERROR_COLS, row))["Error ID"]:
                       dict(zip(ERROR_COLS, row)) for row in final[1]}
    snap = load_register(
        lint, audit / "_run/snapshots/code_b6a/code_error_register.md", ERROR_COLS)
    snapshot_by_id = {}
    if snap:
        snapshot_by_id = {dict(zip(ERROR_COLS, row))["Error ID"]:
                          dict(zip(ERROR_COLS, row)) for row in snap[1]}

    summary = audit / "code_error_recheck_summary.md"
    summary_text = read_text(lint, summary) or ""
    lineage_rows = tables_by_cols(
        lint, summary, summary_text, LINEAGE_COLS, "split lineage")
    lineage, split_originals = {}, set()
    for row in lineage_rows:
        key = (row["Channel"], row["Source ID"], row["Witness ID"])
        mapping = mapping_by_key.get(key)
        if mapping is None or mapping["Error ID"] != row["Original Error ID"]:
            lint.fail(f"{summary}: lineage row names an unmapped original/witness")
            continue
        full = (row["Original Error ID"], *key)
        if full in lineage:
            lint.fail(f"{summary}: mapped witness {'/'.join(key)} appears twice in lineage")
        lineage[full] = row["Descendant Error ID"]
        split_originals.add(row["Original Error ID"])
    for original in split_originals:
        expected = {(row["Channel"], row["Source ID"], row["Witness ID"])
                    for row in mapped.get(original, [])}
        covered = {key[1:] for key in lineage if key[0] == original}
        if covered != expected:
            lint.fail(f"{summary}: split lineage for {original} does not exactly cover its mapped witnesses")
        descendants = {value for key, value in lineage.items() if key[0] == original}
        if len(descendants) < 2:
            lint.fail(f"{summary}: split {original} must have at least two non-empty descendants")
        for descendant in descendants:
            if descendant not in final_by_id:
                lint.fail(f"{summary}: split descendant {descendant} is absent from final register")

    boundary_path = audit / "_run/code_b6a/witness_outcomes.md"
    boundary_text = read_text(lint, boundary_path) or ""
    post_rows = tables_by_cols(
        lint, boundary_path, boundary_text, POST_WITNESS_COLS,
        "post-boundary witness outcomes", required=True)
    post_by_key = {}
    for row in post_rows:
        key = tuple(row[field] for field in ("Channel", "Source ID", "Witness ID"))
        if key in post_by_key:
            lint.fail(f"{boundary_path}: duplicate post-boundary key {'/'.join(key)}")
        post_by_key[key] = row
        if key not in mapping_by_key:
            lint.fail(f"{boundary_path}: post-boundary row {'/'.join(key)} is not mapped")
    if set(post_by_key) != set(mapping_by_key):
        lint.fail(f"{boundary_path}: post-boundary keys do not exactly close detector mapping")
    dismissal_tables = tables_by_cols(
        lint, boundary_path, boundary_text, ["Error ID"], "assembled dismissals")
    assembled = {row["Error ID"] for row in dismissal_tables}

    receipt_path = audit / "_run/code_b6a/dismissal_receipts.md"
    receipt_text = read_text(lint, receipt_path) or ""
    receipts = tables_by_cols(lint, receipt_path, receipt_text, RECEIPT_COLS,
                              "dismissal receipts")
    receipt_ids = [row["Receipt ID"] for row in receipts]
    check_unique(lint, receipt_path, receipt_ids, "Receipt ID")

    effective_groups = {}
    for original, mapping_rows in mapped.items():
        eid = original
        dispositions = ledgers_by_id.get(eid, [])
        if len(dispositions) != 1:
            lint.fail(f"{plan}: mapped Error ID {eid} has {len(dispositions)} ledger "
                      "rows (expected exactly one disposition)")
            continue
        path, ledger = dispositions[0]
        evidence = ledger.get("Evidence Checked", "")
        for source_id in sorted({row["Source ID"] for row in mapping_rows}):
            if not _names_source(evidence, source_id):
                lint.fail(f"{path}: mapped source ID {source_id} missing from {eid} "
                          "Evidence Checked")
        for mapping in mapping_rows:
            full = (eid, mapping["Channel"], mapping["Source ID"], mapping["Witness ID"])
            effective = lineage.get(full, eid)
            effective_groups.setdefault(effective, []).append((mapping, path, ledger))

    for eid, group in effective_groups.items():
        mapping_rows = [item[0] for item in group]
        path, ledger = group[0][1], group[0][2]
        final_row = final_by_id.get(eid)
        if final_row is None:
            lint.fail(f"{plan}: mapped Error ID {eid} absent from final staging register")
            continue
        status = final_row.get("Status", "")
        verdict = ledger.get("Verdict", "")
        expected_disposition = expected_code_disposition(ledger)
        if expected_disposition is None:
            lint.fail(f"{path}: mapped Error ID {eid} has invalid code verdict '{verdict}'")
            continue
        wanted, wanted_severity = expected_disposition
        if status != wanted:
            lint.fail(f"{path}: mapped Error ID {eid} verdict '{verdict}' requires "
                      f"final status '{wanted}', found final status '{status}'")
        if final_row.get("Severity", "") != wanted_severity:
            lint.fail(f"{path}: mapped Error ID {eid} final Severity "
                      f"{final_row.get('Severity')!r} disagrees with applied proposal {wanted_severity!r}")
        for column, replacement in parse_field_patches(lint, path, ledger).items():
            if final_row.get(column) != replacement:
                lint.fail(f"{path}: mapped Error ID {eid} field patch for {column} was not applied")

        keys = {(row["Channel"], row["Source ID"], row["Witness ID"])
                for row in mapping_rows}
        if verdict == "not_error":
            if eid not in assembled:
                lint.fail(f"{boundary_path}: mapped not_error {eid} is absent from Assembled dismissals")
            allowed_records = set(list_cell(ledger.get("Verification Record IDs", "")))
            covered = set()
            for receipt in receipts:
                key = tuple(receipt[field] for field in
                            ("Channel", "Source ID", "Witness ID"))
                if (key in keys and receipt["Record ID"] in allowed_records
                        and receipt["Accepted (yes/no)"] == "yes"):
                    covered.add(key)
                    post = post_by_key.get(key, {})
                    if receipt["Receipt ID"] not in list_cell(post.get("Receipt IDs", "")):
                        lint.fail(f"{boundary_path}: witness {'/'.join(key)} does not bind qualifying receipt {receipt['Receipt ID']}")
            if covered != keys:
                lint.fail(f"{receipt_path}: mapped not_error {eid} lacks qualifying receipt coverage")
        elif eid in assembled:
            lint.fail(f"{boundary_path}: Assembled dismissals lists non-not_error mapped row {eid}")

        if verdict == "duplicate":
            target = ledger.get("Duplicate Target", "").strip()
            if target not in mapped:
                lint.fail(f"{path}: guarded duplicate {eid} target {target!r} is not mechanically mapped")
                continue
            target_row = snapshot_by_id.get(target) or final_by_id.get(target)
            source_row = snapshot_by_id.get(eid) or final_row
            if target_row is None or target_row.get("Error Type") != source_row.get("Error Type"):
                lint.fail(f"{path}: duplicate {eid} and target {target} do not share Error Type")
            source_mechanisms = {post_by_key.get(key, {}).get("Mechanism") for key in keys}
            target_keys = {(row["Channel"], row["Source ID"], row["Witness ID"])
                           for row in mapped[target]}
            target_mechanisms = {post_by_key.get(key, {}).get("Mechanism") for key in target_keys}
            if source_mechanisms != target_mechanisms:
                lint.fail(f"{path}: duplicate {eid} mechanism differs from target {target}")
            coverage = (target_row or {}).get("Code Location", "") + " " + (target_row or {}).get("Code/Data Source", "")
            for mapping in mapping_rows:
                if mapping["Site Anchor"].rsplit(":", 1)[0] not in coverage:
                    lint.fail(f"{path}: duplicate target {target} does not cover {mapping['Site Anchor']}")

    # A split is legal only when its active witnesses actually disagree.
    for original in split_originals:
        tuples = set()
        for row in mapped[original]:
            key = (row["Channel"], row["Source ID"], row["Witness ID"])
            post = post_by_key.get(key, {})
            if post.get("Mechanism") not in {None, "—", "-", ""}:
                tuples.add((post.get("Mechanism"), post.get("Verdict"),
                            post.get("Proposed Severity"), post.get("Duplicate Target")))
        if len(tuples) < 2:
            lint.fail(f"{summary}: split {original} is not justified by distinct active mechanisms")


def canon_ids(lint, audit, stream):
    ids = set()
    if stream == "claims":
        for f, cols, idc in [("claims_register.md", CLAIMS_COLS, "Claim ID"), ("output_register.md", OUTPUT_COLS, "Output ID")]:
            reg = load_register(lint, audit / f, cols)
            if reg:
                ids |= set(col(reg[1], cols, idc))
    else:
        reg = load_register(lint, audit / "code_error_register.md", ERROR_COLS)
        if reg:
            ids |= set(col(reg[1], ERROR_COLS, "Error ID"))
    return ids


# --- U8 (a): the b4 mandatory-recheck inventory computed from canon ----------
#
# A recall guarantee upstream of everything U1 fixes: the recheck can only
# adjudicate rows that actually entered it. The REQUIRED inventory (rows that
# MUST be rechecked) mirrors the pipeline's mechanical inventory rule exactly:
#   * claims  (pipeline-claims.md b4): every severity-bearing row OR every
#     `unclear` row — `Severity != "" OR status == unclear`.
#   * code (pipeline-code-errors.md b4): every `candidate`, plus every
#     `confirmed` with Severity >= 3 — `status == candidate OR (status ==
#     confirmed AND severity >= 3)`.
# The SUBSTANTIVE subset (used only by the `deep` single-ID-cluster rule) is a
# subset of the required set with the SAME predicate on the code side and, on
# the claims side, the U7 definition (issue-flagged OR `unclear`, i.e. exactly
# the required predicate — sampled clean `confirmed` rows are never required and
# never substantive). Sampled clean `confirmed` rows are neither required nor
# substantive here: they may be added to the inventory by the conductor and may
# be grouped, so this check does not fail on their presence, only on the absence
# of a REQUIRED id and on wrong-typed / over-clustered ids.


def _sev_int(sev):
    try:
        return int(sev)
    except (TypeError, ValueError):
        return 0


def required_recheck_ids(lint, audit, stream):
    """Return (required_ids: set, substantive_ids: set) for the b4 inventory.

    `required` = rows the inventory MUST contain (a recall floor).
    `substantive` = the subset that must sit in its own single-ID cluster at
    `deep` depth (U7 definition, mirrored exactly)."""
    required, substantive = set(), set()
    if stream == "claims":
        reg = load_register(lint, audit / "claims_register.md", CLAIMS_COLS)
        if reg:
            for r in reg[1]:
                d = dict(zip(CLAIMS_COLS, r))
                cid, sev, st = d["Claim ID"], d["Severity"], d["Status"]
                if sev or st == "unclear":
                    required.add(cid)
                    substantive.add(cid)  # issue-flagged OR unclear == U7 substantive
    else:
        reg = load_register(lint, audit / "code_error_register.md", ERROR_COLS)
        if reg:
            for r in reg[1]:
                d = dict(zip(ERROR_COLS, r))
                eid, sev, st = d["Error ID"], d["Severity"], d["Status"]
                is_candidate = st == "candidate"
                is_conf_high = st == "confirmed" and _sev_int(sev) >= 3
                if is_candidate or is_conf_high:
                    required.add(eid)
                    substantive.add(eid)  # candidate OR confirmed>=3 == U7 substantive
    return required, substantive


def adjudicator_minted_ids(lint, audit):
    """Return every C-ID minted by a validated reject-and-resolve verdict."""
    path = audit / "_run/claims_adjudication_verdicts.md"
    if not path.is_file():
        return set()
    text = read_text(lint, path) or ""
    minted = set()
    for headers, rows, _line in parse_tables(text):
        if not {"Verdict", "Minted C-ID"} <= set(headers):
            continue
        for raw in rows:
            if len(raw) != len(headers):
                lint.fail(f"{path}: malformed adjudication verdict row")
                continue
            row = dict(zip(headers, raw))
            if row["Verdict"] == "reject_and_resolve":
                minted.add(row["Minted C-ID"])
    return minted


def check_conventions_artifact(lint, audit):
    """Advisory well-formedness check on the optional b3c shared-conventions
    artifact `audit/_run/conventions.md`, consumed by the b4-code recheck grep.

    Non-blocking by design: absent → silent (a package with no multi-site
    convention correctly produces no artifact); present → warn (never fail) if it
    is not a table with the expected header or carries an out-of-vocabulary
    Category. The verdicts come from the code-stream grep, not from lint."""
    path = audit / "_run" / "conventions.md"
    if not path.is_file() or path.stat().st_size == 0:
        return
    try:
        text = path.read_text(encoding="utf-8")
    except (UnicodeDecodeError, OSError):
        lint.warn(f"{path}: could not read the conventions artifact (unreadable)")
        return
    table = None
    for headers, rows, _ in parse_tables(text):
        if headers == CONVENTIONS_COLS:
            table = rows
            break
    if table is None:
        lint.warn(
            f"{path}: no table with the expected header "
            f"{' | '.join(CONVENTIONS_COLS)} (conventions grep may skip it)"
        )
        return
    for r in table:
        d = dict(zip(CONVENTIONS_COLS, r))
        cat = d.get("Category", "")
        if cat and cat not in CONVENTION_CATEGORIES:
            lint.warn(
                f"{path}: convention '{d.get('Convention', '')}' has "
                f"out-of-vocabulary Category '{cat}'"
            )


def stage_b4(lint, audit, stream, manifest=None):
    if stream == "code":
        check_conventions_artifact(lint, audit)  # U2 (preserved — must run first, code only)
    parsed = parse_recheck_plan(lint, audit, stream)
    if parsed is None:
        return
    plan, inventory, clusters = parsed
    if stream == "code":
        check_detector_mapping_b4(lint, audit, plan, inventory)
    canon = canon_ids(lint, audit, stream)
    # U8 (a): the required inventory computed from canon (a recall floor).
    required, substantive = required_recheck_ids(lint, audit, stream)
    adjudicated = adjudicator_minted_ids(lint, audit) if stream == "claims" else set()
    required |= adjudicated
    substantive |= adjudicated
    id_letter = "C" if stream == "claims" else "E"  # the inventory's own ID letter
    inv_ids = {row.get("ID", "") for row in inventory}
    assignments = {}
    for c in clusters:
        for i in re.findall(r"[CEO]-\d{4}", c["Assigned IDs"]):
            assignments.setdefault(i, []).append(c["Cluster ID"])
    # (a1) every REQUIRED id must actually be in the inventory (recall guarantee).
    for i in sorted(required - inv_ids):
        lint.fail(f"{plan}: required recheck ID {i} absent from the b4 inventory "
                  f"(a mandatory row would silently skip the recheck)")
    for row in inventory:
        i = row.get("ID", "")
        if i in adjudicated and row.get("Reason") != "adjudicated_handoff":
            lint.fail(
                f"{plan}: adjudicator-minted {i} requires Reason adjudicated_handoff"
            )
        # (a2) wrong-typed inventory IDs: a claims recheck inventory carries only
        #      C- ids (outputs are never rechecked directly); a code inventory only E-.
        if i and i[0] != id_letter:
            lint.fail(f"{plan}: wrong-typed inventory ID {i} in the {stream} recheck plan "
                      f"(expected a {id_letter}- ID)")
        if i not in canon:
            lint.fail(f"{plan}: inventory ID {i} not found in canonical registers")
        n = len(assignments.get(i, []))
        if n != 1:
            lint.fail(f"{plan}: inventory ID {i} assigned to {n} clusters (expected exactly 1)")
    # (a3) cluster size ceiling, at every depth.
    depth = (manifest or {}).get("review_depth", "standard")
    for c in clusters:
        ids = re.findall(r"[CEO]-\d{4}", c["Assigned IDs"])
        if len(ids) > 8:
            lint.fail(f"{plan}: cluster {c['Cluster ID']} groups {len(ids)} IDs (max 8)")
        # (a4) at deep, a cluster may hold at most one SUBSTANTIVE id.
        if depth == "deep":
            n_sub = sum(1 for i in ids if i in substantive)
            if n_sub > 1:
                lint.fail(f"{plan}: cluster {c['Cluster ID']} groups {n_sub} substantive IDs "
                          f"at deep depth (each substantive ID needs its own cluster)")
    check_unique(lint, plan, [c["Shard File"] for c in clusters], "cluster Shard File")


def stage_b5(lint, audit, stream, shard, manifest, supplementary=False):
    parsed = parse_recheck_plan(lint, audit, stream, supplementary)
    if parsed is None:
        return
    plan, inventory, clusters = parsed
    if shard is None:
        if supplementary and not inventory and not clusters:
            return
        lint.fail(f"{'b5s' if supplementary else 'b5'} requires --shard")
        return
    cluster = None
    for c in clusters:
        if c["Shard File"].strip("`") == str(shard) or str(shard).endswith(c["Shard File"].strip("`")):
            cluster = c
            break
    if cluster is None:
        lint.fail(f"{shard}: not found in the recheck plan's cluster table")
        return
    assigned = set(re.findall(r"[CEO]-\d{4}", cluster["Assigned IDs"]))
    text = read_text(lint, shard)
    if text is None:
        return
    ledger_cols = LEDGER_COLS if stream == "claims" else CODE_LEDGER_COLS
    rows = table_by_cols(lint, shard, text, ledger_cols, "recheck ledger")
    if rows is None:
        return
    verdicts = CLAIMS_VERDICTS if stream == "claims" else ERROR_VERDICTS
    ladder = int(manifest.get("ladder_level", 1)) if manifest else 1
    # U8 (c): verdicts whose ledger row must carry blocker text in a proposal column.
    BLOCKER_VERDICTS = {"blocked", "confirmation_needed", "deferred"}
    seen = []
    evidence_levels = set()
    for r in rows:
        d = dict(zip(ledger_cols, r))
        seen.append(d["ID"])
        if d["ID"] not in assigned:
            lint.fail(f"{shard}: ledger contains unassigned/new ID {d['ID']} (no register IDs may be minted at recheck)")
        if d["Verdict"] not in verdicts:
            lint.fail(f"{shard}: {d['ID']} invalid verdict '{d['Verdict']}'")
        level = d["Evidence Level"]
        if level not in EVIDENCE_LEVELS:
            lint.fail(f"{shard}: {d['ID']} invalid evidence level '{level}'")
        else:
            # U8 (c): an evidence level whose minimum ladder exceeds the run's
            # ladder cannot have been produced — fail it.
            min_ladder = EVIDENCE_MIN_LADDER[level]
            if min_ladder > ladder:
                lint.fail(f"{shard}: {d['ID']} evidence level '{level}' needs ladder "
                          f">= {min_ladder} but the run is at ladder {ladder}")
        evidence_levels.add(level)
        # U8 (c): Evidence Checked must be non-empty on every ledger row.
        if not d["Evidence Checked"].strip():
            lint.fail(f"{shard}: {d['ID']} has an empty 'Evidence Checked' "
                      f"(every rechecked row must record what was inspected)")
        # U8 (c): a blocked/confirmation_needed/deferred verdict must name the
        # blocker in either proposal column (Proposed Register Change / Proposed Note).
        if d["Verdict"] in BLOCKER_VERDICTS:
            blocker = (d["Proposed Register Change"].strip()
                       or d["Proposed Note"].strip())
            if not blocker:
                lint.fail(f"{shard}: {d['ID']} verdict '{d['Verdict']}' but neither "
                          f"'Proposed Register Change' nor 'Proposed Note' records a blocker")
    check_unique(lint, shard, seen, "ledger ID")
    if stream == "claims":
        # U4 advisory: identifier anchoring on rows closing `confirmed`.
        check_anchoring_advisory(lint, audit, rows)
    else:
        _validate_code_adjudication_shard(
            lint, audit, shard, text, rows, assigned, supplementary)
    for i in assigned:
        if i not in seen:
            lint.fail(f"{shard}: assigned ID {i} missing from ledger")
    if ladder >= 2 and evidence_levels == {"static_source_verified"}:
        lint.warn(f"{shard}: ladder level {ladder} but every check is static_source_verified")
    stages = manifest.get("stages", {}) if isinstance(manifest, dict) else {}
    if supplementary or f"{stream}_b6a" in stages:
        entries, _coverage = typed_shard_footer(
            lint, shard, text, stream, recheck=True)
        # Recheck observations are deliberately not register rows.  Their
        # typed footer is reconciled at b6a (main) or b6b (supplementary).
        for entry in entries:
            if entry["Kind"] == "candidate" and entry["_ids"]:
                lint.fail(f"{shard}: recheck candidate {entry['Entry ID']} cannot name register IDs")


def supplementary_detector_mappings(lint, audit):
    """Project detector witnesses onto b6a split descendants assigned at b5s."""
    mappings = parse_detector_mappings(lint, audit)
    summary = audit / "code_error_recheck_summary.md"
    text = read_text(lint, summary) or ""
    lineage_rows = tables_by_cols(
        lint, summary, text, LINEAGE_COLS, "split lineage")
    descendants = {
        (row["Original Error ID"], row["Channel"], row["Source ID"],
         row["Witness ID"]): row["Descendant Error ID"]
        for row in lineage_rows
    }
    projected = []
    for mapping in mappings:
        key = (mapping["Error ID"], mapping["Channel"], mapping["Source ID"],
               mapping["Witness ID"])
        descendant = descendants.get(key)
        if descendant:
            projected.append({**mapping, "Error ID": descendant})
    return projected


def _validate_code_adjudication_shard(lint, audit, shard, text, rows, assigned,
                                       supplementary=False):
    parsed_tables = parse_tables(text)
    ledger_table_count = sum(
        1 for headers, _table_rows, _line in parsed_tables
        if headers == CODE_LEDGER_COLS
    )
    if ledger_table_count != 1:
        lint.fail(f"{shard}: expected exactly one structured recheck ledger table, "
                  f"found {ledger_table_count}")
    for marker in ("Witness outcomes", "Verification records"):
        count = len(re.findall(
            rf"(?m)^###\s+{re.escape(marker)}\s*$", text))
        if count != 1:
            lint.fail(f"{shard}: expected exactly one '### {marker}' marker, found {count}")
    outcome_table_count = sum(
        1 for headers, _table_rows, _line in parsed_tables
        if headers == WITNESS_OUTCOME_COLS
    )
    if outcome_table_count != 1:
        lint.fail(f"{shard}: expected exactly one witness outcomes table, "
                  f"found {outcome_table_count}")

    ledger_rows = [dict(zip(CODE_LEDGER_COLS, row)) for row in rows]
    mappings = (supplementary_detector_mappings(lint, audit)
                if supplementary else parse_detector_mappings(lint, audit))
    mapped_ids = {row["Error ID"] for row in mappings}
    mappings_by_id = {}
    mapping_by_key = {}
    for row in mappings:
        mappings_by_id.setdefault(row["Error ID"], []).append(row)
        mapping_by_key[(row["Channel"], row["Source ID"], row["Witness ID"])] = row
    outcomes = tables_by_cols(
        lint, shard, text, WITNESS_OUTCOME_COLS, "witness outcomes", required=True)
    records = []
    records += tables_by_cols(
        lint, shard, text, MF_VERIFICATION_COLS, "MF verification records")
    records += tables_by_cols(
        lint, shard, text, PROBE_VERIFICATION_COLS, "DU/CV verification records")
    outcome_by_key, record_by_id = {}, {}
    for outcome in outcomes:
        key = tuple(outcome[field] for field in ("Channel", "Source ID", "Witness ID"))
        if key in outcome_by_key:
            lint.fail(f"{shard}: duplicate witness outcome {'/'.join(key)}")
        outcome_by_key[key] = outcome
        mapping = mapping_by_key.get(key)
        if mapping is None:
            lint.fail(f"{shard}: witness outcome {'/'.join(key)} is not mechanically mapped")
            continue
        try:
            mechanism.canonicalize_mechanism(
                outcome["Mech Class"], outcome["Mech Object"],
                outcome["Mech Relation"], outcome["Mech Expected"],
                outcome["Mech Actual"], register="code_errors",
                anchor=mapping["Site Anchor"], projection=mechanism.EMPTY_PROJECTION,
            )
        except mechanism.MechanismSchemaError as exc:
            lint.fail(f"{shard}: witness {'/'.join(key)} mechanism invalid: {exc}")
    for record in records:
        record_id = record["Record ID"]
        if not record_id or blank_cell(record_id):
            lint.fail(f"{shard}: verification record has an empty Record ID")
        elif record_id in record_by_id:
            lint.fail(f"{shard}: duplicate verification Record ID {record_id}")
        else:
            record_by_id[record_id] = record
        channel = record["Channel"]
        if (channel == "MF") != ("File Digest (sha256)" in record):
            lint.fail(f"{shard}: verification record {record_id} uses the wrong channel-typed schema")
        key = tuple(record[field] for field in ("Channel", "Source ID", "Witness ID"))
        if key not in mapping_by_key:
            lint.fail(f"{shard}: verification record {record_id} names unmapped key {'/'.join(key)}")

    for row in ledger_rows:
        eid, verdict = row["ID"], row["Verdict"]
        parse_field_patches(lint, shard, row)
        if eid not in mapped_ids:
            continue
        mapped = mappings_by_id[eid]
        expected_keys = {(m["Channel"], m["Source ID"], m["Witness ID"])
                         for m in mapped}
        expected_witnesses = {m["Witness ID"] for m in mapped}
        proposed_status = row["Proposed Status"]
        proposed_severity = row["Proposed Severity"]
        duplicate_target = row["Duplicate Target"]
        required_outcomes = verdict in {"confirmed_error", "not_error", "duplicate"}
        actual_keys = {key for key in outcome_by_key if key in expected_keys}
        declared_witnesses = set(list_cell(row["Outcome Witness IDs"]))
        if required_outcomes:
            if actual_keys != expected_keys:
                missing = sorted(expected_keys - actual_keys)
                extra = sorted(actual_keys - expected_keys)
                lint.fail(f"{shard}: {eid} witness-outcome coverage mismatch "
                          f"(missing={missing}, extra={extra})")
            if declared_witnesses != expected_witnesses:
                lint.fail(f"{shard}: {eid} Outcome Witness IDs do not exactly cover mapped witnesses")
        else:
            if actual_keys or declared_witnesses:
                lint.fail(f"{shard}: {eid} verdict '{verdict}' must carry no witness outcomes")

        if verdict == "confirmed_error":
            if proposed_status != "confirmed":
                lint.fail(f"{shard}: {eid} confirmed_error requires Proposed Status 'confirmed'")
            if proposed_severity not in {"1", "2", "3", "4"}:
                lint.fail(f"{shard}: {eid} confirmed_error requires Proposed Severity 1-4")
            if row["Accepted Error Type"] not in ERROR_TYPES:
                lint.fail(f"{shard}: {eid} confirmed_error requires a closed Accepted Error Type")
            if blank_cell(row["Accepted Mechanism"]):
                lint.fail(f"{shard}: {eid} confirmed_error requires Accepted Mechanism")
            if not blank_cell(duplicate_target):
                lint.fail(f"{shard}: {eid} confirmed_error forbids Duplicate Target")
        elif verdict == "not_error":
            if proposed_status != "not_error":
                lint.fail(f"{shard}: {eid} not_error requires Proposed Status 'not_error'")
            if not blank_cell(proposed_severity):
                lint.fail(f"{shard}: {eid} not_error forbids Proposed Severity")
            if not blank_cell(duplicate_target):
                lint.fail(f"{shard}: {eid} not_error forbids Duplicate Target")
            record_ids = set(list_cell(row["Verification Record IDs"]))
            covered = set()
            for record_id in record_ids:
                record = record_by_id.get(record_id)
                if record is None:
                    lint.fail(f"{shard}: {eid} names unknown verification Record ID {record_id}")
                    continue
                covered.add(tuple(record[field] for field in
                                  ("Channel", "Source ID", "Witness ID")))
            if covered != expected_keys:
                lint.fail(f"{shard}: {eid} not_error verification records do not cover every mapped key")
        elif verdict == "duplicate":
            target = duplicate_target.strip()
            if target not in mapped_ids:
                lint.fail(f"{shard}: {eid} guarded Duplicate Target {target!r} is not mechanically mapped")
            if proposed_status != f"duplicate_of:{target}":
                lint.fail(f"{shard}: {eid} duplicate Proposed Status must be derived as duplicate_of:{target}")
            if not blank_cell(proposed_severity):
                lint.fail(f"{shard}: {eid} duplicate forbids Proposed Severity")
            if row["Accepted Error Type"] not in ERROR_TYPES or blank_cell(row["Accepted Mechanism"]):
                lint.fail(f"{shard}: {eid} duplicate requires accepted error type and mechanism")
        elif verdict == "confirmation_needed":
            if proposed_status != "confirmation_needed":
                lint.fail(f"{shard}: {eid} confirmation_needed requires matching Proposed Status")
            if proposed_severity not in {"1", "2", "3", "4"}:
                lint.fail(f"{shard}: {eid} confirmation_needed requires Proposed Severity 1-4")
            if not blank_cell(duplicate_target):
                lint.fail(f"{shard}: {eid} confirmation_needed forbids Duplicate Target")
        elif verdict in {"blocked", "deferred"}:
            if proposed_status != "blocked":
                lint.fail(f"{shard}: {eid} {verdict} requires Proposed Status 'blocked'")
            carried = row["Current Severity"]
            if (not blank_cell(carried) and proposed_severity != carried) or (
                    blank_cell(carried) and not blank_cell(proposed_severity)):
                lint.fail(f"{shard}: {eid} {verdict} must carry forward Current Severity")
            if not blank_cell(duplicate_target):
                lint.fail(f"{shard}: {eid} {verdict} forbids Duplicate Target")
            if verdict == "deferred":
                citation = (row["Proposed Register Change"] + " "
                            + row["Proposed Note"]).lower()
                if "off-limit" not in citation:
                    lint.fail(f"{shard}: {eid} deferred requires an off-limits citation")


def register_files(audit, stream):
    if stream == "claims":
        return [("claims_register.md", CLAIMS_COLS, "Claim ID"), ("output_register.md", OUTPUT_COLS, "Output ID")]
    return [("code_error_register.md", ERROR_COLS, "Error ID")]


def _summary_path(audit, stream, supplementary=False):
    name = (SUPPLEMENTARY_SUMMARY_FILES[stream] if supplementary else
            ("claims_recheck_summary.md" if stream == "claims"
             else "code_error_recheck_summary.md"))
    return audit / name


def _summary_footer_report(text):
    lines = []
    pattern = re.compile(
        r"^audit/[^#\s]+#OBS-\d{4}\s+\|\s+"
        r"(?:candidate:(?:[CEO]-\d{4})(?:\s*[;,]\s*[CEO]-\d{4})*|"
        r"late_observation:LO-[CE]-\d{4}|dismissed:\S.*)$"
    )
    for line in (text or "").splitlines():
        if pattern.fullmatch(line.strip()):
            lines.append(line.strip())
    return {"footer_dispositions": lines}


def _manifest_shard_status(manifest, stage_key, raw):
    stages = manifest.get("stages", {}) if isinstance(manifest, dict) else {}
    stage = stages.get(stage_key, {}) if isinstance(stages, dict) else {}
    shards = stage.get("shards", {}) if isinstance(stage, dict) else {}
    wire = normalized_audit_path(raw)
    value = shards.get(wire, shards.get(raw, {})) if isinstance(shards, dict) else {}
    return value.get("status") if isinstance(value, dict) else None


def _recheck_footer_entries(lint, audit, stream, clusters, manifest, stage_key):
    entries = {}
    for cluster in clusters:
        raw = cluster.get("Shard File", "")
        if _manifest_shard_status(manifest, stage_key, raw) == "blocked":
            continue
        wire = normalized_audit_path(raw)
        path = audit_path(audit, wire)
        text = read_text(lint, path)
        if text is None:
            continue
        parsed, _coverage = typed_shard_footer(
            lint, path, text, stream, recheck=True)
        for entry in parsed:
            key = (wire, entry["Entry ID"])
            if key in entries:
                lint.fail(f"{path}: duplicate typed-footer join key {wire}#{entry['Entry ID']}")
            entries[key] = entry
    return entries


def _reconcile_recheck_footers(lint, audit, stream, clusters, manifest,
                               stage_key, summary, allowed_candidates=None,
                               late_rows=None):
    expected = _recheck_footer_entries(
        lint, audit, stream, clusters, manifest, stage_key)
    actual = {}
    for disposition in parse_footer_dispositions(
            lint, summary, _summary_footer_report(read_text(lint, summary) or "")):
        key = disposition["key"]
        if key in actual:
            lint.fail(f"{summary}: duplicate footer disposition for {key[0]}#{key[1]}")
        actual[key] = disposition
    for key in sorted(set(expected) - set(actual)):
        lint.fail(f"{summary}: typed footer entry {key[0]}#{key[1]} is undispositioned")
    for key in sorted(set(actual) - set(expected)):
        lint.fail(f"{summary}: disposition has no typed footer entry {key[0]}#{key[1]}")
    candidate_ids = set()
    late_by_id = {row["LO ID"]: row for row in (late_rows or [])}
    for key in sorted(set(expected) & set(actual)):
        entry, disposition = expected[key], actual[key]
        action = disposition["action"]
        if entry["Kind"] == "candidate" and action == "dismissed":
            # Explicit dismissal is the operator-ratified alternative to minting.
            continue
        if allowed_candidates is not None:
            if action == "late_observation":
                lint.fail(f"{summary}: main-wave footer {key[0]}#{key[1]} cannot become a late observation")
            if action == "candidate":
                candidate_ids.update(disposition["ids"])
                extra = set(disposition["ids"]) - set(allowed_candidates)
                if extra:
                    lint.fail(f"{summary}: footer disposition cites non-discovery IDs {sorted(extra)}")
        else:
            if action == "candidate":
                lint.fail(f"{summary}: b6b cannot mint register rows from {key[0]}#{key[1]}")
            if entry["Kind"] == "candidate" and action not in {"late_observation", "dismissed"}:
                lint.fail(f"{summary}: b5s candidate must become a late observation or dismissal")
            if action == "late_observation":
                row = late_by_id.get(disposition["lo_id"])
                source = f"{key[0]}#{key[1]}"
                if row is None or row.get("Source Shard") != source:
                    lint.fail(f"{summary}: {source} does not join exactly to {disposition['lo_id']}")
    return candidate_ids


def _ranges_from_text(text):
    return [parsed for a, b in RANGE_RE.findall(text or "")
            if (parsed := parse_range(f"{a}–{b}"))]


def _previous_ranges(lint, audit, stream):
    names = (["claims_review_plan.md", "claims_second_read_plan.md"]
             if stream == "claims" else
             ["code_error_review_plan.md", "code_error_second_read_plan.md"])
    ranges = []
    for name in names:
        path = audit / "plans" / name
        text = read_text(lint, path)
        if text is not None:
            ranges.extend(_ranges_from_text(text))
    if stream == "code":
        path = audit / "_run/detector_mapping.md"
        text = read_text(lint, path)
        if text is not None:
            ranges.extend(_ranges_from_text(text))
    return ranges


def _validate_supplementary_plan(lint, audit, stream, inventory, clusters,
                                 new_ids, discovery_ids):
    plan = recheck_plan_path(audit, stream, True)
    text = read_text(lint, plan) or ""
    declared = [parse_range(f"{a}–{b}") for a, b in DISCOVERY_RANGE_RE.findall(text)]
    declared = [item for item in declared if item]
    check_identifier_exhaustion(lint, plan, declared)
    check_disjoint(lint, plan, _previous_ranges(lint, audit, stream) + declared)
    by_letter = {letter: set() for letter in "CEO"}
    for item in discovery_ids:
        by_letter[item[0]].add(item)
    for letter in "CEO":
        spans = [span for span in declared if span[0] == letter]
        capacity = sum(end - start + 1 for _l, start, end in spans)
        if capacity != len(by_letter[letter]):
            lint.fail(f"{plan}: {letter} discovery range capacity {capacity} does not equal accepted count {len(by_letter[letter])}")
        for item in sorted(by_letter[letter]):
            if not in_ranges(item, spans):
                lint.fail(f"{plan}: discovery {item} is outside its declared range")
    inv_ids = [row.get("ID", "") for row in inventory]
    expected_inventory = {item for item in new_ids if item.startswith(("C-", "E-"))}
    if set(inv_ids) != expected_inventory or len(inv_ids) != len(set(inv_ids)):
        lint.fail(f"{plan}: supplementary inventory must exactly equal new C/E rows; expected={sorted(expected_inventory)}, actual={sorted(inv_ids)}")
    assignments = {}
    expected_dir = f"audit/{SUPPLEMENTARY_SHARD_DIRS[stream]}/"
    for cluster in clusters:
        wire = normalized_audit_path(cluster.get("Shard File", ""))
        if not wire.startswith(expected_dir) or not wire.endswith(".md"):
            lint.fail(f"{plan}: supplementary shard path {wire!r} must live under {expected_dir}")
        for item in re.findall(r"[CEO]-\d{4}", cluster.get("Assigned IDs", "")):
            assignments.setdefault(item, []).append(cluster.get("Cluster ID", ""))
    for item in sorted(expected_inventory):
        if len(assignments.get(item, [])) != 1:
            lint.fail(f"{plan}: supplementary inventory ID {item} assigned {len(assignments.get(item, []))} times")
    for item in sorted(set(assignments) - expected_inventory):
        lint.fail(f"{plan}: cluster assigns non-inventory ID {item}")


def _lineage_descendants(lint, summary, text):
    rows = tables_by_cols(lint, summary, text, LINEAGE_COLS, "split lineage")
    return {row["Descendant Error ID"] for row in rows}


def _output_discovery_lifecycle(lint, summary, text, rows, discovery_ids):
    discovered = {item for item in discovery_ids if item.startswith("O-")}
    table = tables_by_cols(
        lint, summary, text, OUTPUT_DISCOVERY_COLS, "output discovery dispositions")
    by_id = {}
    for row in table:
        output_id = row["Output ID"]
        if output_id in by_id:
            lint.fail(f"{summary}: duplicate output discovery disposition {output_id}")
        by_id[output_id] = row
        wanted = OUTPUT_FROM_CLAIM_VERDICT.get(row["Claim Verdict"])
        if wanted is None or row["Output Status"] != wanted:
            lint.fail(f"{summary}: output {output_id} verdict/status mapping is invalid")
    final = {dict(zip(OUTPUT_COLS, row))["Output ID"]: dict(zip(OUTPUT_COLS, row))
             for row in rows}
    for output_id in sorted(discovered):
        status = final.get(output_id, {}).get("Status")
        structural = status == "orphan"
        mapped = output_id in by_id
        if structural == mapped:
            lint.fail(f"{summary}: output discovery {output_id} must take exactly one claim-disposition or orphan branch")
        if mapped:
            record = by_id[output_id]
            if final[output_id].get("Claim IDs") != record["Claim ID"]:
                lint.fail(f"{summary}: output discovery {output_id} claim join disagrees with register")
            if status != record["Output Status"]:
                lint.fail(f"{summary}: output discovery {output_id} status disagrees with mapping")


def _stage_b6_pre_u6_fixture(lint, audit, stream, manifest):
    """Preserve historical direct-fixture coverage below the production seam.

    Initialized runs always contain ``<stream>_b6a`` and therefore never take
    this path.  It exists only so U1–U5 tests keep proving the boundary they
    originally reviewed while their shared builder migrates stage names.
    """
    check_manifest_worker_shards(
        lint, manifest, f"{stream}_b5", recheck_plan_path(audit, stream),
        "Cluster ID")
    snap = audit / "_run/snapshots" / SNAP_KEY[f"b6a-{stream}"]
    staging = audit / "_staging"
    summary = _summary_path(audit, stream)
    stext = read_text(lint, summary)
    splits = 0
    if stext is not None:
        first = re.search(r"Splits declared:\s*(\d+)", stext)
        second = re.search(r"Merges declared:\s*(\d+)", stext)
        if not first or not second:
            lint.fail(f"{summary}: missing machine-readable split/merge counts")
        else:
            splits = int(first.group(1))
    total_new = 0
    if stream == "claims":
        _alloc, coord = parse_plan(
            lint, audit / "plans/claims_review_plan.md", "Worker ID")
    else:
        _alloc, coord = parse_plan(
            lint, audit / "plans/code_error_review_plan.md", "Chunk ID")
    for fname, cols, id_col in register_files(audit, stream):
        final_path = staging / fname
        final = load_register(lint, final_path, cols)
        before = load_register(lint, snap / fname, cols)
        if final is None or before is None:
            continue
        final_ids = col(final[1], cols, id_col)
        before_ids = set(col(before[1], cols, id_col))
        check_unique(lint, final_path, final_ids, id_col)
        for item in sorted(before_ids - set(final_ids)):
            lint.fail(f"{final_path}: row {item} deleted at recheck merge")
        new = set(final_ids) - before_ids
        for item in sorted(new):
            if not coord or not in_ranges(item, coord):
                lint.fail(f"{final_path}: new ID {item} outside the merge-coordinator range")
        total_new += len(new)
        if fname == "claims_register.md":
            check_claims_rows(lint, final_path, final[1], final=True)
            check_adjudication_advisory(lint, final_path, final[1])
        elif fname == "output_register.md":
            check_output_rows(lint, final_path, final[1], final=True)
        else:
            check_error_rows(lint, final_path, final[1], final=True)
            for status in col(final[1], cols, "Status"):
                if status == "candidate":
                    lint.fail(f"{final_path}: 'candidate' status must not survive the recheck merge")
    if total_new != splits:
        lint.fail(f"recheck merge: {total_new} new row(s) but 'Splits declared: {splits}'")
    if stream == "code":
        check_detector_mapping_b6(
            lint, audit, audit / "_staging/code_error_register.md")


def stage_b6(lint, audit, stream, manifest):
    """Compatibility entry point for historical in-process unit tests only."""
    return _stage_b6_pre_u6_fixture(lint, audit, stream, manifest)


def stage_b6a(lint, audit, stream, manifest):
    stages = manifest.get("stages", {}) if isinstance(manifest, dict) else {}
    if f"{stream}_b6a" not in stages:
        # Design call 8: the b6 -> b6a rename is a rename, not an alias.  A
        # manifest without the b6a key predates the U6 supplementary contract
        # and is a restart; the pre-U6 fixture path stays reachable only via
        # the in-process helper for the historical U1-U5 test fixtures.
        lint.fail(
            f"manifest has no {stream}_b6a stage entry: this run predates the "
            "b6a/b5s/b6b supplementary contract and must restart (design call 8)"
        )
        return
    main_plan = recheck_plan_path(audit, stream)
    check_manifest_worker_shards(lint, manifest, f"{stream}_b5", main_plan, "Cluster ID")
    parsed_main = parse_recheck_plan(lint, audit, stream)
    parsed_supp = parse_recheck_plan(lint, audit, stream, True)
    if parsed_main is None or parsed_supp is None:
        return
    _main_path, _main_inventory, main_clusters = parsed_main
    _supp_path, supp_inventory, supp_clusters = parsed_supp
    snap = audit / "_run/snapshots" / SNAP_KEY[f"b6a-{stream}"]
    summary = _summary_path(audit, stream)
    text = read_text(lint, summary) or ""
    counts_match = DISCOVERY_COUNTS_RE.search(text)
    if not counts_match:
        lint.fail(f"{summary}: missing exact 'Discoveries declared: C=<n>; O=<n>; E=<n>' line")
        declared_counts = {letter: -1 for letter in "COE"}
    else:
        declared_counts = dict(zip("COE", map(int, counts_match.groups())))
    if not re.search(r"(?m)^Splits declared:\s*\d+\s*$", text) or not re.search(
            r"(?m)^Merges declared:\s*\d+\s*$", text):
        lint.fail(f"{summary}: missing machine-readable split/merge counts")
    descendants = _lineage_descendants(lint, summary, text) if stream == "code" else set()
    new_ids, new_rows = set(), {}
    final_rows_by_file = {}
    for fname, cols, id_col in register_files(audit, stream):
        frozen_post_b6a = audit / "_run/snapshots" / f"{stream}_b6b" / fname
        final_path = frozen_post_b6a if frozen_post_b6a.is_file() else audit / fname
        final = load_register(lint, final_path, cols)
        before = load_register(lint, snap / fname, cols)
        if final is None or before is None:
            continue
        final_rows, before_rows = final[1], before[1]
        final_rows_by_file[fname] = final_rows
        final_by_id = {dict(zip(cols, row))[id_col]: dict(zip(cols, row)) for row in final_rows}
        before_ids = set(col(before_rows, cols, id_col))
        final_ids = set(final_by_id)
        for item in sorted(before_ids - final_ids):
            lint.fail(f"{final_path}: row {item} deleted at b6a (rows are never deleted)")
        current_new = final_ids - before_ids
        new_ids.update(current_new)
        new_rows.update({item: final_by_id[item] for item in current_new})
        check_unique(lint, final_path, list(final_by_id), id_col)
        if fname == "claims_register.md":
            check_claims_rows(lint, final_path, final_rows, final=True)
            for item in current_new:
                if final_by_id[item]["Status"] not in {"inconsistent", "unclear"}:
                    lint.fail(f"{final_path}: discovery {item} must be inconsistent or unclear")
        elif fname == "output_register.md":
            check_output_rows(lint, final_path, final_rows, final=True)
            for item in current_new:
                if final_by_id[item]["Status"] not in {"mapped", "orphan", "unclear", "inconsistent"}:
                    lint.fail(f"{final_path}: discovery {item} has illegal output status")
        else:
            check_error_rows(lint, final_path, final_rows, final=True)
            for item, row in final_by_id.items():
                if row["Status"] == "candidate" and item not in current_new:
                    lint.fail(f"{final_path}: only a b6a discovery may remain candidate")
            for item in current_new:
                if final_by_id[item]["Status"] != "candidate":
                    lint.fail(f"{final_path}: code discovery {item} must be candidate")
    discovery_ids = new_ids - descendants
    actual_counts = {letter: sum(item.startswith(letter + "-") for item in discovery_ids)
                     for letter in "COE"}
    if declared_counts != actual_counts:
        lint.fail(f"{summary}: discovery counts {declared_counts} disagree with rows {actual_counts}")
    _validate_supplementary_plan(
        lint, audit, stream, supp_inventory, supp_clusters, new_ids, discovery_ids)
    footer_ids = _reconcile_recheck_footers(
        lint, audit, stream, main_clusters, manifest, f"{stream}_b5", summary,
        allowed_candidates=discovery_ids)
    if footer_ids != discovery_ids:
        lint.fail(f"{summary}: footer-minted discovery IDs do not exactly equal declared discoveries; expected={sorted(discovery_ids)}, actual={sorted(footer_ids)}")
    if stream == "claims":
        output_rows = final_rows_by_file.get("output_register.md")
        if output_rows is not None:
            _output_discovery_lifecycle(
                lint, summary, text, output_rows, discovery_ids)
    else:
        check_detector_mapping_b6(lint, audit)


def _all_recheck_ledgers(lint, audit, stream, supplementary):
    root = audit / (SUPPLEMENTARY_SHARD_DIRS[stream] if supplementary else
                    ("_recheck" if stream == "claims" else "_code_error_recheck"))
    cols = LEDGER_COLS if stream == "claims" else CODE_LEDGER_COLS
    found = []
    if not root.is_dir():
        return found
    for path in sorted(root.rglob("*.md")):
        text = read_text(lint, path) or ""
        for headers, rows, _line in parse_tables(text):
            if headers == cols:
                found.extend((path, dict(zip(cols, row))) for row in rows
                             if len(row) == len(cols))
    return found


def _claim_disposition_allowed(row, final):
    verdict = row.get("Verdict")
    status, severity = final.get("Status"), final.get("Severity")
    if verdict in {"substantiated", "substantiated_but_reframe"}:
        return status == "inconsistent" and severity in {"1", "2", "3", "4"}
    if verdict == "row_note_only":
        return status in {"confirmed", "mapped"} and severity == "1"
    if verdict == "not_substantiated":
        return status in {"confirmed", "mapped", "unclear"} and blank_cell(severity)
    if verdict == "confirmation_needed":
        return status == "confirmation_needed"
    if verdict == "blocked":
        return status == "blocked"
    return False


def _late_observation_rows(lint, audit, stream, path=None):
    path = path or audit / f"late_observations_{stream}.md"
    text = read_text(lint, path)
    if text is None:
        return path, [], []
    tables = [(headers, rows) for headers, rows, _line in parse_tables(text)]
    matches = [rows for headers, rows in tables if headers == LO_COLS]
    if len(matches) > 1:
        lint.fail(f"{path}: expected at most one late-observations table")
    rows = []
    if matches:
        for raw in matches[0]:
            if len(raw) != len(LO_COLS):
                lint.fail(f"{path}: malformed late-observation row")
                continue
            rows.append(dict(zip(LO_COLS, raw)))
    elif LO_EMPTY not in text:
        lint.fail(f"{path}: requires a table or exact '{LO_EMPTY}'")
    disp_matches = [rows for headers, rows in tables if headers == LO_DISPOSITION_COLS]
    dispositions = []
    if disp_matches:
        if len(disp_matches) != 1:
            lint.fail(f"{path}: expected exactly one dispositions table")
        for raw in disp_matches[0]:
            if len(raw) == len(LO_DISPOSITION_COLS):
                dispositions.append(dict(zip(LO_DISPOSITION_COLS, raw)))
    elif rows and LO_DISPOSITIONS_EMPTY not in text:
        lint.fail(f"{path}: late observations require a Dispositions table")
    prefix = "LO-C-" if stream == "claims" else "LO-E-"
    for index, row in enumerate(rows, start=1):
        if row["LO ID"] != f"{prefix}{index:04d}":
            lint.fail(f"{path}: late-observation IDs must be sequential from {prefix}0001")
        if not re.fullmatch(r"audit/[^#]+#OBS-\d{4}", row["Source Shard"]):
            lint.fail(f"{path}: {row['LO ID']} Source Shard must be path#OBS-####")
        if blank_cell(row["Anchor"]) or blank_cell(row["Observation"]):
            lint.fail(f"{path}: {row['LO ID']} requires Anchor and Observation")
    if {row["LO ID"] for row in dispositions} != {row["LO ID"] for row in rows}:
        lint.fail(f"{path}: dispositions must cover every late observation exactly once")
    check_unique(lint, path, [row["LO ID"] for row in dispositions], "LO disposition")
    for row in dispositions:
        if not valid_lo_state(row["Prior State"]):
            lint.fail(f"{path}: {row['LO ID']} invalid prior state {row['Prior State']!r}")
        if not valid_lo_state(row["State"]):
            lint.fail(f"{path}: {row['LO ID']} invalid disposition state {row['State']!r}")
        elif not valid_lo_transition(row["Prior State"], row["State"]):
            lint.fail(
                f"{path}: illegal recorded disposition transition for {row['LO ID']}: "
                f"{row['Prior State']!r} -> {row['State']!r}")
    return path, rows, dispositions


def valid_lo_state(state):
    return bool(
        state == "pending" or state == "acknowledged_unverified"
        or re.fullmatch(r"qa_commissioned:QA-\d{4}", state)
        or re.fullmatch(r"qa_closed:QA-\d{4}:(?:conclusive|inconclusive)", state)
        or re.fullmatch(r"minted:BC-\d{4}", state)
    )


def valid_lo_transition(before, after):
    if before == after:
        return True
    if before == "pending":
        return bool(after == "acknowledged_unverified"
                    or re.fullmatch(r"qa_commissioned:QA-\d{4}", after)
                    or re.fullmatch(r"minted:BC-\d{4}", after))
    commissioned = re.fullmatch(r"qa_commissioned:(QA-\d{4})", before)
    if commissioned:
        return after in {
            f"qa_closed:{commissioned.group(1)}:conclusive",
            f"qa_closed:{commissioned.group(1)}:inconclusive",
        }
    if before == "acknowledged_unverified" or re.fullmatch(
            r"qa_closed:QA-\d{4}:conclusive", before):
        return bool(re.fullmatch(r"minted:BC-\d{4}", after))
    if re.fullmatch(r"qa_closed:QA-\d{4}:inconclusive", before):
        return bool(re.fullmatch(r"qa_commissioned:QA-\d{4}", after)
                    or re.fullmatch(r"minted:BC-\d{4}", after))
    return False


def _check_detector_materiality(lint, audit):
    mappings = parse_detector_mappings(lint, audit)
    if not mappings:
        return
    ledgers = (_all_recheck_ledgers(lint, audit, "code", False)
               + _all_recheck_ledgers(lint, audit, "code", True))
    by_id = {}
    for path, row in ledgers:
        by_id.setdefault(row.get("ID", ""), []).append((path, row))
    final = load_register(
        lint, audit / "code_error_register.md", ERROR_COLS, allow_extra=True)
    final_by_id = ({dict(zip(final[0], row))["Error ID"]: dict(zip(final[0], row))
                    for row in final[1]} if final else {})
    summary = audit / "code_error_recheck_summary.md"
    summary_text = read_text(lint, summary) or ""
    lineage_rows = tables_by_cols(
        lint, summary, summary_text, LINEAGE_COLS, "split lineage")
    descendants = {}
    for row in lineage_rows:
        descendants.setdefault(row["Original Error ID"], set()).add(
            row["Descendant Error ID"])
    detector_ids = set()
    for original in {row["Error ID"] for row in mappings}:
        detector_ids.update(descendants.get(original, {original}))
    for error_id in sorted(detector_ids):
        final_row = final_by_id.get(error_id, {})
        final_status = final_row.get("Status", "")
        if final_status == "not_error" or final_status.startswith("duplicate_of:"):
            continue
        dispositions = by_id.get(error_id, [])
        if len(dispositions) != 1:
            continue
        path, ledger = dispositions[0]
        proposed = ledger.get("Proposed Severity", "")
        if ledger.get("Verdict") in {"confirmation_needed", "blocked", "deferred"}:
            match = re.search(
                r"\[materiality_reassessment\]\s+severity=([1-4]);\s+basis=(\S.*)",
                ledger.get("Proposed Note", ""))
            if not match:
                lint.fail(f"{path}: detector-minted {error_id} carry-forward verdict requires '[materiality_reassessment] severity=<1-4>; basis=<text>' in Proposed Note")
            elif ledger.get("Verdict") != "confirmation_needed" \
                    and not blank_cell(proposed) and proposed != match.group(1):
                lint.fail(f"{path}: detector-minted {error_id} materiality severity disagrees with Proposed Severity")
            elif final_row.get("Severity", "") != match.group(1):
                lint.fail(f"{path}: detector-minted {error_id} materiality severity was not applied at b6b")
        elif proposed not in {"1", "2", "3", "4"}:
            lint.fail(f"{path}: active detector-minted {error_id} lacks an explicit severity ruling")


def _reject_mixed_cells(lint, paths):
    """Refuse a persisted merge aggregate without banning prose about MIXED."""
    for path in paths:
        text = read_text(lint, path)
        if text is None:
            continue
        for headers, rows, _line in parse_tables(text):
            for row in rows:
                if any(cell.strip() == mechanism.MIXED for cell in row):
                    lint.fail(f"{path}: MIXED cannot survive b6b")
                    break


def _check_duplicate_chains(lint, path, rows, cols, id_col):
    """Require every duplicate to name one present, terminal canonical row."""
    by_id = {dict(zip(cols, row))[id_col]: dict(zip(cols, row)) for row in rows}
    for row_id, row in by_id.items():
        match = re.fullmatch(r"duplicate_of:([CEO]-\d{4})", row.get("Status", ""))
        if not match:
            continue
        target = match.group(1)
        target_row = by_id.get(target)
        if target == row_id or target_row is None:
            lint.fail(f"{path}: duplicate {row_id} names absent/self target {target}")
        elif str(target_row.get("Status", "")).startswith("duplicate_of:"):
            lint.fail(f"{path}: duplicate chain {row_id} -> {target} is not resolved to a terminal row")


def stage_b6b(lint, audit, stream, manifest):
    parsed = parse_recheck_plan(lint, audit, stream, True)
    if parsed is None:
        return
    plan, inventory, clusters = parsed
    check_manifest_worker_shards(
        lint, manifest, f"{stream}_b5s", plan, "Cluster ID", clusters)
    snap = audit / "_run/snapshots" / SNAP_KEY[f"b6b-{stream}"]
    final_by_all = {}
    final_register_paths = []
    for fname, cols, id_col in register_files(audit, stream):
        # b7/b8 (and later bC) may legally evolve canon after b6b.  Prefer the
        # first immutable pre-stage image that preserves the post-b6b ID set.
        later_snapshots = ("bC", "b7", "b8")
        final_path = next((
            audit / "_run/snapshots" / stage / fname
            for stage in later_snapshots
            if (audit / "_run/snapshots" / stage / fname).is_file()
        ), audit / fname)
        final = load_register(lint, final_path, cols, allow_extra=True)
        before = load_register(lint, snap / fname, cols)
        if final is None or before is None:
            continue
        final_register_paths.append(final_path)
        final_rows = [
            [dict(zip(final[0], row)).get(column, "") for column in cols]
            for row in final[1]
        ]
        final_ids = set(col(final_rows, cols, id_col))
        before_ids = set(col(before[1], cols, id_col))
        if final_ids != before_ids:
            lint.fail(f"{final_path}: b6b cannot mint or delete register rows")
        final_by_all.update({dict(zip(cols, row))[id_col]: dict(zip(cols, row))
                             for row in final_rows})
        if fname == "claims_register.md":
            check_claims_rows(lint, final_path, final_rows, final=True)
        elif fname == "output_register.md":
            check_output_rows(lint, final_path, final_rows, final=True)
        else:
            check_error_rows(lint, final_path, final_rows, final=True)
            for row in final_rows:
                if dict(zip(cols, row))["Status"] == "candidate":
                    lint.fail(f"{final_path}: candidate status cannot survive b6b")
        # Every register: a duplicate_of must name a present, terminal row
        # (the 241ce5e recheck-merge guarantee, restored at b6b).
        _check_duplicate_chains(lint, final_path, final_rows, cols, id_col)
    ledgers = _all_recheck_ledgers(lint, audit, stream, True)
    by_id = {}
    for path, row in ledgers:
        by_id.setdefault(row.get("ID", ""), []).append((path, row))
    inventory_ids = {row.get("ID", "") for row in inventory}
    for item in sorted(inventory_ids):
        dispositions = by_id.get(item, [])
        if len(dispositions) != 1:
            lint.fail(f"{plan}: supplementary ID {item} has {len(dispositions)} ledger rows; expected exactly one")
            continue
        path, ledger = dispositions[0]
        final = final_by_all.get(item)
        if final is None:
            lint.fail(f"{plan}: supplementary ID {item} vanished before b6b")
            continue
        if stream == "code":
            expected = expected_code_disposition(ledger)
            if expected is None or (final.get("Status"), final.get("Severity")) != expected:
                lint.fail(f"{path}: supplementary ID {item} final disposition disagrees with its ledger")
        elif not _claim_disposition_allowed(ledger, final):
            lint.fail(f"{path}: supplementary claim {item} final disposition disagrees with its ledger")
    summary = _summary_path(audit, stream, True)
    read_text(lint, summary)
    lo_path, lo_rows, _dispositions = _late_observation_rows(lint, audit, stream)
    _reconcile_recheck_footers(
        lint, audit, stream, clusters, manifest, f"{stream}_b5s", summary,
        late_rows=lo_rows)
    footer_report = _summary_footer_report(read_text(lint, summary) or "")
    cited_late_ids = {
        disposition["lo_id"]
        for disposition in parse_footer_dispositions(lint, summary, footer_report)
        if disposition["action"] == "late_observation"
    }
    artifact_late_ids = {row["LO ID"] for row in lo_rows}
    if cited_late_ids != artifact_late_ids:
        lint.fail(
            f"{lo_path}: late-observation rows do not exactly equal b5s footer dispositions; "
            f"expected={sorted(cited_late_ids)}, actual={sorted(artifact_late_ids)}")
    sources = [row["Source Shard"] for row in lo_rows]
    check_unique(lint, lo_path, sources, "late-observation Source Shard")
    if stream == "code":
        _check_detector_materiality(lint, audit)
        receipt_path = audit / "_run/code_b6b/dismissal_receipts.md"
        witness_path = audit / "_run/code_b6b/witness_outcomes.md"
        zero_surfaces = {
            receipt_path: (
                "# Supplementary dismissal receipts\n\n"
                "No supplementary dismissal receipts were required.\n"),
            witness_path: (
                "# Supplementary witness outcomes\n\n"
                "No supplementary mapped witness outcomes.\n"),
        }
        for evidence, expected in zero_surfaces.items():
            text = read_text(lint, evidence)
            zero_line = expected.splitlines()[-1]
            if text is not None and zero_line in text and text != expected:
                lint.fail(f"{evidence}: supplementary evidence must use the exact explicit-zero form")
        receipt_text = read_text(lint, receipt_path) or ""
        if "No supplementary dismissal receipts were required." not in receipt_text:
            tables_by_cols(
                lint, receipt_path, receipt_text, RECEIPT_COLS,
                "supplementary dismissal receipts", required=True)
        witness_text = read_text(lint, witness_path) or ""
        if "No supplementary mapped witness outcomes." not in witness_text:
            tables_by_cols(
                lint, witness_path, witness_text, POST_WITNESS_COLS,
                "supplementary witness outcomes", required=True)
        _reject_mixed_cells(lint, [
            *(path for path, _row in ledgers),
            summary,
            lo_path,
            *zero_surfaces,
            *final_register_paths,
        ])


def bc_old_value_hash(register, row_id, field, old_value):
    """Hash the documented bC old-value preimage."""
    preimage = f"{register}\0{row_id}\0{field}\0{old_value}".encode("utf-8")
    return hashlib.sha256(preimage).hexdigest()


def _bc_plan(lint, audit):
    path = audit / "plans/late_observation_corrections.md"
    text = read_text(lint, path)
    if text is None:
        return path, [], []
    matches = [(rows, line) for headers, rows, line in parse_tables(text)
               if headers == BC_PLAN_COLS]
    if len(matches) != 1:
        lint.fail(f"{path}: expected exactly one bC plan table with columns {' | '.join(BC_PLAN_COLS)}")
        rows = []
    else:
        rows = []
        for index, raw in enumerate(matches[0][0], start=1):
            if len(raw) != len(BC_PLAN_COLS):
                lint.fail(f"{path}: malformed plan row {index}")
                continue
            row = dict(zip(BC_PLAN_COLS, raw))
            if not re.fullmatch(r"BC-\d{4}", row["BC ID"]):
                lint.fail(f"{path}: invalid BC ID {row['BC ID']!r}")
            if not re.fullmatch(r"LO-[CE]-\d{4}", row["LO ID"]):
                lint.fail(f"{path}: invalid LO ID {row['LO ID']!r}")
            if row["Register"] not in {"claims", "output", "code_error"}:
                lint.fail(f"{path}: invalid Register {row['Register']!r}")
            if row["Operation"] not in {"new_row", "patch"}:
                lint.fail(f"{path}: invalid Operation {row['Operation']!r}")
            try:
                payload = json.loads(row["Payload JSON"])
            except json.JSONDecodeError as exc:
                lint.fail(f"{path}: {row['BC ID']} Payload JSON is invalid: {exc}")
                payload = None
            if not isinstance(payload, dict):
                lint.fail(f"{path}: {row['BC ID']} Payload JSON must be an object")
            row["_payload"] = payload or {}
            rows.append(row)
    ranges = [parse_range(f"{a}–{b}") for a, b in BC_RANGE_RE.findall(text)]
    return path, rows, [item for item in ranges if item]


def _check_bc_declared_cells(lint, final_path, headers, before_by_id,
                             final_by_id, expected_patches):
    for row_id in set(before_by_id) & set(final_by_id):
        for field in headers:
            before_value = before_by_id[row_id][field]
            final_value = final_by_id[row_id][field]
            expected = expected_patches.get((row_id, field), before_value)
            if final_value != expected:
                lint.fail(f"{final_path}: undeclared bC cell change at {row_id}.{field}")


def _bc_new_row_matches(payload, final_row, headers, rewrite_pairs):
    """Compare a bC payload before or after the authorized b8 rewrite pass."""
    if set(payload) != set(headers):
        return False
    by_base = dict(rewrite_pairs)
    by_original = {original: base for base, original in rewrite_pairs}
    for field in headers:
        if field in by_base:
            original = by_base[field]
            observed = final_row.get(original, "")
            if blank_cell(observed):
                observed = final_row.get(field, "")
            if payload.get(field) != observed:
                return False
        elif field in by_original and blank_cell(payload.get(field, "")):
            if final_row.get(field, "") not in {payload.get(field, ""),
                                                 payload.get(by_original[field], "")}:
                return False
        elif payload.get(field) != final_row.get(field):
            return False
    return True


def stage_bC(lint, audit, manifest):
    plan_path, plan_rows, ranges = _bc_plan(lint, audit)
    check_identifier_exhaustion(lint, plan_path, ranges)
    mode = manifest.get("mode") if isinstance(manifest, dict) else None
    streams = ("code",) if mode == "code_errors_only" else ("claims", "code")
    prior = []
    for stream in streams:
        prior.extend(_previous_ranges(lint, audit, stream))
        supplementary = recheck_plan_path(audit, stream, True)
        if supplementary.is_file():
            prior.extend(_ranges_from_text(supplementary.read_text(encoding="utf-8")))
    check_disjoint(lint, plan_path, prior + ranges)
    # The artifacts are the immutable provenance home; register schemas remain frozen.
    lo_states = {}
    for stream in streams:
        path = audit / f"late_observations_{stream}.md"
        if not path.is_file():
            continue
        _path, rows, dispositions = _late_observation_rows(lint, audit, stream)
        snapshot_path = audit / "_run/snapshots/bC" / path.name
        _snapshot_path, old_rows, old_dispositions = _late_observation_rows(
            lint, audit, stream, snapshot_path)
        if old_rows != rows:
            lint.fail(f"{path}: bC cannot add, delete, or edit late-observation evidence")
        old_states = {row["LO ID"]: row["State"] for row in old_dispositions}
        for disposition in dispositions:
            old_state = old_states.get(disposition["LO ID"])
            if old_state is None or disposition["Prior State"] != old_state \
                    or not valid_lo_transition(old_state, disposition["State"]):
                lint.fail(
                    f"{path}: illegal disposition transition for {disposition['LO ID']}: "
                    f"{old_state!r} -> {disposition['State']!r}")
        known = {row["LO ID"] for row in rows}
        for disposition in dispositions:
            lo_states[disposition["LO ID"]] = disposition["State"]
        for row in plan_rows:
            if row["LO ID"].startswith("LO-C-") == (stream == "claims") \
                    and row["LO ID"] not in known:
                lint.fail(f"{plan_path}: plan names unknown late observation {row['LO ID']}")
    bc_to_lo = {}
    operation_keys = set()
    for row in plan_rows:
        previous = bc_to_lo.setdefault(row["BC ID"], row["LO ID"])
        if previous != row["LO ID"]:
            lint.fail(f"{plan_path}: {row['BC ID']} is reused for more than one LO ID")
        if lo_states.get(row["LO ID"]) != f"minted:{row['BC ID']}":
            lint.fail(f"{plan_path}: {row['LO ID']} disposition must be minted:{row['BC ID']}")
        key = (row["Register"], row["Operation"], row["Row ID"],
               row["_payload"].get("field", ""))
        if key in operation_keys:
            lint.fail(f"{plan_path}: duplicate correction operation {key}")
        operation_keys.add(key)
    register_contracts = {
        "claims": ("claims_register.md", CLAIMS_COLS, "Claim ID"),
        "output": ("output_register.md", OUTPUT_COLS, "Output ID"),
        "code_error": ("code_error_register.md", ERROR_COLS, "Error ID"),
    }
    if mode == "code_errors_only" and any(
            row["Register"] in {"claims", "output"} for row in plan_rows):
        lint.fail(f"{plan_path}: code-errors-only bC cannot edit claims/output registers")
    new_ids = set()
    touched_by_bc = {}
    for register, (filename, base_cols, id_col) in register_contracts.items():
        if mode == "code_errors_only" and register != "code_error":
            continue
        before_path = audit / "_run/snapshots/bC" / filename
        final_path = audit / filename
        before = load_register(lint, before_path, base_cols, allow_extra=True)
        final = load_register(lint, final_path, base_cols, allow_extra=True)
        if before is None or final is None:
            continue
        before_headers, before_rows = before
        final_headers, final_rows = final
        if before_headers != final_headers:
            lint.fail(f"{final_path}: bC cannot change register columns")
            continue
        before_by_id = {dict(zip(before_headers, row))[id_col]: dict(zip(before_headers, row))
                        for row in before_rows}
        final_by_id = {dict(zip(final_headers, row))[id_col]: dict(zip(final_headers, row))
                       for row in final_rows}
        if set(before_by_id) - set(final_by_id):
            lint.fail(f"{final_path}: bC cannot delete rows")
        register_plans = [row for row in plan_rows if row["Register"] == register]
        rewrite_pairs = REWRITE_PAIRS.get(filename, [])
        new_plans = {row["Row ID"]: row for row in register_plans
                     if row["Operation"] == "new_row"}
        actual_new = set(final_by_id) - set(before_by_id)
        if actual_new != set(new_plans):
            lint.fail(f"{final_path}: new-row delta disagrees with bC plan; expected={sorted(new_plans)}, actual={sorted(actual_new)}")
        new_ids.update(actual_new)
        for row_id in actual_new:
            row = new_plans[row_id]
            if row["Old Value SHA256"] not in {"", "-", "—"}:
                lint.fail(f"{plan_path}: new row {row_id} forbids an old-value hash")
            if not _bc_new_row_matches(
                    row["_payload"], final_by_id[row_id], final_headers,
                    rewrite_pairs):
                lint.fail(f"{plan_path}: new row {row_id} payload does not exactly equal the final register row")
            if row_id != row["_payload"].get(id_col):
                lint.fail(f"{plan_path}: new row {row_id} payload ID disagrees")
            touched_by_bc.setdefault(row["BC ID"], set()).add(register)
        expected_patches = {}
        for row in register_plans:
            if row["Operation"] != "patch":
                continue
            payload = row["_payload"]
            field = payload.get("field")
            if (register, field) not in {("claims", "Output IDs"), ("output", "Claim IDs")}:
                lint.fail(f"{plan_path}: patch {row['BC ID']} may change only reciprocal C↔O link columns")
                continue
            if set(payload) != {"field", "new_value"}:
                lint.fail(f"{plan_path}: patch {row['BC ID']} payload must contain only field/new_value")
                continue
            if row["Row ID"] not in before_by_id:
                lint.fail(f"{plan_path}: patch target {row['Row ID']} is absent from the bC snapshot")
                continue
            old = before_by_id[row["Row ID"]].get(field, "")
            wanted_hash = bc_old_value_hash(register, row["Row ID"], field, old)
            if row["Old Value SHA256"] != wanted_hash:
                lint.fail(f"{plan_path}: patch {row['BC ID']} old-value hash disagrees with the bC snapshot")
            expected_patches[(row["Row ID"], field)] = payload["new_value"]
            touched_by_bc.setdefault(row["BC ID"], set()).add(register)
        _check_bc_declared_cells(
            lint, final_path, final_headers, before_by_id, final_by_id,
            expected_patches)
        projected_final_rows = [
            [dict(zip(final_headers, row)).get(column, "") for column in base_cols]
            for row in final_rows
        ]
        if register == "claims":
            check_claims_rows(lint, final_path, projected_final_rows, final=True)
        elif register == "output":
            check_output_rows(lint, final_path, projected_final_rows, final=True)
        else:
            check_error_rows(lint, final_path, projected_final_rows, final=True)
    by_letter = {letter: {item for item in new_ids if item.startswith(letter + "-")}
                 for letter in "CEO"}
    for letter in "CEO":
        spans = [span for span in ranges if span[0] == letter]
        capacity = sum(end - start + 1 for _l, start, end in spans)
        if capacity != len(by_letter[letter]):
            lint.fail(f"{plan_path}: bC {letter} range capacity {capacity} does not equal new-row count {len(by_letter[letter])}")
        for row_id in by_letter[letter]:
            if not in_ranges(row_id, spans):
                lint.fail(f"{plan_path}: bC row {row_id} falls outside its declared range")
    for bc_id, registers in touched_by_bc.items():
        if "output" in registers and "claims" not in registers:
            lint.fail(f"{plan_path}: output correction {bc_id} requires a companion claims edit")
    if mode != "code_errors_only":
        claims = load_register(lint, audit / "claims_register.md", CLAIMS_COLS, allow_extra=True)
        outputs = load_register(lint, audit / "output_register.md", OUTPUT_COLS, allow_extra=True)
        if claims and outputs:
            check_bidirectional(
                lint, claims[1], claims[0], "Claim ID", "Output IDs",
                outputs[1], outputs[0], "Output ID", "Claim IDs", "bC C<->O")


def non_link_identical(lint, st_rows, st_cols, sn_rows, base_cols, idc, link_col, rewrite_pairs, label):
    """Every non-link column must still match the b7 snapshot. Replay-aware: after b8 has run,
    staging carries the `*Original` columns and its base rewrite-pair columns are rewritten in
    place — so when a rewrite-pair's `*Original` column is present, compare the snapshot's base
    value against that `*Original` (the frozen text) rather than the rewritten base column. Every
    other non-link column compares directly. Staging rows are read under their own header order
    (`st_cols`) so extra `*Original` columns never misalign the values."""
    st = {dict(zip(st_cols, r))[idc]: dict(zip(st_cols, r)) for r in st_rows}
    sn = {dict(zip(base_cols, r))[idc]: dict(zip(base_cols, r)) for r in sn_rows}
    if set(st) != set(sn):
        lint.fail(f"{label}: ID sets differ between staging and snapshot")
        return
    rewrite_base = {base: orig for base, orig in rewrite_pairs}
    for i, row in st.items():
        for c in base_cols:
            if c == link_col:
                continue
            orig_col = rewrite_base.get(c)
            if orig_col and orig_col in st_cols:
                if row.get(orig_col, "") != sn[i][c]:
                    lint.fail(f"{label}: {i} rewrite-pair '{c}' — '{orig_col}' does not match the b7 snapshot")
            elif row.get(c, "") != sn[i][c]:
                lint.fail(f"{label}: {i} non-link column '{c}' changed at cross-link")


def confirmed_conflict_links(claim_rows, error_rows, claim_cols=None, error_cols=None):
    """Links pairing a confirmed claim with a confirmed code error (status conflicts)."""
    ccols = claim_cols or CLAIMS_COLS
    ecols = error_cols or ERROR_COLS
    err_status = {
        d["Error ID"]: d.get("Status")
        for d in (dict(zip(ecols, r)) for r in error_rows if not is_example_row(r))
        if d.get("Error ID")
    }
    pairs = []
    for r in claim_rows:
        if is_example_row(r):
            continue
        d = dict(zip(ccols, r))
        if d.get("Status") != "confirmed" or not d.get("Claim ID"):
            continue
        for eid in ids_in(d.get("Related Error IDs", ""), "E"):
            if err_status.get(eid) == "confirmed":
                pairs.append((d["Claim ID"], eid))
    return pairs


# SC-01 overlap-conflict advisory (plan 2026-07-07-001, U5). Code-line citation
# forms observed in real registers: the documented colon form (`path:21-23`,
# `path:11,16-20`) plus the space-L form (`path L21-23`) seen as worker drift.
# RANGE_RE above matches identifier ranges (C-0001–C-0005), not code lines.
CODE_LOC_RE = re.compile(
    r"([A-Za-z0-9_][A-Za-z0-9_./-]*\.[A-Za-z0-9_]+)"  # repo-relative path with extension
    r"(?::(?=\d)|\s+L\s*)"  # colon form (digit must follow — 'path: 3 vars' is prose) or space-L
    r"(\d+(?:\s*[–—-]\s*\d+)?(?:\s*,\s*\d+(?:\s*[–—-]\s*\d+)?)*)"
)
_SPAN_RE = re.compile(r"(\d+)(?:\s*[–—-]\s*(\d+))?")


def code_line_ranges(cell):
    """Extract ``(path, (lo, hi))`` tuples from a citation cell.

    Handles the colon form (``path:21-23``, ``path:11,16-20``) and the space-L
    form (``path L21-23``), comma-separated multi-ranges, hyphen/en-dash/
    em-dash, and backticked paths. Ranged-only matching (KTD-4): a bare file
    with no line spec contributes no range and never overlaps — whole-file
    coverage belongs to the b7 worker rule, not this parser. Single lines
    yield ``(n, n)``; reversed bounds are normalized.
    """
    out = []
    # strip (not blank) backticks so `path`:21-23 keeps the colon abutting the path
    for m in CODE_LOC_RE.finditer((cell or "").replace("`", "")):
        path = m.group(1)
        for span in _SPAN_RE.finditer(m.group(2)):
            lo = int(span.group(1))
            hi = int(span.group(2)) if span.group(2) else lo
            out.append((path, (min(lo, hi), max(lo, hi))))
    return out


def overlapping_confirmed_pairs(claim_rows, error_rows, claim_cols=None, error_cols=None):
    """Confirmed-claim↔confirmed-error pairs whose cited code lines overlap.

    The claim side parses the claims register's Code/Data Source cell; the
    error side parses the code-error register's Code Location cell (the ranged
    column — the error Code/Data Source carries bare script paths that parse
    to no ranges). Example rows are skipped. Pairs already linked via Related
    Error IDs are the hard check's territory (``confirmed_conflict_links``);
    callers subtract that set before advising.
    """
    ccols = claim_cols or CLAIMS_COLS
    ecols = error_cols or ERROR_COLS
    err_ranges = []
    for r in error_rows:
        if is_example_row(r):
            continue
        d = dict(zip(ecols, r))
        if d.get("Status") != "confirmed" or not d.get("Error ID"):
            continue
        spans = code_line_ranges(d.get("Code Location", ""))
        if spans:
            err_ranges.append((d["Error ID"], spans))
    pairs = []
    for r in claim_rows:
        if is_example_row(r):
            continue
        d = dict(zip(ccols, r))
        if d.get("Status") != "confirmed" or not d.get("Claim ID"):
            continue
        cspans = code_line_ranges(d.get("Code/Data Source", ""))
        if not cspans:
            continue
        for eid, espans in err_ranges:
            if any(cp == ep and clo <= ehi and elo <= chi
                   for cp, (clo, chi) in cspans
                   for ep, (elo, ehi) in espans):
                pairs.append((d["Claim ID"], eid))
    return pairs


# NOTE (U6/U7 open question): the "downstream-use severities must cite a search" rule is NOT
# mechanized here. Detecting that a severity "rests on downstream use" requires interpreting free
# text (Issue/Error Description), which is not cleanly lintable; it stays review guidance
# (registers.md severity rubric), not a lint check.
def escalated_mapped_links(claim_rows, error_rows, claim_cols=None, error_cols=None):
    """Links pairing a `mapped` claim with a `confirmed` code error (escalated mapped claims).
    The cross-linker only links a claim to an error that breaks what it asserts, so a
    mapped-claim↔confirmed-error link is the located-but-unverified miss to escalate for a
    second look. Unlike a status conflict, such a pair may legitimately survive to b8."""
    ccols = claim_cols or CLAIMS_COLS
    ecols = error_cols or ERROR_COLS
    err_status = {
        d["Error ID"]: d.get("Status")
        for d in (dict(zip(ecols, r)) for r in error_rows if not is_example_row(r))
        if d.get("Error ID")
    }
    pairs = []
    for r in claim_rows:
        if is_example_row(r):
            continue
        d = dict(zip(ccols, r))
        if d.get("Status") != "mapped" or not d.get("Claim ID"):
            continue
        for eid in ids_in(d.get("Related Error IDs", ""), "E"):
            if err_status.get(eid) == "confirmed":
                pairs.append((d["Claim ID"], eid))
    return pairs


def severity_divergence_links(claim_rows, error_rows, claim_cols=None, error_cols=None):
    """Links whose two rows carry filled, differing severities (severity divergences)."""
    ccols = claim_cols or CLAIMS_COLS
    ecols = error_cols or ERROR_COLS
    err_sev = {
        d["Error ID"]: d.get("Severity")
        for d in (dict(zip(ecols, r)) for r in error_rows if not is_example_row(r))
        if d.get("Error ID")
    }
    pairs = []
    for r in claim_rows:
        if is_example_row(r):
            continue
        d = dict(zip(ccols, r))
        if not d.get("Severity") or not d.get("Claim ID"):
            continue
        for eid in ids_in(d.get("Related Error IDs", ""), "E"):
            es = err_sev.get(eid)
            if es and es != d["Severity"]:
                pairs.append((d["Claim ID"], eid))
    return pairs


def check_pairs_listed(lint, summary, section, pairs, stage, reason):
    m = re.search(
        rf"^##\s*{re.escape(section)}\b(.*?)(?=^##\s|\Z)", summary or "", re.M | re.S
    )
    lines = (m.group(1) if m else "").split("\n")
    for cid, eid in pairs:
        if not any(cid in ln and eid in ln for ln in lines):
            lint.fail(
                f"{stage}: {reason} ({cid} <-> {eid}) but no line under "
                f"'## {section}' in register_cross_link_summary.md lists the pair"
            )


def stage_b7(lint, audit):
    snap = audit / "_run" / "snapshots" / "b7"
    staging = audit / "_staging"
    # Replay mode: load staging with allow_extra=True so a post-b8 re-run (staging carries the
    # `*Original` rewrite columns) still finds the table; the snapshot keeps the exact base cols.
    c = load_register(lint, staging / "claims_register.md", CLAIMS_COLS, allow_extra=True)
    e = load_register(lint, staging / "code_error_register.md", ERROR_COLS, allow_extra=True)
    cs = load_register(lint, snap / "claims_register.md", CLAIMS_COLS)
    es = load_register(lint, snap / "code_error_register.md", ERROR_COLS)
    if None in (c, e, cs, es):
        return
    # the allow_extra path does not drop schema example rows — filter them here so ID sets match
    c_headers, c_rows = list(c[0]), [r for r in c[1] if not is_example_row(r)]
    e_headers, e_rows = list(e[0]), [r for r in e[1] if not is_example_row(r)]
    non_link_identical(
        lint, c_rows, c_headers, cs[1], CLAIMS_COLS, "Claim ID", "Related Error IDs",
        REWRITE_PAIRS["claims_register.md"], "b7 claims",
    )
    non_link_identical(
        lint, e_rows, e_headers, es[1], ERROR_COLS, "Error ID", "Related Claim IDs",
        REWRITE_PAIRS["code_error_register.md"], "b7 errors",
    )
    check_bidirectional(
        lint, c_rows, c_headers, "Claim ID", "Related Error IDs",
        e_rows, e_headers, "Error ID", "Related Claim IDs", "b7 C<->E",
    )
    summary = read_text(lint, audit / "register_cross_link_summary.md")
    check_pairs_listed(
        lint, summary, "Status conflicts",
        confirmed_conflict_links(c_rows, e_rows, c_headers, e_headers),
        "b7", "confirmed claim linked to confirmed error",
    )
    check_pairs_listed(
        lint, summary, "Escalated mapped claims",
        escalated_mapped_links(c_rows, e_rows, c_headers, e_headers),
        "b7", "mapped claim linked to confirmed error",
    )
    check_pairs_listed(
        lint, summary, "Severity divergences",
        severity_divergence_links(c_rows, e_rows, c_headers, e_headers),
        "b7", "linked pair with differing severities",
    )
    # SC-01 overlap-conflict advisory (U5): warn on confirmed-vs-confirmed pairs
    # whose cited code lines overlap but that no one linked — the case the hard
    # checks above cannot see (they only inspect links a worker already created).
    # Advisory only; the exit code is driven by lint.errors alone.
    linked = set(confirmed_conflict_links(c_rows, e_rows, c_headers, e_headers))
    for cid, eid in overlapping_confirmed_pairs(c_rows, e_rows, c_headers, e_headers):
        if (cid, eid) in linked:
            continue
        lint.warn(
            f"overlap-conflict: confirmed claim {cid} cites code lines overlapping "
            f"confirmed error {eid}, but the pair is not linked and listed under "
            f"'## Status conflicts' — adjudicate against the b7 status-conflict rule "
            f"(registers.md, Cross-link consistency); overlap alone is not proof of conflict"
        )


REWRITE_PAIRS = {
    "claims_register.md": [("Issue Description", "Issue Description Original")],
    "code_error_register.md": [
        ("Error Description", "Error Description Original"),
        ("Why It Matters", "Why It Matters Original"),
    ],
}


def stage_b8(lint, audit, manifest):
    snap = audit / "_run" / "snapshots" / "b8"
    staging = audit / "_staging"
    mode = (manifest or {}).get("mode", "replication")
    files = ["code_error_register.md"] if mode == "code_errors_only" else ["claims_register.md", "code_error_register.md"]
    if mode != "code_errors_only":
        # the rewrite never touches the output register, but `listed` must be gone by now
        out_reg = load_register(lint, audit / "output_register.md", OUTPUT_COLS)
        if out_reg is not None:
            for r in out_reg[1]:
                d = dict(zip(OUTPUT_COLS, r))
                if d["Status"] == "listed":
                    lint.fail(f"{audit / 'output_register.md'}: {d['Output ID']} still 'listed' at b8 (transient status)")
    base_cols = {"claims_register.md": CLAIMS_COLS, "code_error_register.md": ERROR_COLS}
    idc = {"claims_register.md": "Claim ID", "code_error_register.md": "Error ID"}
    for f in files:
        st = load_register(lint, staging / f, base_cols[f], allow_extra=True)
        sn = load_register(lint, snap / f, base_cols[f])
        if st is None or sn is None:
            continue
        st_headers, st_rows = st
        st_rows = [r for r in st_rows if not is_example_row(r)]
        _, sn_rows = sn
        if f == "claims_register.md":
            # U1 advisory: b8 also lints a FINAL claims register. Read under the
            # staging header order (extra `*Original` cols) so field lookup aligns.
            check_adjudication_advisory(lint, staging / f, st_rows, list(st_headers))
            # U5 advisory: filename-parameter reconciliation on blocked rows
            # (attaches at finalize per KTD-4).
            check_filename_parameter_advisory(lint, staging / f, st_rows, list(st_headers))
        if any(h in ("Notes", "Notes Original") for h in st_headers):
            lint.fail(f"{staging / f}: Notes columns are forbidden")
        for new_col, orig_col in REWRITE_PAIRS[f]:
            if orig_col not in st_headers:
                lint.fail(f"{staging / f}: missing column '{orig_col}'")
        if len(st_rows) != len(sn_rows):
            lint.fail(f"{staging / f}: row count changed at rewrite ({len(st_rows)} vs {len(sn_rows)})")
            continue
        st_by = {dict(zip(st_headers, r))[idc[f]]: dict(zip(st_headers, r)) for r in st_rows}
        sn_by = {dict(zip(base_cols[f], r))[idc[f]]: dict(zip(base_cols[f], r)) for r in sn_rows}
        if set(st_by) != set(sn_by):
            lint.fail(f"{staging / f}: ID set changed at rewrite")
            continue
        frozen = [c for c in base_cols[f] if c not in [p[0] for p in REWRITE_PAIRS[f]]]
        for i, row in st_by.items():
            for c in frozen:
                if row.get(c, "") != sn_by[i][c]:
                    lint.fail(f"{staging / f}: {i} column '{c}' changed at rewrite")
            for new_col, orig_col in REWRITE_PAIRS[f]:
                orig_val, prior = row.get(orig_col, ""), sn_by[i][new_col]
                if orig_val != prior:
                    lint.fail(f"{staging / f}: {i} '{orig_col}' does not preserve prior text")
                if bool(row.get(new_col, "")) != bool(orig_val):
                    lint.fail(f"{staging / f}: {i} blankness pairing violated for '{new_col}'")
    if mode != "code_errors_only":
        st_c = load_register(lint, staging / "claims_register.md", CLAIMS_COLS, allow_extra=True)
        st_e = load_register(lint, staging / "code_error_register.md", ERROR_COLS, allow_extra=True)
        if st_c is not None and st_e is not None:
            for cid, eid in confirmed_conflict_links(st_c[1], st_e[1], st_c[0], st_e[0]):
                lint.fail(
                    f"{staging / 'claims_register.md'}: confirmed claim {cid} still linked to "
                    f"confirmed error {eid} at b8 (unresolved status conflict)"
                )
            summary = read_text(lint, audit / "register_cross_link_summary.md")
            # A mapped-claim↔confirmed-error pair may legitimately survive to b8 (unlike a status
            # conflict), but a surviving one must remain documented — b8 checks only that it stays
            # listed, not any status outcome. Verifying the second look actually ran (a recheck
            # ledger entry) is a prose obligation on the conductor (pipeline-finalize b7 step 5).
            check_pairs_listed(
                lint, summary, "Escalated mapped claims",
                escalated_mapped_links(st_c[1], st_e[1], st_c[0], st_e[0]),
                "b8", "mapped claim linked to confirmed error",
            )
            check_pairs_listed(
                lint, summary, "Severity divergences",
                severity_divergence_links(st_c[1], st_e[1], st_c[0], st_e[0]),
                "b8", "linked pair with differing severities",
            )


def export_expected_headers(reg_headers, add_potential_issue):
    """The header row export_xlsx.py writes for a sheet, from the register's own
    headers. Mirrors export_xlsx.drop_and_augment: drop every `*Original` column
    (order preserved), then — on Paper Claims — insert `Potential Issue` right
    after `Status`."""
    keep = [h for h in reg_headers if not h.endswith("Original")]
    if add_potential_issue and "Status" in keep:
        keep = list(keep)
        keep.insert(keep.index("Status") + 1, "Potential Issue")
    return keep


def check_staging_frozen(lint, audit, mode):
    """U8 (d): a `done` b8 leaves `_staging/` populated as the frozen b8 state.
    Export must not precede b8 — so the frozen registers must exist and be
    non-empty at b9."""
    staging = audit / "_staging"
    files = ["code_error_register.md"]
    if mode != "code_errors_only":
        files.insert(0, "claims_register.md")
    for f in files:
        p = staging / f
        if not p.is_file() or p.stat().st_size == 0:
            lint.fail(f"{p}: b8 must leave a non-empty frozen register in _staging/ "
                      f"(else export precedes b8 — rerun b8 before b9)")


def stage_b9(lint, audit, manifest):
    try:
        from openpyxl import load_workbook
    except ImportError:
        lint.fail("b9 requires openpyxl")
        return
    wb_path = audit / "code_review.xlsx"
    if not wb_path.is_file():
        lint.fail(f"missing {wb_path}")
        return
    wb = load_workbook(wb_path, read_only=True)
    mode = (manifest or {}).get("mode", "replication")
    expect = {
        "Overview", "Code Errors", "Late observations (unverified)",
        "Late observation coverage",
    } | ({"Paper Claims"} if mode == "replication" else set())
    if mode == "replication" and (manifest or {}).get("paper_source_set"):
        expect.add("Handoff ledger")
    if set(wb.sheetnames) != expect:
        lint.fail(f"workbook sheets {wb.sheetnames} != expected {sorted(expect)}")
    if "Handoff ledger" in expect and "Handoff ledger" in wb.sheetnames:
        ledger_path = audit / "_run/handoff_ledger.json"
        try:
            ledger = json.loads(ledger_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            lint.fail(f"{ledger_path}: cannot verify exported handoff ledger ({exc})")
            ledger = {"H": [], "X": []}
        columns = [
            "Obligation ID", "Kind", "Source Shard", "Anchor",
            "Destination Worker", "Terminal State", "Covering Claim ID",
            "Disposition",
        ]
        expected_rows = []
        for entry in sorted(ledger.get("H", []) + ledger.get("X", []),
                            key=lambda row: row.get("id", "")):
            disposition = entry.get("disposition") or {}
            anchor = entry.get("anchor", "")
            if isinstance(anchor, dict):
                end = anchor.get("end_line", anchor.get("start_line", ""))
                lines = (str(anchor.get("start_line", ""))
                         if end == anchor.get("start_line")
                         else f"{anchor.get('start_line', '')}-{end}")
                anchor = f"{anchor.get('source_path', '')}:{lines}"
            expected_rows.append([
                str(entry.get("id", "")), str(entry.get("kind", "")),
                str(entry.get("source_shard", "")), str(anchor),
                str(entry.get("destination_worker", "")), str(entry.get("state", "")),
                str(entry.get("covering_c_id") or ""),
                str(disposition.get("reason", "")),
            ])
        data = list(wb["Handoff ledger"].values)
        actual_headers = [str(value) if value is not None else "" for value in data[0]] if data else []
        actual_rows = [[str(value) if value is not None else "" for value in row]
                       for row in data[1:]] if data else []
        if actual_headers != columns or actual_rows != expected_rows:
            lint.fail("Handoff ledger: sheet does not exactly match handoff_ledger.json")
    # U8 (d): the frozen b8 staging registers must still be present and non-empty.
    check_staging_frozen(lint, audit, mode)
    checks = [("Code Errors", "code_error_register.md", ERROR_COLS, "Error ID", ["Error Description", "Why It Matters"])]
    if mode == "replication":
        checks.append(("Paper Claims", "claims_register.md", CLAIMS_COLS, "Claim ID", ["Issue Description", "Potential Issue"]))
    for sheet, f, cols, idc, required in checks:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        data = list(ws.values)
        if not data:
            lint.fail(f"{sheet}: sheet is empty")
            continue
        headers = [str(h) if h is not None else "" for h in data[0]]
        if any(h.endswith("Original") for h in headers):
            lint.fail(f"{sheet}: *_Original column leaked into export")
        for req in required:
            if req not in headers:
                lint.fail(f"{sheet}: required author-facing column '{req}' missing")
        if idc not in headers:
            lint.fail(f"{sheet}: ID column '{idc}' missing")
            continue
        reg = load_register(lint, audit / f, cols, allow_extra=True)
        if reg is None:
            continue
        reg_headers, reg_rows = reg
        reg_rows = [r for r in reg_rows if not is_example_row(r)]
        # U8 (d): exact header parity with the register minus `*Original` (+
        # `Potential Issue` after `Status` on Paper Claims).
        expected_headers = export_expected_headers(
            list(reg_headers), add_potential_issue=(sheet == "Paper Claims"))
        if headers != expected_headers:
            lint.fail(f"{sheet}: export header parity failure — {headers} != {expected_headers}")
        reg_ids = set(col(reg_rows, list(reg_headers), idc)) if idc in reg_headers else set()
        id_i = headers.index(idc)
        sheet_rows = [r for r in data[1:] if r[id_i] is not None]
        sheet_ids = {str(r[id_i]) for r in sheet_rows}
        if sheet_ids != reg_ids:
            lint.fail(f"{sheet}: ID set differs from {f} ({len(sheet_ids)} vs {len(reg_ids)})")
        if len(sheet_rows) != len(reg_rows):
            lint.fail(f"{sheet}: {len(sheet_rows)} rows vs {len(reg_rows)} register rows")
        # U8 (d): Potential Issue = TRUE iff Severity non-empty, per row.
        if sheet == "Paper Claims" and "Potential Issue" in headers and "Severity" in headers:
            pi_i, sev_i = headers.index("Potential Issue"), headers.index("Severity")
            id_j = headers.index(idc)
            for r in sheet_rows:
                sev = str(r[sev_i]).strip() if r[sev_i] is not None else ""
                pi = str(r[pi_i]).strip() if r[pi_i] is not None else ""
                expected = "TRUE" if sev else "FALSE"
                if pi != expected:
                    lint.fail(f"{sheet}: {r[id_j]} Potential Issue '{pi}' disagrees with "
                              f"Severity ('{sev}'): must be '{expected}' "
                              f"(Potential Issue = TRUE iff Severity non-empty)")
    expected_lo = []
    streams = ("claims", "code") if mode == "replication" else ("code",)
    for stream in streams:
        artifact = audit / f"late_observations_{stream}.md"
        b6b_state = manifest.get("stages", {}).get(
            f"{stream}_b6b", {}).get("status")
        if b6b_state == "blocked" and not artifact.is_file():
            # Blocked collection is represented on the coverage surface; it
            # is never silently converted into a zero-observation artifact.
            continue
        # Pending dispositions are legal at b9: the export publishes them on
        # the explicitly unverified sheet.  The completion-report gate lives
        # on `certify_stage.py close-run`, after the Phase-4 first batch.
        _path, rows, _dispositions = _late_observation_rows(lint, audit, stream)
        expected_lo.extend([[stream] + [row[column] for column in LO_COLS]
                            for row in rows])
    expected_lo.sort(key=lambda row: row[1])
    lo_sheet = "Late observations (unverified)"
    if lo_sheet in wb.sheetnames:
        data = list(wb[lo_sheet].values)
        headers = [str(value) if value is not None else "" for value in data[0]] if data else []
        actual = [[str(value) if value is not None else "" for value in row]
                  for row in data[1:]] if data else []
        if headers != ["Stream"] + LO_COLS:
            lint.fail(f"{lo_sheet}: wrong headers {headers}")
        if actual != expected_lo:
            lint.fail(f"{lo_sheet}: rows do not exactly equal stable LO aggregation")
    coverage_path = audit / "_run/late_observation_coverage.md"
    coverage_text = read_text(lint, coverage_path) or ""
    coverage_rows = tables_by_cols(
        lint, coverage_path, coverage_text,
        ["Stream", "Required", "b6b State", "Collection State", "Artifact Head",
         "Blocker Evidence IDs"], "late observation coverage", required=True)
    expected_coverage = []
    for stream in ("claims", "code"):
        required = mode == "replication" or stream == "code"
        entry = manifest.get("stages", {}).get(f"{stream}_b6b", {}) if required else {}
        state = entry.get("status", "not present") if required else "not applicable"
        artifact = audit / f"late_observations_{stream}.md"
        if not required:
            collection = "not required"
        elif state == "blocked":
            collection = "degraded"
        elif state == "done" and artifact.is_file() and artifact.stat().st_size:
            collection = "collected"
        else:
            collection = "incomplete"
        expected_coverage.append({
            "Stream": stream, "Required": "yes" if required else "no",
            "b6b State": state, "Collection State": collection,
            "Artifact Head": "not recorded", "Blocker Evidence IDs": "none recorded",
        })
    if coverage_rows != expected_coverage:
        lint.fail(f"{coverage_path}: coverage cells do not derive exactly from manifest state")
    coverage_sheet = "Late observation coverage"
    if coverage_sheet in wb.sheetnames:
        data = list(wb[coverage_sheet].values)
        headers = [str(value) if value is not None else "" for value in data[0]] if data else []
        actual = [[str(value) if value is not None else "" for value in row]
                  for row in data[1:]] if data else []
        expected_headers = [
            "Stream", "Required", "b6b State", "Collection State", "Artifact Head",
            "Blocker Evidence IDs",
        ]
        expected_values = [[row[column] for column in expected_headers]
                           for row in expected_coverage]
        if headers != expected_headers or actual != expected_values:
            lint.fail(f"{coverage_sheet}: sheet does not match derived coverage artifact")


# --------------------------------------------------------------- main

STAGES = (
    ["b0"]
    + [f"b{n}-{s}" for n in range(1, 4) for s in ("claims", "code")]
    + [f"b3b-{s}" for s in ("claims", "code")]
    + [f"b{n}-{s}" for n in range(4, 6) for s in ("claims", "code")]
    + [f"{stage}-{stream}" for stage in ("b6a", "b5s", "b6b")
       for stream in ("claims", "code")]
    + ["bC", "b7", "b8", "b9"]
)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--stage", required=True, choices=STAGES)
    ap.add_argument("--shard", type=Path, default=None)
    ap.add_argument("--audit-dir", type=Path, default=Path("audit"))
    args = ap.parse_args()

    lint = Lint()
    audit = args.audit_dir
    manifest = {}
    mpath = audit / "_run" / "manifest.json"
    if mpath.is_file():
        try:
            manifest = json.loads(mpath.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            lint.fail(f"{mpath}: invalid JSON")

    stage = args.stage
    n, _, stream = stage.partition("-")
    if stage == "b0":
        stage_b0(lint, audit, manifest)
    elif n == "b1":
        stage_b1(lint, audit, stream, manifest)
    elif n == "b2":
        stage_b2(lint, audit, stream, args.shard)
    elif n == "b3":
        stage_b3(lint, audit, stream, manifest)
    elif n == "b3b":
        if args.shard is not None:
            stage_b3b_shard(lint, audit, stream, args.shard)
        else:
            stage_b3b(lint, audit, stream, manifest)
    elif n == "b4":
        stage_b4(lint, audit, stream, manifest)
    elif n == "b5":
        stage_b5(lint, audit, stream, args.shard, manifest)
    elif n == "b6a":
        stage_b6a(lint, audit, stream, manifest)
    elif n == "b5s":
        stage_b5(lint, audit, stream, args.shard, manifest, supplementary=True)
    elif n == "b6b":
        stage_b6b(lint, audit, stream, manifest)
    elif stage == "bC":
        stage_bC(lint, audit, manifest)
    elif stage == "b7":
        stage_b7(lint, audit)
    elif stage == "b8":
        stage_b8(lint, audit, manifest)
    elif stage == "b9":
        stage_b9(lint, audit, manifest)
    return lint.finish(stage)


if __name__ == "__main__":
    sys.exit(main())
