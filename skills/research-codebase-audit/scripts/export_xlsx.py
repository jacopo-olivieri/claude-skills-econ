#!/usr/bin/env python3
"""Export audit registers to the author-facing Excel workbook.

Deterministic formatting/export step (b9) — never an LLM. Reads the canonical
Markdown registers and the run manifest, writes ``audit/code_review.xlsx``:

- ``Overview``: sheet guide, status legends (statuses actually present),
  variable legends, degraded-confidence warnings from the manifest.
- ``Paper Claims`` (full replication mode only): claims register minus
  ``*_Original`` columns, plus computed ``Potential Issue`` (TRUE iff
  Severity non-empty).
- ``Code Errors``: code-error register minus ``*_Original`` columns.

Usage:
    export_xlsx.py --audit-dir audit [--manifest audit/_run/manifest.json]
                   --mode replication|code_errors_only [-o OUTPUT.xlsx]
"""

import argparse
import json
import os
import re
import sys
import tempfile
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    print("ERROR: export_xlsx.py requires openpyxl (pip install openpyxl)", file=sys.stderr)
    sys.exit(2)

# ---------------------------------------------------------------- md tables


def parse_md_table(text):
    """Parse the first Markdown table in *text* -> (headers, rows)."""
    lines = [l for l in text.split("\n")]
    for i, line in enumerate(lines):
        if line.lstrip().startswith("|") and i + 1 < len(lines) and re.match(
            r"^\s*\|[\s:|-]+\|\s*$", lines[i + 1]
        ):
            headers = split_row(line)
            rows = []
            for row_line in lines[i + 2:]:
                if not row_line.lstrip().startswith("|"):
                    break
                rows.append(split_row(row_line))
            return headers, rows
    return None, []


def split_row(line):
    line = line.strip()
    if line.startswith("|"):
        line = line[1:]
    if line.endswith("|"):
        line = line[:-1]
    cells, cur, esc = [], [], False
    for ch in line:
        if esc:
            cur.append(ch)
            esc = False
        elif ch == "\\":
            cur.append(ch)
            esc = True
        elif ch == "|":
            cells.append("".join(cur).strip())
            cur = []
        else:
            cur.append(ch)
    cells.append("".join(cur).strip())
    return cells


# ---------------------------------------------------------------- legends

DUPLICATE_MEANING = "Duplicate of the referenced row; kept for traceability."

CLAIMS_STATUS_MEANINGS = {
    "confirmed": "The claim is supported by the code, data construction, or documentation reviewed, within the run's evidence limits.",
    "mapped": "The producing code or data was identified, but the claim could not be fully verified without running restricted parts of the pipeline.",
    "unclear": "The claim could not be verified from the available materials (missing or restricted data or scripts).",
    "inconsistent": "The claim appears to conflict with the code, data construction, or shipped outputs.",
    "confirmation_needed": "A second-pass check could not settle this row; author input or a fuller run is needed.",
    "blocked": "The check could not be performed (restricted data, environment limits, or agreed review boundaries); the blocker is documented.",
}

ERROR_STATUS_MEANINGS = {
    "candidate": "Possible error, not yet resolved by the second pass.",
    "confirmed": "The reviewers believe this is a real error, within the run's evidence limits.",
    "not_error": "Reviewed and judged not to be an active error.",
    "confirmation_needed": "A second-pass check could not settle this row; author input or a fuller run is needed.",
    "blocked": "The check could not be performed; the blocker is documented.",
}

CLAIMS_COLUMN_MEANINGS = {
    "Claim ID": "Stable identifier for the claim row.",
    "Paper Context": "Where the claim lives in the paper, in human terms (section > subsection > paragraph or note cue).",
    "Paper Quote": "Verbatim quote from the manuscript containing the claimed fact (searchable with ctrl-F).",
    "Used in Text": "TRUE if the claimed number/object is actually used in the paper; FALSE if it exists only in code or unused artifacts.",
    "Claim Type": "Kind of claim (quantitative result, sample count, specification, robustness, transcription, ...).",
    "Claim Text": "The assertion that was checked.",
    "Code/Data Source": "Script(s), dataset(s), or documentation supporting or contradicting the claim.",
    "Output IDs": "Related paper tables/figures/outputs (output register IDs).",
    "Status": "Verification result — see the status legend.",
    "Potential Issue": "TRUE if the review flagged a potential issue on this row.",
    "Severity": "1 (cosmetic) to 4 (could change a headline result); filled only on flagged rows.",
    "Issue Description": "Author-facing description: what the paper says, what the code shows, why it matters.",
    "Blocked Check": "For a blocked claim, what stayed checkable from visible material (filenames, headers, shapes) and what that check found.",
    "Related Error IDs": "Directly related code-error rows, if any.",
}

ERROR_COLUMN_MEANINGS = {
    "Error ID": "Stable identifier for the error row.",
    "Error Type": "Category of the error (paths, filters, merges, seeds, standard errors, weights, PII, ...).",
    "Code/Data Source": "Script, dataset, config, or output contract involved.",
    "Code Location": "File and line range(s) anchoring the concern.",
    "Status": "Review result — see the status legend.",
    "Severity": "1 (cosmetic) to 4 (could change a headline result).",
    "Error Description": "Author-facing description of what appears wrong.",
    "Why It Matters": "Author-facing consequence for the pipeline, outputs, or paper.",
    "Related Claim IDs": "Directly related paper-claim rows, if any.",
}

SHEET_GUIDE = [
    (
        "Paper Claims",
        "A register of analytical or data-driven claims made in the paper. Each row is a "
        "single claim: the claim text, its source, the verification result, and the "
        "author-facing issue description where something was flagged.",
    ),
    (
        "Code Errors",
        "A register of potential coding errors found in the scripts. Each row is a single "
        "error: its type, location, the author-facing description, and why it matters.",
    ),
    (
        "Late observations (unverified)",
        "Observations raised during the single supplementary wave. These are explicitly unverified and never mutate the registers without an approved bC correction plan.",
    ),
    (
        "Late observation coverage",
        "Per-stream collection coverage derived from certified b6b stage states. A blocked b6b is degraded coverage, never proof of zero observations.",
    ),
]

LO_COLS = ["LO ID", "Source Shard", "Anchor", "Observation"]
LO_SHEET_COLS = ["Stream"] + LO_COLS
LO_COVERAGE_COLS = [
    "Stream", "Required", "b6b State", "Collection State", "Artifact Head",
    "Blocker Evidence IDs",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")


# ---------------------------------------------------------------- sheets


def unescape(cell):
    return cell.replace("\\|", "|").replace("\\\\", "\\")


# Characters that make a spreadsheet treat a cell's text as a formula/command when
# it leads the value: the classic CSV/formula-injection set. This workbook is sent to
# paper authors, so a leading ``=HYPERLINK(...)`` (or ``+``/``-``/``@``/tab/CR) must
# export as inert text, not execute. Prefixing a single apostrophe forces Excel/Sheets
# to treat the whole cell as literal text.
_FORMULA_LEADERS = ("=", "+", "-", "@", "\t", "\r")


def excel_safe(value):
    """Neutralise formula/command injection in a string cell.

    Returns *value* unchanged unless it is a string beginning with a formula-leader
    character, in which case a leading ``'`` is prepended so the cell is inert text.
    Non-string values (ints, floats, None) pass through untouched — only text cells
    can be interpreted as formulas.
    """
    if isinstance(value, str) and value.startswith(_FORMULA_LEADERS):
        return "'" + value
    return value


def coerce_warnings(value):
    """Normalise a manifest ``warnings`` value to a list of strings.

    The schema expects a list, but a hand-edited or partial manifest can carry a
    bare string, ``None``, or some other type. Rather than crash on the Overview
    sheet, coerce defensively:

    - a list  -> each element stringified (already the common case);
    - a string -> a one-element list (treated as a single warning);
    - anything else (``None``, dict, number) -> an empty list (no warnings shown).
    """
    if isinstance(value, list):
        return [str(w) for w in value]
    if isinstance(value, str):
        return [value] if value else []
    return []


def drop_and_augment(headers, rows, drop_originals=True, add_potential_issue=False):
    rows = [r for r in rows if not (r and re.fullmatch(r"[CEO]-0000", r[0]))]  # schema example rows
    keep = [i for i, h in enumerate(headers) if not (drop_originals and h.endswith("Original"))]
    out_headers = [headers[i] for i in keep]
    out_rows = [[unescape(r[i]) if i < len(r) else "" for i in keep] for r in rows]
    if add_potential_issue:
        sev_i = out_headers.index("Severity")
        stat_i = out_headers.index("Status")
        out_headers.insert(stat_i + 1, "Potential Issue")
        for r in out_rows:
            r.insert(stat_i + 1, "TRUE" if r[sev_i].strip() else "FALSE")
    return out_headers, out_rows


def write_data_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title)
    ws.append([excel_safe(h) for h in headers])
    for c in ws[1]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = WRAP
    for row in rows:
        ws.append([excel_safe(v) for v in row])
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = WRAP
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for i, h in enumerate(headers, 1):
        longest = max([len(h)] + [len(str(r[i - 1])) for r in rows] or [10])
        ws.column_dimensions[get_column_letter(i)].width = min(70, max(12, longest // 2 + 8))
    return ws


def write_overview(wb, sheets_present, status_usage, warnings):
    ws = wb.create_sheet("Overview", 0)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 110
    r = 1

    def title(text):
        nonlocal r
        ws.cell(row=r, column=1, value=text).font = Font(bold=True, size=12)
        r += 2

    def table(pairs, head=("", "")):
        nonlocal r
        if head[0]:
            ws.cell(row=r, column=1, value=excel_safe(head[0])).font = Font(bold=True)
            ws.cell(row=r, column=2, value=excel_safe(head[1])).font = Font(bold=True)
            r += 1
        for k, v in pairs:
            ws.cell(row=r, column=1, value=excel_safe(k)).alignment = WRAP
            ws.cell(row=r, column=2, value=excel_safe(v)).alignment = WRAP
            r += 1
        r += 1

    title("Code review — overview")
    table([(s, d) for s, d in SHEET_GUIDE if s in sheets_present], ("Sheet", "Purpose"))

    def legend_pairs(statuses, meanings):
        pairs, saw_dup = [], False
        for s in statuses:
            if s.startswith("duplicate_of:"):
                saw_dup = True
            else:
                pairs.append((s, meanings.get(s, "")))
        if saw_dup:
            pairs.append(("duplicate_of:<ID>", DUPLICATE_MEANING))
        return pairs

    if "Paper Claims" in sheets_present:
        title("Paper Claims status legend")
        table(legend_pairs(status_usage["claims"], CLAIMS_STATUS_MEANINGS), ("Status", "Meaning"))
    title("Code Errors status legend")
    table(legend_pairs(status_usage["errors"], ERROR_STATUS_MEANINGS), ("Status", "Meaning"))
    if "Paper Claims" in sheets_present:
        title("Claims variable legend")
        table(
            [(c, CLAIMS_COLUMN_MEANINGS.get(c, "")) for c in status_usage["claims_cols"]],
            ("Column", "Meaning"),
        )
    title("Code Errors variable legend")
    table(
        [(c, ERROR_COLUMN_MEANINGS.get(c, "")) for c in status_usage["errors_cols"]],
        ("Column", "Meaning"),
    )
    title("Degraded-confidence warnings")
    if warnings:
        table([(f"W{i+1}", w) for i, w in enumerate(warnings)], ("", "Warning"))
    else:
        table([("", "None — review preconditions were met.")])


def _md_table(columns, rows):
    lines = ["| " + " | ".join(columns) + " |",
             "| " + " | ".join(["---"] * len(columns)) + " |"]
    lines.extend("| " + " | ".join(str(cell).replace("|", "\\|") for cell in row) + " |"
                 for row in rows)
    return "\n".join(lines) + "\n"


def _write_atomic(path, payload):
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, temp_name = tempfile.mkstemp(prefix=f".{path.name}.", dir=path.parent)
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as handle:
            handle.write(payload)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temp_name, path)
    except BaseException:
        Path(temp_name).unlink(missing_ok=True)
        raise


def late_observation_rows(audit, mode):
    rows = []
    streams = ("claims", "code") if mode == "replication" else ("code",)
    for stream in streams:
        path = audit / f"late_observations_{stream}.md"
        if not path.is_file():
            continue
        headers, parsed = parse_md_table(path.read_text(encoding="utf-8"))
        if headers == LO_COLS:
            rows.extend([[stream] + row for row in parsed if len(row) == len(LO_COLS)])
    return sorted(rows, key=lambda row: row[1])


def derive_late_observation_coverage(audit, manifest, mode):
    rows = []
    for stream in ("claims", "code"):
        required = mode == "replication" or stream == "code"
        key = f"{stream}_b6b"
        entry = manifest.get("stages", {}).get(key, {}) if required else {}
        state = entry.get("status", "not present") if required else "not applicable"
        artifact = audit / f"late_observations_{stream}.md"
        if not required:
            collection = "not required"
        elif state == "blocked":
            collection = "degraded"
        elif state == "done" and artifact.is_file() and artifact.stat().st_size:
            collection = "collected"
        else:
            collection = "incomplete"
        rows.append([
            stream, "yes" if required else "no", state, collection,
            "not recorded", "none recorded",
        ])
    path = audit / "_run/late_observation_coverage.md"
    _write_atomic(path, "# Late observation coverage\n\n" + _md_table(LO_COVERAGE_COLS, rows))
    return rows


# ---------------------------------------------------------------- main


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--audit-dir", type=Path, default=Path("audit"))
    ap.add_argument("--manifest", type=Path, default=None)
    ap.add_argument("--mode", choices=["replication", "code_errors_only"], required=True)
    ap.add_argument("-o", "--output", type=Path, default=None)
    args = ap.parse_args()

    audit = args.audit_dir
    manifest_path = args.manifest or audit / "_run" / "manifest.json"
    warnings = []
    manifest = {}
    if manifest_path.is_file():
        try:
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError as exc:
            print(f"ERROR: invalid manifest JSON in {manifest_path}: {exc}", file=sys.stderr)
            return 1
        warnings = coerce_warnings(manifest.get("warnings", []))

    errors_md = (audit / "code_error_register.md").read_text(encoding="utf-8")
    e_headers, e_rows = parse_md_table(errors_md)
    if not e_headers:
        print("ERROR: could not parse code_error_register.md", file=sys.stderr)
        return 1

    wb = Workbook()
    wb.remove(wb.active)
    sheets_present, status_usage = [
        "Code Errors", "Late observations (unverified)", "Late observation coverage"
    ], {}

    if args.mode == "replication":
        claims_md = (audit / "claims_register.md").read_text(encoding="utf-8")
        c_headers, c_rows = parse_md_table(claims_md)
        if not c_headers:
            print("ERROR: could not parse claims_register.md", file=sys.stderr)
            return 1
        ch, cr = drop_and_augment(c_headers, c_rows, add_potential_issue=True)
        sheets_present.insert(0, "Paper Claims")
        status_usage["claims"] = sorted({r[ch.index("Status")] for r in cr if r[ch.index("Status")]})
        status_usage["claims_cols"] = ch

    eh, er = drop_and_augment(e_headers, e_rows)
    status_usage["errors"] = sorted({r[eh.index("Status")] for r in er if r[eh.index("Status")]})
    status_usage["errors_cols"] = eh
    lo_rows = late_observation_rows(audit, args.mode)
    coverage_rows = derive_late_observation_coverage(
        audit, manifest, args.mode)

    write_overview(wb, sheets_present, status_usage, warnings)
    if args.mode == "replication":
        write_data_sheet(wb, "Paper Claims", ch, cr)
    write_data_sheet(wb, "Code Errors", eh, er)
    write_data_sheet(wb, "Late observations (unverified)", LO_SHEET_COLS, lo_rows)
    write_data_sheet(wb, "Late observation coverage", LO_COVERAGE_COLS, coverage_rows)

    out = args.output or audit / "code_review.xlsx"
    Path(out).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"OK: wrote {out} ({', '.join(sheets_present)})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
