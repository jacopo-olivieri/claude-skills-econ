"""U7a Part-I tests: deterministic claim-handoff spine and Tier-1 drills."""

import json
import shutil
import subprocess
import sys
from pathlib import Path

import pytest

import regbuild as rb

anchor = rb.load_script("anchor_resolver")
adjudication = rb.load_script("claims_adjudication")
crossref = rb.load_script("build_crossref_inventory")
handoffs = rb.load_script("build_handoff_ledger")
ch = rb.load_script("claim_handoffs")
cs = rb.load_script("certify_stage")
lint = rb.load_script("lint_registers")
paper_sources = rb.load_script("paper_sources")
score_fixture = rb.load_script("score_fixture")
score_replay = rb.load_script("score_replay")
mech = rb.load_script("mechanism_schema")

pytestmark = pytest.mark.u7


def cli(root, command, *args):
    return rb.run_script(
        "certify_stage.py", command, "--package-root", root, *args,
    )


def write_manifest(root, **extra):
    run = root / "audit/_run"
    run.mkdir(parents=True, exist_ok=True)
    manifest = {
        "mode": "replication", "ladder_level": 1,
        "scope_exclusions": [], "off_limits": [],
        "effort_map": dict(cs.dispatch_tracking.DEFAULT_EFFORT_MAP),
    }
    manifest.update(extra)
    (run / "manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")


def source_package(tmp_path, *, override=None):
    root = tmp_path / "package"
    (root / "paper").mkdir(parents=True)
    (root / "paper/main.tex").write_text(
        "\\documentclass{article}\n"
        "\\begin{document}\n"
        "\\input{body}\n"
        "\\include{appendix}\n"
        "\\end{document}\n", encoding="utf-8")
    (root / "paper/body.tex").write_text(
        "\\newcommand{\\speedfig}{Figure~\\ref{fig:speed}}.\n"
        "Calculated and reference speeds overlap substantially \\speedfig.\n"
        "A spotter also notes the appendix's reference-speed assertion.\n",
        encoding="utf-8")
    (root / "paper/appendix.tex").write_text(
        "\\begin{figure}\n"
        "Reference speeds occupy a separate range.\n"
        "\\caption{Calculated and reference speeds}\n"
        "\\label{fig:speed}\n"
        "\\end{figure}\n", encoding="utf-8")
    extra = {"paper_source_path": "paper/main.tex"}
    if override is not None:
        extra["allocation_override"] = override
    write_manifest(root, **extra)
    cs.init_run(root)
    return root


def allocation_rows(root):
    manifest = json.loads((root / "audit/_run/manifest.json").read_text())
    rows = []
    for index, entry in enumerate(manifest["paper_source_set"], start=1):
        count = len(Path(entry["audit_path"]).read_text().splitlines())
        rows.append({
            "Worker ID": f"W{index}", "Paper Scope": f"file {index}",
            "Paper File": entry["source_path"], "Line Intervals": f"1-{count}",
            "Likely Code Scope": "none", "Shard File": f"audit/_work/w{index}.md",
            "Claim ID Range": f"C-{index * 100:04d}–C-{index * 100 + 49:04d}",
            "Output ID Range": f"O-{index * 100:04d}–O-{index * 100 + 29:04d}",
            "H ID Range": f"H-{index * 100:04d}–H-{index * 100 + 49:04d}",
            "Review Focus": "all assertions",
        })
    return rows


def write_plan(root, rows=None):
    rows = allocation_rows(root) if rows is None else rows
    table = rb.md_table(ch.CLAIMS_PLAN_COLS, [[row[col] for col in ch.CLAIMS_PLAN_COLS]
                                              for row in rows])
    path = root / "audit/plans/claims_review_plan.md"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "# Claims plan\n\n" + table + "\n"
        "Merge-coordinator range: C-8000–C-8049\n"
        "Merge-coordinator range: O-8000–O-8049\n"
        "Adjudication range: C-8100–C-8149\n",
        encoding="utf-8",
    )
    return path, rows


def claims_shard(claim_rows=(), *, handoff_rows=None, x_rows=None):
    body = rb.register_text("Claims", lint.CLAIMS_COLS, list(claim_rows))
    body += "\n" + rb.register_text("Outputs", lint.OUTPUT_COLS, [])
    body += "\n### Handoffs\n\n"
    body += ("No handoffs.\n" if handoff_rows is None
             else rb.md_table(ch.HANDOFF_COLS, handoff_rows))
    body += "\n### Cross-reference coverage\n\n"
    body += ("No assigned cross-references.\n" if x_rows is None
             else rb.md_table(ch.X_COVERAGE_COLS, x_rows))
    body += "\n### Coverage\n\nEvery assertion in scope was reviewed.\n"
    body += "\n### Footer dispositions\n\n" + rb.md_table(lint.FOOTER_COLS, [])
    return body


def prepare_b2(root):
    manifest = json.loads((root / "audit/_run/manifest.json").read_text())
    inventory = json.loads((root / "audit/_run/crossref_inventory.json").read_text())
    assignments = json.loads((root / "audit/_run/crossref_assignments.json").read_text())[
        "assignments"]
    rows, _ = ch.load_claims_allocations(root / "audit/plans/claims_review_plan.md")
    appendix = next(entry for entry in manifest["paper_source_set"]
                    if entry["source_path"].endswith("appendix.tex"))
    appendix_worker = next(row["Worker ID"] for row in rows
                           if row["Paper File"] == appendix["source_path"])
    body_worker = next(iter(assignments.values()))
    for row in rows:
        handoff_rows = None
        x_rows = None
        if row["Worker ID"] == body_worker:
            handoff_rows = [[
                row["H ID Range"].split("–")[0],
                f"{appendix['source_path']}:2",
                "Reference speeds occupy a separate range.",
                "The appendix asserts that reference speeds occupy a separate range.",
                "Figure A2",
            ]]
            assigned = [entry for entry in inventory["entries"]
                        if assignments[entry["id"]] == body_worker]
            x_rows = [[
                entry["id"], "disposition", "bare_pointer",
                "sentence: Figure pointer; no_checkable_predicate: pointer only",
                "—", "—",
            ] for entry in assigned]
        shard = root / row["Shard File"]
        shard.parent.mkdir(parents=True, exist_ok=True)
        shard.write_text(claims_shard(handoff_rows=handoff_rows, x_rows=x_rows),
                         encoding="utf-8")
    return rows, body_worker, appendix_worker


def write_empty_canon_and_report(root):
    audit = root / "audit"
    (audit / "claims_register.md").write_text(
        rb.register_text("Claims", lint.CLAIMS_COLS, []), encoding="utf-8")
    (audit / "output_register.md").write_text(
        rb.register_text("Outputs", lint.OUTPUT_COLS, []), encoding="utf-8")
    report = {
        "claims_register.md": {"shard_rows": 0, "dedup_removed": 0, "added": 0,
                               "conflicts": [], "coverage_gaps": [], "blocked_shards": []},
        "output_register.md": {"shard_rows": 0, "dedup_removed": 0, "added": 0,
                               "conflicts": [], "coverage_gaps": [], "blocked_shards": []},
        "footer_dispositions": [],
    }
    (audit / "_run/merge_report_claims.json").write_text(
        json.dumps(report, indent=2), encoding="utf-8")


def prepared_b3(tmp_path, certify=False):
    root = source_package(tmp_path)
    write_plan(root)
    crossref.build(root, root / "audit")
    if certify:
        cs.start_stage(root, "claims_b1")
        cs.finish_stage(root, "claims_b1", "done")
    rows, body_worker, appendix_worker = prepare_b2(root)
    cs.start_stage(root, "claims_b2")
    for row in rows:
        cs.set_shard(root, "claims_b2", row["Shard File"], "done")
    cs.finish_stage(root, "claims_b2", "done")
    write_empty_canon_and_report(root)
    handoffs.build(root, root / "audit", "claims_b3")
    return root, rows, body_worker, appendix_worker


PROSE_CONTEXT = 'Appendix > Figure A2 > sentence beginning "Reference speeds"'


def add_b3b_resolution(root, rows, body_worker):
    audit = root / "audit"
    ledger = json.loads((audit / "_run/snapshots/claims_b3/handoff_ledger.json").read_text())
    forwarded = next(entry for entry in ledger["H"] if entry["state"] == "forwarded")
    plan = (
        "# Claims second-read plan\n\n"
        "| Worker ID | File/Section Scope | Shard File | Claim ID Range | Output ID Range | Reason | Known Findings | Assigned Handoff IDs |\n"
        "| --- | --- | --- | --- | --- | --- | --- | --- |\n"
        f"| R1 | appendix | `audit/_work_second_read/r1.md` | C-2000–C-2049 | O-2000–O-2029 | handoff | — | {forwarded['id']} |\n"
    )
    (audit / "plans/claims_second_read_plan.md").write_text(plan, encoding="utf-8")
    # Paper Context stays the registers.md prose locator; the machine anchor
    # travels on the resolution entry's Covering Range/Quote cells (C-F1).
    claim = rb.claims_row(
        "C-2000", context=PROSE_CONTEXT, quote=forwarded["quote"],
        text=forwarded["asserted_substance"], status="confirmed",
    )
    resolution = [[
        forwarded["id"], forwarded["anchor"], forwarded["quote"],
        forwarded["asserted_substance"], forwarded["referenced_objects"],
        "resolved", "C-2000", "—", forwarded["anchor"], forwarded["quote"],
    ]]
    shard = audit / "_work_second_read/r1.md"
    shard.parent.mkdir(parents=True, exist_ok=True)
    body = rb.register_text("Claims", lint.CLAIMS_COLS, [claim])
    body += "\n" + rb.register_text("Outputs", lint.OUTPUT_COLS, [])
    body += "\n### Handoffs\n\n" + rb.md_table(ch.HANDOFF_RESOLUTION_COLS, resolution)
    body += "\n### Coverage\n\nEvery assigned handoff was resolved.\n"
    body += "\n### Footer dispositions\n\n" + rb.md_table(
        lint.FOOTER_COLS, [["OBS-0001", "candidate", "C-2000", "row retained", ""]]
    )
    shard.write_text(body, encoding="utf-8")
    snap = audit / "_run/snapshots/claims_b3b"
    snap.mkdir(parents=True, exist_ok=True)
    (snap / "claims_register.md").write_text(
        rb.register_text("Claims", lint.CLAIMS_COLS, []), encoding="utf-8")
    (snap / "output_register.md").write_text(
        rb.register_text("Outputs", lint.OUTPUT_COLS, []), encoding="utf-8")
    (audit / "claims_register.md").write_text(
        rb.register_text("Claims", lint.CLAIMS_COLS, [claim]), encoding="utf-8")
    report = {
        "claims_register.md": {"shard_rows": 1, "dedup_removed": 0, "added": 1,
                               "conflicts": [], "coverage_gaps": [], "blocked_shards": []},
        "output_register.md": {"shard_rows": 0, "dedup_removed": 0, "added": 0,
                               "conflicts": [], "coverage_gaps": [], "blocked_shards": []},
        "footer_dispositions": [
            "audit/_work_second_read/r1.md#OBS-0001 | candidate:C-2000"
        ],
    }
    (audit / "_run/merge_report_claims_b3b.json").write_text(
        json.dumps(report, indent=2), encoding="utf-8")
    cs.start_stage(root, "claims_b3b")
    cs.set_shard(root, "claims_b3b", "audit/_work_second_read/r1.md", "done")
    handoffs.build(root, audit, "claims_b3b")
    return forwarded, shard


def prepare_adjudication(tmp_path):
    root, rows, body, _appendix = prepared_b3(tmp_path)
    add_b3b_resolution(root, rows, body)
    audit = root / "audit"
    snap = audit / "_run/snapshots/claims_adjudication"
    snap.mkdir(parents=True, exist_ok=True)
    shutil.copy2(audit / "claims_register.md", snap / "claims_register.md")
    shutil.copy2(audit / "output_register.md", snap / "output_register.md")
    shutil.copy2(audit / "_run/handoff_ledger.json", snap / "handoff_ledger.json")
    worklist = adjudication.build_worklist(root, audit, "claims_adjudication")
    return root, worklist


def freeze_lineage_inputs(root):
    audit = root / "audit"
    snap = audit / "_run/snapshots/claims_adjudication_lineage"
    snap.mkdir(parents=True, exist_ok=True)
    shutil.copy2(audit / "claims_register.md", snap / "claims_register.md")
    shutil.copy2(audit / "_run/handoff_ledger.json", snap / "handoff_ledger.json")


def write_adjudication_verdicts(root, worklist, overrides=None):
    overrides = overrides or {}
    rows = []
    for item in worklist["items"]:
        verdict = overrides.get(item["id"])
        if verdict is None:
            verdict = ("capture_confirmed" if item["work_kind"] == "mapping"
                       else "disposition_accepted")
        rows.append([
            item["id"], item["work_kind"], verdict, "fresh adjudicator checked capture",
            *(["—"] * (len(adjudication.ADJUDICATION_VERDICT_COLS) - 4)),
        ])
    path = root / "audit/_run/claims_adjudication_verdicts.md"
    path.write_text(rb.md_table(adjudication.ADJUDICATION_VERDICT_COLS, rows),
                    encoding="utf-8")
    return path


# ---------------------------------------------------------------- intake/b1


def test_multifile_intake_builds_pinned_twins_and_preserves_lines(tmp_path):
    root = source_package(tmp_path)
    manifest = json.loads((root / "audit/_run/manifest.json").read_text())
    assert [Path(entry["source_path"]).name for entry in manifest["paper_source_set"]] == [
        "main.tex", "body.tex", "appendix.tex",
    ]
    assert all("audit/_run/paper_twins" in entry["audit_path"]
               for entry in manifest["paper_source_set"])
    paper_sources.validate_source_set(root, manifest)


@pytest.mark.parametrize("command", ["subfile", "includestandalone"])
def test_intake_refuses_unsupported_inclusion_with_line(command, tmp_path):
    root = tmp_path / "package"
    (root / "paper").mkdir(parents=True)
    (root / "paper/child.tex").write_text("child\n", encoding="utf-8")
    (root / "paper/main.tex").write_text(
        f"first\n\\{command}{{child}}\n", encoding="utf-8")
    write_manifest(root, paper_source_path="paper/main.tex")
    with pytest.raises(cs.CertificationError, match=r"unsupported inclusion syntax.*main.tex:2"):
        cs.init_run(root)


def test_intake_refuses_unknown_macro_that_resolves_tex(tmp_path):
    root = tmp_path / "package"
    (root / "paper").mkdir(parents=True)
    (root / "paper/child.tex").write_text("child\n", encoding="utf-8")
    (root / "paper/main.tex").write_text("\\mystery{child}\n", encoding="utf-8")
    write_manifest(root, paper_source_path="paper/main.tex")
    with pytest.raises(cs.CertificationError, match=r"unsupported inclusion-like macro.*main.tex:1"):
        cs.init_run(root)


def test_init_refuses_paperless_replication_without_override(tmp_path):
    root = tmp_path / "package"
    write_manifest(root)
    with pytest.raises(cs.CertificationError, match="paper_source_path"):
        cs.init_run(root)


def test_b1_exact_partition_rejects_gap_and_accepts_complete(tmp_path):
    root = source_package(tmp_path)
    _path, rows = write_plan(root)
    assert rb.lint(rb.AuditDir(root), "b1-claims").returncode == 0
    rows[0]["Line Intervals"] = "2-5"
    write_plan(root, rows)
    result = rb.lint(rb.AuditDir(root), "b1-claims")
    assert result.returncode == 1
    assert "unowned line" in result.stdout


def test_allocation_override_validates_purpose_and_exact_headers(tmp_path):
    with pytest.raises(cs.CertificationError, match="purpose"):
        source_package(tmp_path / "bad", override={"purpose": "gate", "allocation": []})
    bad_row = {column: "x" for column in reversed(ch.CLAIMS_PLAN_COLS)}
    with pytest.raises(cs.CertificationError, match="exact ordered b1 headers"):
        source_package(tmp_path / "shape", override={
            "purpose": "fixture", "allocation": [bad_row],
        })
    good = source_package(tmp_path / "good", override={
        "purpose": "development", "allocation": [],
    })
    manifest = json.loads((good / "audit/_run/manifest.json").read_text())
    assert manifest["allocation_override"]["purpose"] == "development"


# --------------------------------------------------------- inventory/anchors


def test_inventory_expands_local_macro_and_assigns_anchor_owner(tmp_path):
    root = source_package(tmp_path)
    _path, rows = write_plan(root)
    crossref.build(root, root / "audit")
    inventory = json.loads((root / "audit/_run/crossref_inventory.json").read_text())
    assert len(inventory["entries"]) == 1
    entry = inventory["entries"][0]
    assert entry["id"] == "X-0001"
    assert entry["referenced_float_labels"] == ["fig:speed"]
    assignments = json.loads((root / "audit/_run/crossref_assignments.json").read_text())
    body_worker = next(row["Worker ID"] for row in rows
                       if row["Paper File"].endswith("body.tex"))
    assert assignments["assignments"] == {"X-0001": body_worker}
    crossref.check(root, root / "audit")


def test_inventory_digest_and_byte_compare_reject_replan_and_edit(tmp_path):
    root = source_package(tmp_path)
    plan, rows = write_plan(root)
    crossref.build(root, root / "audit")
    rows[0]["Paper Scope"] = "renamed scope"
    write_plan(root, rows)
    with pytest.raises(crossref.CrossrefError, match="stale, or edited"):
        crossref.check(root, root / "audit")
    write_plan(root, rows)
    crossref.build(root, root / "audit")
    inventory_path = root / "audit/_run/crossref_inventory.json"
    payload = json.loads(inventory_path.read_text())
    payload["entries"][0]["destination_worker"] = "W999"
    inventory_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    with pytest.raises(crossref.CrossrefError, match="stale, or edited"):
        crossref.check(root, root / "audit")


def test_inventory_records_unexpanded_package_macro_warning(tmp_path):
    root = source_package(tmp_path)
    body = root / "paper/body.tex"
    body.write_text("A package indirection names \\speedfigure.\n", encoding="utf-8")
    # Re-init refreshes the source set/twins after the deliberate source edit.
    cs.init_run(root, clear_stale_marker=True)
    write_plan(root)
    crossref.build(root, root / "audit")
    inventory = json.loads((root / "audit/_run/crossref_inventory.json").read_text())
    assert inventory["entries"] == []
    assert "unexpanded package macro \\speedfigure" in inventory["macro_warnings"][0]["macro"]


def test_inventory_emits_hard_coded_printed_reference_without_target(tmp_path):
    root = source_package(tmp_path)
    (root / "paper/body.tex").write_text(
        "Calculated and reference speeds overlap in Figure A2.\n", encoding="utf-8"
    )
    cs.init_run(root, clear_stale_marker=True)
    write_plan(root)
    crossref.build(root, root / "audit")
    entries = json.loads((root / "audit/_run/crossref_inventory.json").read_text())["entries"]
    assert len(entries) == 1
    assert entries[0]["kind"] == "printed_reference"


@pytest.mark.parametrize("mutation", ["delete", "replan", "edit"])
def test_verify_run_refuses_inventory_mutations_tier1_drill(mutation, tmp_path):
    root = source_package(tmp_path)
    _plan, rows = write_plan(root)
    crossref.build(root, root / "audit")
    cs.start_stage(root, "claims_b1")
    cs.finish_stage(root, "claims_b1", "done")
    inventory_path = root / "audit/_run/crossref_inventory.json"
    if mutation == "delete":
        inventory_path.unlink()
    elif mutation == "replan":
        rows[0]["Paper Scope"] = "changed after certification"
        write_plan(root, rows)
    else:
        payload = json.loads(inventory_path.read_text())
        payload["entries"][0]["destination_worker"] = "W999"
        inventory_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    result = cli(root, "verify-run")
    assert result.returncode == 1
    assert "crossref" in result.stderr.lower() or "claims allocation" in result.stderr.lower()


def test_anchor_resolver_normalizes_multiline_and_refuses_ambiguity(tmp_path):
    root = source_package(tmp_path)
    manifest = json.loads((root / "audit/_run/manifest.json").read_text())
    body = next(entry for entry in manifest["paper_source_set"]
                if entry["source_path"].endswith("body.tex"))
    resolved = anchor.resolve_quote(
        manifest["paper_source_set"], f"{body['source_path']}:2-3",
        "Calculated and reference speeds overlap substantially \\speedfig. A spotter",
        root,
    )
    assert resolved["start_line"] == 2
    Path(body["audit_path"]).write_text("same quote\nsame quote\n", encoding="utf-8")
    with pytest.raises(anchor.AnchorError, match="2 occurrences"):
        anchor.resolve_quote(
            manifest["paper_source_set"], f"{body['source_path']}:1-2", "same quote", root
        )
    with pytest.raises(anchor.AnchorError, match="at most five"):
        anchor.parse_anchor(f"{body['source_path']}:1-6")


# ---------------------------------------------------------- shard/ledger CLI


def test_b2_shard_requires_terminal_x_and_valid_h_anchor(tmp_path):
    root = source_package(tmp_path)
    write_plan(root)
    crossref.build(root, root / "audit")
    rows, body_worker, _ = prepare_b2(root)
    body = next(row for row in rows if row["Worker ID"] == body_worker)
    shard = root / body["Shard File"]
    assert rb.lint(rb.AuditDir(root), "b2-claims", shard).returncode == 0
    shard.write_text(shard.read_text().replace("X-0001", "X-9999"), encoding="utf-8")
    failed = rb.lint(rb.AuditDir(root), "b2-claims", shard)
    assert failed.returncode == 1
    assert "do not equal assignment" in failed.stdout


def test_b3_ledger_exact_sets_and_forwarded_state(tmp_path):
    root, _rows, _body, appendix = prepared_b3(tmp_path)
    ledger = json.loads((root / "audit/_run/snapshots/claims_b3/handoff_ledger.json").read_text())
    assert [entry["state"] for entry in ledger["H"]] == ["forwarded"]
    assert ledger["H"][0]["destination_worker"] == appendix
    assert [entry["state"] for entry in ledger["X"]] == ["disposition"]
    handoffs.check(root, root / "audit", "claims_b3")


def test_b3_verify_run_rederivation_rejects_deleted_and_flipped_ledger(tmp_path):
    root, _rows, _body, _appendix = prepared_b3(tmp_path)
    cs.start_stage(root, "claims_b3")
    cs.finish_stage(root, "claims_b3", "done")
    ledger_path = root / "audit/_run/snapshots/claims_b3/handoff_ledger.json"
    original = json.loads(ledger_path.read_text())
    deleted = dict(original)
    deleted["H"] = []
    ledger_path.write_text(json.dumps(deleted, indent=2), encoding="utf-8")
    assert cli(root, "verify-run").returncode == 1
    flipped = original
    flipped["H"][0]["state"] = "satisfied"
    ledger_path.write_text(json.dumps(flipped, indent=2), encoding="utf-8")
    result = cli(root, "verify-run")
    assert result.returncode == 1
    assert "handoff ledger" in result.stderr.lower()


def test_b3b_column_exactly_covers_forwarded_and_resolution_contains(tmp_path):
    root, rows, body, _appendix = prepared_b3(tmp_path)
    forwarded, shard = add_b3b_resolution(root, rows, body)
    assert rb.lint(rb.AuditDir(root), "b3b-claims", shard).returncode == 0
    plan = root / "audit/plans/claims_second_read_plan.md"
    plan.write_text(plan.read_text().replace(forwarded["id"], "—"), encoding="utf-8")
    failed = rb.lint(rb.AuditDir(root), "b3b-claims", shard)
    assert failed.returncode == 1
    assert "exactly cover forwarded" in failed.stdout


def test_b3b_verbatim_quote_mismatch_refuses_in_lint_and_builder(tmp_path):
    root, rows, body, _appendix = prepared_b3(tmp_path)
    forwarded, shard = add_b3b_resolution(root, rows, body)
    # Mutate only the trailing Covering Quote cell of the resolution row; the
    # copied filing cells and the cited C-row stay intact, so the verbatim
    # Paper-Quote equality check is the one that must fire (C-F4C).
    text = shard.read_text()
    tail = f"| — | {forwarded['anchor']} | {forwarded['quote']} |"
    assert text.count(tail) == 1
    shard.write_text(text.replace(
        tail, f"| — | {forwarded['anchor']} | Calculated and reference speeds |",
    ), encoding="utf-8")
    failed = rb.lint(rb.AuditDir(root), "b3b-claims", shard)
    assert failed.returncode == 1
    assert "not verbatim Paper Quote" in failed.stdout
    with pytest.raises(handoffs.LedgerError, match="Paper Quote verbatim"):
        handoffs.derive(root, root / "audit", "claims_b3b")


def test_b3b_clause_b_cannot_close_clause_a(tmp_path):
    root, rows, body, _appendix = prepared_b3(tmp_path)
    forwarded, shard = add_b3b_resolution(root, rows, body)
    manifest = json.loads((root / "audit/_run/manifest.json").read_text())
    appendix = next(entry for entry in manifest["paper_source_set"]
                    if entry["source_path"].endswith("appendix.tex"))
    twin = Path(appendix["audit_path"])
    twin.write_text(twin.read_text().replace(
        "Reference speeds occupy a separate range.",
        "Reference speeds occupy a separate range. Calculated speeds occupy another range.",
    ), encoding="utf-8")
    text = shard.read_text()
    # Change only the C-row and resolution covering cells; copied H filing
    # cells remain clause A verbatim.
    text = text.replace(
        f"| C-2000 | {PROSE_CONTEXT} | {forwarded['quote']} |",
        f"| C-2000 | {PROSE_CONTEXT} | Calculated speeds occupy another range. |",
    )
    text = text.replace(
        f"| resolved | C-2000 | — | {forwarded['anchor']} | {forwarded['quote']} |",
        f"| resolved | C-2000 | — | {forwarded['anchor']} | Calculated speeds occupy another range. |",
    )
    shard.write_text(text, encoding="utf-8")
    failed = rb.lint(rb.AuditDir(root), "b3b-claims", shard)
    assert failed.returncode == 1
    assert "does not contain the obligation assertion" in failed.stdout


def test_deterministic_chain_reaches_b3b_then_close_run_refuses_pending_tail(tmp_path):
    root, rows, body, _appendix = prepared_b3(tmp_path, certify=True)
    cs.start_stage(root, "claims_b3")
    cs.finish_stage(root, "claims_b3", "done")
    add_b3b_resolution(root, rows, body)
    cs.finish_stage(root, "claims_b3b", "done")
    ledger = json.loads((root / "audit/_run/handoff_ledger.json").read_text())
    assert ledger["H"][0]["state"] == "resolved"
    result = cli(root, "close-run")
    assert result.returncode == 1
    assert "claims_adjudication is pending" in result.stderr
    assert (root / "audit/_run/RUNNING").is_file()


def test_blocked_operator_decisions_are_exact_joined_and_cannot_bypass_stage_refusal(tmp_path):
    root, _rows, _body, _appendix = prepared_b3(tmp_path)
    ledger_path = root / "audit/_run/handoff_ledger.json"
    ledger = json.loads(ledger_path.read_text())
    ledger["H"][0]["state"] = "blocked_fallback"
    ledger_path.write_text(json.dumps(ledger, indent=2), encoding="utf-8")
    decisions = [{
        "id": ledger["H"][0]["id"], "decision": "accept_blocked",
        "reason": "operator accepts the documented dead worker", "date": "2026-07-20",
    }]
    (root / "audit/_run/handoff_blocked_decisions.json").write_text(
        json.dumps(decisions, indent=2), encoding="utf-8")
    result = cli(root, "close-run")
    assert result.returncode == 1
    assert "claims_adjudication is pending" in result.stderr
    decisions.append({"id": "X-9999", "decision": "accept_blocked",
                      "reason": "extra", "date": "2026-07-20"})
    (root / "audit/_run/handoff_blocked_decisions.json").write_text(
        json.dumps(decisions, indent=2), encoding="utf-8")
    result = cli(root, "close-run")
    assert "unknown obligation X-9999" in result.stderr


def test_close_run_refuses_forged_manifest_with_absent_adjudication_key(tmp_path):
    root, _rows, _body, _appendix = prepared_b3(tmp_path)
    manifest_path = root / "audit/_run/manifest.json"
    manifest = json.loads(manifest_path.read_text())
    del manifest["stages"]["claims_adjudication"]
    manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    result = cli(root, "close-run")
    assert result.returncode == 1
    assert "claims_adjudication is absent" in result.stderr


def test_inventory_keeps_fig_abbreviation_in_one_sentence(tmp_path):
    root = source_package(tmp_path)
    (root / "paper/body.tex").write_text(
        "Speeds overlap broadly in Fig. 2 of the appendix.\n", encoding="utf-8")
    cs.init_run(root, clear_stale_marker=True)
    write_plan(root)
    crossref.build(root, root / "audit")
    entries = json.loads(
        (root / "audit/_run/crossref_inventory.json").read_text())["entries"]
    assert len(entries) == 1
    assert entries[0]["kind"] == "printed_reference"
    assert entries[0]["referencing_sentence"] == (
        "Speeds overlap broadly in Fig. 2 of the appendix.")


def test_b2_ambiguous_quote_refused_through_lint_cli(tmp_path):
    root = source_package(tmp_path)
    write_plan(root)
    crossref.build(root, root / "audit")
    rows, body_worker, _appendix = prepare_b2(root)
    manifest = json.loads((root / "audit/_run/manifest.json").read_text())
    appendix = next(entry for entry in manifest["paper_source_set"]
                    if entry["source_path"].endswith("appendix.tex"))
    twin = Path(appendix["audit_path"])
    twin.write_text(twin.read_text().replace(
        "Reference speeds occupy a separate range.",
        "Reference speeds occupy a separate range. "
        "Reference speeds occupy a separate range.",
    ), encoding="utf-8")
    body = next(row for row in rows if row["Worker ID"] == body_worker)
    failed = rb.lint(rb.AuditDir(root), "b2-claims", root / body["Shard File"])
    assert failed.returncode == 1
    assert "2 occurrences" in failed.stdout


def test_b3_populated_canon_matches_prose_context_rows_by_containment(tmp_path):
    root = source_package(tmp_path)
    write_plan(root)
    crossref.build(root, root / "audit")
    rows, _body, _appendix = prepare_b2(root)
    cs.start_stage(root, "claims_b2")
    for row in rows:
        cs.set_shard(root, "claims_b2", row["Shard File"], "done")
    cs.finish_stage(root, "claims_b2", "done")
    write_empty_canon_and_report(root)
    audit = root / "audit"
    covering = rb.claims_row(
        "C-0100", context=PROSE_CONTEXT,
        quote="Reference speeds occupy a separate range.",
        text="the appendix states reference speeds sit apart", status="confirmed")
    (audit / "claims_register.md").write_text(
        rb.register_text("Claims", lint.CLAIMS_COLS, [covering]), encoding="utf-8")
    handoffs.build(root, audit, "claims_b3")
    ledger = json.loads((audit / "_run/handoff_ledger.json").read_text())
    assert [(entry["state"], entry["covering_c_id"])
            for entry in ledger["H"]] == [("satisfied", "C-0100")]
    noncontaining = rb.claims_row(
        "C-0100", context=PROSE_CONTEXT,
        quote="\\caption{Calculated and reference speeds}",
        text="caption text only", status="confirmed")
    (audit / "claims_register.md").write_text(
        rb.register_text("Claims", lint.CLAIMS_COLS, [noncontaining]),
        encoding="utf-8")
    handoffs.build(root, audit, "claims_b3")
    ledger = json.loads((audit / "_run/handoff_ledger.json").read_text())
    assert [entry["state"] for entry in ledger["H"]] == ["forwarded"]


def test_score_fixture_missing_override_fails_p25_split_check(tmp_path):
    audit = tmp_path / "audit"
    (audit / "_run").mkdir(parents=True)
    for fname, cols in (("claims_register.md", lint.CLAIMS_COLS),
                        ("output_register.md", lint.OUTPUT_COLS),
                        ("code_error_register.md", lint.ERROR_COLS)):
        (audit / fname).write_text(
            rb.register_text(fname, cols, []), encoding="utf-8")
    expected_path = tmp_path / "expected.json"
    expected_path.write_text(json.dumps({
        "must_find": [{"id": "P-25", "kind": "claim", "min_severity": 2,
                       "mechanism": "overlap plant"}],
        "must_not_find": [],
    }), encoding="utf-8")
    result = rb.run_script(
        "score_fixture.py", "--audit-dir", audit, "--expected", expected_path)
    assert result.returncode == 1
    assert "U7a allocation split: FAIL" in result.stdout
    assert "allocation_override" in result.stdout


def test_score_fixture_split_assertion_rejects_one_worker(tmp_path):
    root = source_package(tmp_path)
    rows = allocation_rows(root)
    manifest = json.loads((root / "audit/_run/manifest.json").read_text())
    paper = next(entry for entry in manifest["paper_source_set"]
                 if entry["source_path"].endswith("main.tex"))
    # The production P-25 source is a separate fixture; this unit directly
    # proves the scorer's one-worker refusal using a copied P-25-shaped paper.
    Path(paper["source_path"]).write_text(
        "The calculated and reference speeds show substantial overlap (Figure A2).\n"
        "\\label{fig:speed-overlap}\n", encoding="utf-8")
    manifest["paper_source_set"][0]["source_path"] = str(Path(paper["source_path"]))
    (root / "audit/_run/manifest.json").write_text(json.dumps(manifest), encoding="utf-8")
    rows = [rows[0]]
    rows[0]["Paper File"] = str(Path(paper["source_path"]))
    rows[0]["Line Intervals"] = "1-2"
    write_plan(root, rows)
    status, note = score_fixture.check_u7_allocation_split(root / "audit")
    assert status == "FAIL"
    assert "fewer than two workers" in note


def test_score_fixture_split_assertion_accepts_distinct_workers(tmp_path):
    root = tmp_path / "package"
    paper = root / "paper/paper.tex"
    paper.parent.mkdir(parents=True)
    paper.write_text(
        "The calculated and reference speeds show substantial overlap (Figure A2).\n"
        "\\label{fig:speed-overlap}\n", encoding="utf-8"
    )
    audit = root / "audit"
    (audit / "_run").mkdir(parents=True)
    source_set = [{
        "source_path": str(paper), "source_sha256": "unused",
        "audit_path": str(paper), "audit_sha256": "unused",
    }]
    (audit / "_run/manifest.json").write_text(
        json.dumps({"paper_source_set": source_set}), encoding="utf-8"
    )
    rows = []
    for worker, interval, shard, base in (
            ("W1", "1-1", "audit/_work/w1.md", 100),
            ("W2", "2-2", "audit/_work/w2.md", 200)):
        rows.append({
            "Worker ID": worker, "Paper Scope": worker, "Paper File": str(paper),
            "Line Intervals": interval, "Likely Code Scope": "none",
            "Shard File": shard, "Claim ID Range": f"C-{base:04d}–C-{base+49:04d}",
            "Output ID Range": f"O-{base:04d}–O-{base+29:04d}",
            "H ID Range": f"H-{base:04d}–H-{base+49:04d}", "Review Focus": "all",
        })
    write_plan(root, rows)
    status, note = score_fixture.check_u7_allocation_split(audit)
    assert status == "PASS"
    assert "W1" in note and "W2" in note


# ---------------------------------------------------------- U7b adjudication


def test_adjudication_done_exactly_closes_mapping_and_disposition(tmp_path):
    root, worklist = prepare_adjudication(tmp_path)
    assert {item["work_kind"] for item in worklist["items"]} == {
        "mapping", "disposition",
    }
    write_adjudication_verdicts(root, worklist)
    adjudication.apply_done(root, root / "audit", "claims_adjudication")
    ledger = json.loads((root / "audit/_run/handoff_ledger.json").read_text())
    states = {row["id"]: row["state"] for row in ledger["H"] + ledger["X"]}
    mapping = next(item for item in worklist["items"] if item["work_kind"] == "mapping")
    disposition = next(item for item in worklist["items"]
                       if item["work_kind"] == "disposition")
    assert states[mapping["id"]] == mapping["state"]
    assert states[disposition["id"]] == "disposition_accepted"
    adjudication.check_done(root, root / "audit", "claims_adjudication")


def test_adjudicator_mint_requires_containment_and_reserved_range_tier1(tmp_path):
    root, worklist = prepare_adjudication(tmp_path)
    item = next(item for item in worklist["items"] if item["work_kind"] == "mapping")
    base = rb.claims_row(
        "C-8100", context=PROSE_CONTEXT,
        quote="Reference speeds occupy a separate range.",
        text="Reference speeds occupy a separate range.", status="confirmed",
    )
    claim = dict(zip(lint.CLAIMS_COLS, base))
    rows = []
    for current in worklist["items"]:
        if current["id"] == item["id"]:
            rows.append([
                current["id"], current["work_kind"], "reject_and_resolve",
                "the old carrier omitted the assertion", "C-8100",
                *[claim[column] for column in lint.CLAIMS_COLS if column != "Claim ID"],
                current["resolved_anchor"]["source_path"] + ":2",
            ])
        else:
            rows.append([
                current["id"], current["work_kind"], "disposition_accepted",
                "the pointer genuinely has no predicate", *(["—"] *
                    (len(adjudication.ADJUDICATION_VERDICT_COLS) - 4)),
            ])
    path = root / "audit/_run/claims_adjudication_verdicts.md"
    path.write_text(rb.md_table(adjudication.ADJUDICATION_VERDICT_COLS, rows),
                    encoding="utf-8")
    adjudication.apply_done(root, root / "audit", "claims_adjudication")
    claims = (root / "audit/claims_register.md").read_text()
    assert "C-8100" in claims

    # Deliberately break the Tier-1 containment check: clause/caption text
    # elsewhere cannot discharge the assertion.
    rows[0][-1] = item["resolved_anchor"]["source_path"] + ":3"
    path.write_text(rb.md_table(adjudication.ADJUDICATION_VERDICT_COLS, rows),
                    encoding="utf-8")
    with pytest.raises(adjudication.AdjudicationError,
                       match="does not contain|not found|0 occurrences"):
        adjudication._validate_verdicts(root, root / "audit", "claims_adjudication", worklist)


def test_verdict_deletion_and_handwritten_acceptance_are_rederived_tier1(tmp_path):
    root, worklist = prepare_adjudication(tmp_path)
    path = write_adjudication_verdicts(root, worklist)
    adjudication.apply_done(root, root / "audit", "claims_adjudication")
    text = path.read_text()
    first = worklist["items"][0]["id"]
    path.write_text("\n".join(line for line in text.splitlines()
                              if not line.startswith(f"| {first} |")) + "\n",
                    encoding="utf-8")
    with pytest.raises(adjudication.AdjudicationError, match="exactly cover"):
        adjudication.check_done(root, root / "audit", "claims_adjudication")

    # A ledger edit is not a receipt: restore the table without the disposition
    # verdict, then hand-write disposition_accepted in the ledger.
    ledger_path = root / "audit/_run/handoff_ledger.json"
    ledger = json.loads(ledger_path.read_text())
    disposition = next(item for item in worklist["items"]
                       if item["work_kind"] == "disposition")
    for row in ledger["H"] + ledger["X"]:
        if row["id"] == disposition["id"]:
            row["state"] = "disposition_accepted"
    ledger_path.write_text(json.dumps(ledger, indent=2), encoding="utf-8")
    with pytest.raises(adjudication.AdjudicationError, match="exactly cover"):
        adjudication.check_done(root, root / "audit", "claims_adjudication")


def test_zero_work_both_stages_and_lineage_mechanical_carry(tmp_path):
    root, worklist = prepare_adjudication(tmp_path)
    write_adjudication_verdicts(root, worklist)
    cs.start_stage(root, "claims_adjudication")
    adjudication.apply_done(root, root / "audit", "claims_adjudication")
    cs.finish_stage(root, "claims_adjudication", "done")
    freeze_lineage_inputs(root)
    cs.start_stage(root, "claims_adjudication_lineage")
    lineage = adjudication.build_worklist(
        root, root / "audit", "claims_adjudication_lineage")
    assert lineage["items"] == []
    assert (root / "audit/_run/claims_adjudication_lineage_verdicts.md").read_text() \
        == "No lineage verdicts.\n"
    adjudication.apply_done(root, root / "audit", "claims_adjudication_lineage")
    cs.finish_stage(root, "claims_adjudication_lineage", "done")


def test_adjudication_zero_work_uses_explicit_worklist_and_unsharded_verdict(tmp_path):
    root, _worklist = prepare_adjudication(tmp_path)
    empty = {"format_version": 1, "stage": "claims_b3b", "H": [], "X": []}
    for relative in (
            "audit/_run/snapshots/claims_adjudication/handoff_ledger.json",
            "audit/_run/snapshots/claims_b3b/handoff_ledger.json",
            "audit/_run/handoff_ledger.json"):
        (root / relative).write_text(json.dumps(empty, indent=2) + "\n", encoding="utf-8")
    path = root / "audit/_run/claims_adjudication_verdicts.md"
    path.unlink(missing_ok=True)
    built = rb.run_script(
        "claims_adjudication.py", root, "--audit-dir", root / "audit",
        "--stage", "claims_adjudication", "--build-worklist")
    assert built.returncode == 0, built.stdout + built.stderr
    worklist = json.loads(
        (root / "audit/_run/claims_adjudication_worklist.json").read_text())
    assert worklist["items"] == []
    assert path.read_text() == "No adjudication verdicts.\n"
    cs.start_stage(root, "claims_adjudication")
    applied = rb.run_script(
        "claims_adjudication.py", root, "--audit-dir", root / "audit",
        "--stage", "claims_adjudication", "--apply")
    assert applied.returncode == 0, applied.stdout + applied.stderr
    cs.finish_stage(root, "claims_adjudication", "done")


def test_blocked_stage_degrades_pending_items_and_close_requires_decisions(tmp_path):
    root, worklist = prepare_adjudication(tmp_path)
    cs.start_stage(root, "claims_adjudication")
    cs.finish_stage(root, "claims_adjudication", "blocked", "worker died after retry")
    ledger_path = root / "audit/_run/handoff_ledger.json"
    ledger = json.loads(ledger_path.read_text())
    assert {row["state"] for row in ledger["H"] + ledger["X"]} == {"blocked_fallback"}

    freeze_lineage_inputs(root)
    adjudication.build_worklist(root, root / "audit", "claims_adjudication_lineage")
    cs.start_stage(root, "claims_adjudication_lineage")
    cs.finish_stage(root, "claims_adjudication_lineage", "blocked", "no lineage work")
    refused = cli(root, "close-run")
    assert refused.returncode == 1
    assert "has no operator decision" in refused.stderr
    decisions = [{
        "id": row["id"], "decision": "accept_blocked", "reason": "operator reviewed blocker",
        "date": "2026-07-20",
    } for row in sorted(ledger["H"] + ledger["X"], key=lambda value: value["id"])]
    (root / "audit/_run/handoff_blocked_decisions.json").write_text(
        json.dumps(decisions, indent=2), encoding="utf-8")
    assert cli(root, "close-run").returncode == 0


def test_blocked_stage_applies_valid_partial_mint_before_degrading_rest(tmp_path):
    root, worklist = prepare_adjudication(tmp_path)
    item = next(current for current in worklist["items"]
                if current["work_kind"] == "mapping")
    claim = dict(zip(lint.CLAIMS_COLS, rb.claims_row(
        "C-8100", context=PROSE_CONTEXT,
        quote="Reference speeds occupy a separate range.",
        text="Reference speeds occupy a separate range.", status="confirmed",
    )))
    path = root / "audit/_run/claims_adjudication_verdicts.md"
    path.write_text(rb.md_table(adjudication.ADJUDICATION_VERDICT_COLS, [[
        item["id"], item["work_kind"], "reject_and_resolve", "old carrier omitted it",
        "C-8100", *[claim[column] for column in lint.CLAIMS_COLS
                    if column != "Claim ID"],
        item["resolved_anchor"]["source_path"] + ":2",
    ]]), encoding="utf-8")

    cs.start_stage(root, "claims_adjudication")
    cs.finish_stage(root, "claims_adjudication", "blocked", "worker died after partial output")
    ledger = json.loads((root / "audit/_run/handoff_ledger.json").read_text())
    by_id = {row["id"]: row for row in ledger["H"] + ledger["X"]}
    assert by_id[item["id"]]["state"] == "resolved"
    assert by_id[item["id"]]["covering_c_id"] == "C-8100"
    assert all(row["state"] == "blocked_fallback" for key, row in by_id.items()
               if key != item["id"])
    assert "C-8100" in (root / "audit/claims_register.md").read_text()
    adjudication.check_blocked(root, root / "audit", "claims_adjudication")


def test_pending_adjudication_stage_refuses_close_run(tmp_path):
    root, _worklist = prepare_adjudication(tmp_path)
    result = cli(root, "close-run")
    assert result.returncode == 1
    assert "claims_adjudication is pending" in result.stderr


def test_lineage_changed_carrier_requires_verdict_and_refusal_blocks_close(tmp_path):
    root, worklist = prepare_adjudication(tmp_path)
    write_adjudication_verdicts(root, worklist)
    cs.start_stage(root, "claims_adjudication")
    adjudication.apply_done(root, root / "audit", "claims_adjudication")
    cs.finish_stage(root, "claims_adjudication", "done")
    stable = root / "audit/_run/snapshots/claims_b6a"
    stable.mkdir(parents=True, exist_ok=True)
    shutil.copy2(root / "audit/claims_register.md", stable / "claims_register.md")
    shutil.copy2(root / "audit/output_register.md", stable / "output_register.md")
    claims_path = root / "audit/claims_register.md"
    claims_path.write_text(claims_path.read_text().replace(
        "The appendix asserts that reference speeds occupy a separate range.",
        "The carrier dropped the original assertion."), encoding="utf-8")
    freeze_lineage_inputs(root)
    cs.start_stage(root, "claims_adjudication_lineage")
    lineage = adjudication.build_worklist(
        root, root / "audit", "claims_adjudication_lineage")
    assert len(lineage["items"]) == 1
    item = lineage["items"][0]
    verdict_path = root / "audit/_run/claims_adjudication_lineage_verdicts.md"
    verdict_path.write_text(rb.md_table(adjudication.LINEAGE_VERDICT_COLS, [[
        item["id"], "equivalence_refused", "the assertion text was removed",
    ]]), encoding="utf-8")
    adjudication.apply_done(root, root / "audit", "claims_adjudication_lineage")
    cs.finish_stage(root, "claims_adjudication_lineage", "done")
    result = cli(root, "close-run")
    assert result.returncode == 1
    assert "lineage equivalence refused" in result.stderr


def test_b4_requires_adjudicated_handoff_reason(tmp_path):
    root, worklist = prepare_adjudication(tmp_path)
    item = next(item for item in worklist["items"] if item["work_kind"] == "mapping")
    claim_values = dict(zip(lint.CLAIMS_COLS, rb.claims_row(
        "C-8100", context=PROSE_CONTEXT,
        quote="Reference speeds occupy a separate range.",
        text="Reference speeds occupy a separate range.", status="confirmed")))
    other = next(current for current in worklist["items"] if current["id"] != item["id"])
    verdicts = root / "audit/_run/claims_adjudication_verdicts.md"
    verdicts.write_text(rb.md_table(adjudication.ADJUDICATION_VERDICT_COLS, [[
        item["id"], "mapping", "reject_and_resolve", "missing assertion", "C-8100",
        *[claim_values[column] for column in lint.CLAIMS_COLS if column != "Claim ID"],
        item["resolved_anchor"]["source_path"] + ":2",
    ], [
        other["id"], other["work_kind"], "disposition_accepted", "valid pointer disposal",
        *(["—"] * (len(adjudication.ADJUDICATION_VERDICT_COLS) - 4)),
    ]]), encoding="utf-8")
    adjudication.apply_done(root, root / "audit", "claims_adjudication")
    plan = root / "audit/plans/claims_recheck_plan.md"
    plan.write_text(
        "# Plan\n\nSee audit_readme.md for vocabulary.\n\n"
        + rb.md_table(["ID", "Reason", "Likely Evidence"], [[
            "C-8100", "issue-flagged", "paper",
        ]]) + "\n" + rb.md_table(
            ["Cluster ID", "Cluster Name", "Assigned IDs", "Shard File"], [[
                "K1", "adjudication", "C-8100", "audit/_recheck/k1.md",
            ]]),
        encoding="utf-8")
    failed = rb.lint(rb.AuditDir(root), "b4-claims")
    assert failed.returncode == 1
    assert "requires Reason adjudicated_handoff" in failed.stdout
    plan.write_text(plan.read_text().replace("issue-flagged", "adjudicated_handoff"),
                    encoding="utf-8")
    passed = rb.lint(rb.AuditDir(root), "b4-claims")
    assert passed.returncode == 0, passed.stdout


def test_adjudication_scorer_requires_positive_controls_and_no_verdict_control():
    sheet = {
        "false_positive_ceiling": 0,
        "expected_verdicts": [
            {"key": "reject", "obligation_id": "H-0001", "verdict": "reject_and_resolve"},
            {"key": "accept", "obligation_id": "X-0001", "verdict": "capture_confirmed"},
            {"key": "carry", "obligation_id": "X-0002", "verdict": None},
        ],
    }
    scored = score_replay.score_adjudication(sheet, [
        {"Obligation ID": "H-0001", "Verdict": "reject_and_resolve"},
        {"Obligation ID": "X-0001", "Verdict": "capture_confirmed"},
    ])
    assert scored["status"] == "score"
    always_reject = score_replay.score_adjudication(sheet, [
        {"Obligation ID": "H-0001", "Verdict": "reject_and_resolve"},
        {"Obligation ID": "X-0001", "Verdict": "reject_and_resolve"},
    ])
    assert always_reject["status"] == "red"


def test_s706_dual_accept_scores_resolver_valid_handoff(tmp_path):
    root = source_package(tmp_path)
    manifest = json.loads((root / "audit/_run/manifest.json").read_text())
    appendix = next(entry for entry in manifest["paper_source_set"]
                    if entry["source_path"].endswith("appendix.tex"))
    expected = {
        "key": "truck-speed-claim", "target_anchor": appendix["source_path"] + ":2",
        "target_quote": "Reference speeds occupy a separate range.",
    }
    handoff = {
        "H ID": "H-0001", "Anchor": expected["target_anchor"],
        "Quote": expected["target_quote"], "Asserted Substance": "reference speeds differ",
        "Referenced Objects": "Figure A2",
    }
    sheet = {"expected_claim_obligations": [expected]}
    scored, routes = score_replay.score_claim_obligations(
        sheet, root, [], [handoff], [])
    assert scored[0]["status"] == "score"
    assert ("filed_h", "H-0001") in routes


def test_b9_exports_and_exactly_lints_handoff_ledger_sheet(tmp_path):
    root, worklist = prepare_adjudication(tmp_path)
    write_adjudication_verdicts(root, worklist)
    adjudication.apply_done(root, root / "audit", "claims_adjudication")
    audit = root / "audit"
    (audit / "code_error_register.md").write_text(
        rb.register_text("Errors", lint.ERROR_COLS, []), encoding="utf-8")
    for stream in ("claims", "code"):
        (audit / f"late_observations_{stream}.md").write_text(
            f"# Late observations — {stream}\n\nNo late observations.\n\n"
            "## Dispositions\n\nNo dispositions.\n", encoding="utf-8")
    staging = audit / "_staging"
    staging.mkdir(exist_ok=True)
    shutil.copy2(audit / "claims_register.md", staging / "claims_register.md")
    shutil.copy2(audit / "code_error_register.md", staging / "code_error_register.md")
    result = rb.run_script(
        "export_xlsx.py", "--audit-dir", audit, "--mode", "replication")
    assert result.returncode == 0, result.stdout + result.stderr
    from openpyxl import load_workbook
    workbook = load_workbook(audit / "code_review.xlsx", read_only=True)
    assert "Handoff ledger" in workbook.sheetnames
    passed = rb.lint(rb.AuditDir(root), "b9")
    assert passed.returncode == 0, passed.stdout + passed.stderr
    ledger = json.loads((audit / "_run/handoff_ledger.json").read_text())
    ledger["H"] = []
    (audit / "_run/handoff_ledger.json").write_text(
        json.dumps(ledger, indent=2), encoding="utf-8")
    failed = rb.lint(rb.AuditDir(root), "b9")
    assert failed.returncode == 1
    assert "does not exactly match" in failed.stdout


# ------------------------------------------------------- U7b phase-D repairs


def _mint_verdict_rows(worklist, item, mint_id, covering_suffix=":2"):
    claim = dict(zip(lint.CLAIMS_COLS, rb.claims_row(
        mint_id, context=PROSE_CONTEXT,
        quote="Reference speeds occupy a separate range.",
        text="Reference speeds occupy a separate range.", status="confirmed",
    )))
    rows = []
    for current in worklist["items"]:
        if current["id"] == item["id"]:
            rows.append([
                current["id"], current["work_kind"], "reject_and_resolve",
                "the old carrier omitted the assertion", mint_id,
                *[claim[column] for column in lint.CLAIMS_COLS
                  if column != "Claim ID"],
                current["resolved_anchor"]["source_path"] + covering_suffix,
            ])
        else:
            rows.append([
                current["id"], current["work_kind"], "disposition_accepted",
                "the pointer genuinely has no predicate",
                *(["—"] * (len(adjudication.ADJUDICATION_VERDICT_COLS) - 4)),
            ])
    return rows


def test_adjudicator_mint_outside_reserved_range_refuses(tmp_path):
    root, worklist = prepare_adjudication(tmp_path)
    item = next(item for item in worklist["items"] if item["work_kind"] == "mapping")
    rows = _mint_verdict_rows(worklist, item, "C-8150")
    (root / "audit/_run/claims_adjudication_verdicts.md").write_text(
        rb.md_table(adjudication.ADJUDICATION_VERDICT_COLS, rows), encoding="utf-8")
    with pytest.raises(adjudication.AdjudicationError,
                       match="outside adjudication range"):
        adjudication._validate_verdicts(
            root, root / "audit", "claims_adjudication", worklist)


def test_certification_and_verify_run_refuse_sabotaged_verdicts_tier1(tmp_path):
    root, worklist = prepare_adjudication(tmp_path)
    item = next(item for item in worklist["items"] if item["work_kind"] == "mapping")
    path = root / "audit/_run/claims_adjudication_verdicts.md"
    cs.start_stage(root, "claims_adjudication")
    # Incomplete verdict table refuses at finish, through the production surface.
    path.write_text(rb.md_table(
        adjudication.ADJUDICATION_VERDICT_COLS,
        _mint_verdict_rows(worklist, item, "C-8100")[:-1]), encoding="utf-8")
    with pytest.raises(cs.CertificationError, match="exactly cover"):
        cs.finish_stage(root, "claims_adjudication", "done")
    # A non-containing mint refuses at finish.
    path.write_text(rb.md_table(
        adjudication.ADJUDICATION_VERDICT_COLS,
        _mint_verdict_rows(worklist, item, "C-8100", ":3")), encoding="utf-8")
    with pytest.raises(cs.CertificationError,
                       match="does not contain|not found|0 occurrences"):
        cs.finish_stage(root, "claims_adjudication", "done")
    # Healthy evidence certifies and verify-run stays quiet.
    healthy = rb.md_table(adjudication.ADJUDICATION_VERDICT_COLS,
                          _mint_verdict_rows(worklist, item, "C-8100"))
    path.write_text(healthy, encoding="utf-8")
    adjudication.apply_done(root, root / "audit", "claims_adjudication")
    cs.finish_stage(root, "claims_adjudication", "done")
    good = cli(root, "verify-run")
    assert good.returncode == 0, good.stdout + good.stderr
    # Post-certification verdict deletion refuses at verify-run.
    path.write_text("\n".join(line for line in healthy.splitlines()
                              if not line.startswith(f"| {item['id']} |")) + "\n",
                    encoding="utf-8")
    deleted = cli(root, "verify-run")
    assert deleted.returncode == 1
    assert "exactly cover" in deleted.stdout + deleted.stderr
    # Post-certification containment corruption refuses at verify-run.
    path.write_text(healthy.replace(
        item["resolved_anchor"]["source_path"] + ":2",
        item["resolved_anchor"]["source_path"] + ":3"), encoding="utf-8")
    corrupted = cli(root, "verify-run")
    assert corrupted.returncode == 1
    assert "CLAIMS ADJUDICATION REFUSED" in corrupted.stdout + corrupted.stderr


def test_lineage_certification_and_verify_run_refuse_sabotage_tier1(tmp_path):
    root, worklist = prepare_adjudication(tmp_path)
    write_adjudication_verdicts(root, worklist)
    cs.start_stage(root, "claims_adjudication")
    adjudication.apply_done(root, root / "audit", "claims_adjudication")
    cs.finish_stage(root, "claims_adjudication", "done")
    stable = root / "audit/_run/snapshots/claims_b6a"
    stable.mkdir(parents=True, exist_ok=True)
    shutil.copy2(root / "audit/claims_register.md", stable / "claims_register.md")
    shutil.copy2(root / "audit/output_register.md", stable / "output_register.md")
    claims_path = root / "audit/claims_register.md"
    claims_path.write_text(claims_path.read_text().replace(
        "The appendix asserts that reference speeds occupy a separate range.",
        "The carrier dropped the original assertion."), encoding="utf-8")
    freeze_lineage_inputs(root)
    cs.start_stage(root, "claims_adjudication_lineage")
    lineage = adjudication.build_worklist(
        root, root / "audit", "claims_adjudication_lineage")
    assert len(lineage["items"]) == 1
    item = lineage["items"][0]
    verdict_path = root / "audit/_run/claims_adjudication_lineage_verdicts.md"
    # A required verdict deleted before finish refuses at finish.
    verdict_path.write_text("No lineage verdicts.\n", encoding="utf-8")
    with pytest.raises(cs.CertificationError, match="exactly cover"):
        cs.finish_stage(root, "claims_adjudication_lineage", "done")
    healthy = rb.md_table(adjudication.LINEAGE_VERDICT_COLS, [[
        item["id"], "equivalence_confirmed", "the carrier still bears the assertion",
    ]])
    verdict_path.write_text(healthy, encoding="utf-8")
    adjudication.apply_done(root, root / "audit", "claims_adjudication_lineage")
    cs.finish_stage(root, "claims_adjudication_lineage", "done")
    good = cli(root, "verify-run")
    assert good.returncode == 0, good.stdout + good.stderr
    # Post-certification verdict mutation refuses at verify-run.
    verdict_path.write_text(healthy.replace(
        "equivalence_confirmed", "capture_confirmed"), encoding="utf-8")
    mutated = cli(root, "verify-run")
    assert mutated.returncode == 1
    assert "invalid lineage verdict" in mutated.stdout + mutated.stderr


def test_dead_carrier_confirmation_refused_and_close_run_rederives_tier1(tmp_path):
    root, worklist = prepare_adjudication(tmp_path)
    write_adjudication_verdicts(root, worklist)
    cs.start_stage(root, "claims_adjudication")
    adjudication.apply_done(root, root / "audit", "claims_adjudication")
    cs.finish_stage(root, "claims_adjudication", "done")
    stable = root / "audit/_run/snapshots/claims_b6a"
    stable.mkdir(parents=True, exist_ok=True)
    shutil.copy2(root / "audit/claims_register.md", stable / "claims_register.md")
    shutil.copy2(root / "audit/output_register.md", stable / "output_register.md")
    claims_path = root / "audit/claims_register.md"
    claims_path.write_text("\n".join(
        line for line in claims_path.read_text().splitlines()
        if not line.startswith("| C-2000 |")) + "\n", encoding="utf-8")
    freeze_lineage_inputs(root)
    cs.start_stage(root, "claims_adjudication_lineage")
    lineage = adjudication.build_worklist(
        root, root / "audit", "claims_adjudication_lineage")
    assert len(lineage["items"]) == 1
    assert lineage["items"][0]["terminal_c_id"] is None
    verdict_path = root / "audit/_run/claims_adjudication_lineage_verdicts.md"
    verdict_path.write_text(rb.md_table(adjudication.LINEAGE_VERDICT_COLS, [[
        lineage["items"][0]["id"], "equivalence_confirmed", "looks equivalent to me",
    ]]), encoding="utf-8")
    with pytest.raises(cs.CertificationError, match="no terminal live carrier"):
        cs.finish_stage(root, "claims_adjudication_lineage", "done")
    # A forged done status cannot slip a hand-confirmed dead carrier past
    # close-run: the dead-end refusal is re-derived from the worklist artifact.
    manifest_path = root / "audit/_run/manifest.json"
    manifest = json.loads(manifest_path.read_text())
    manifest["stages"]["claims_adjudication_lineage"]["status"] = "done"
    manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    result = cli(root, "close-run")
    assert result.returncode == 1
    assert "lineage dead-end with no live carrier" in result.stderr


def test_score_run_adjudication_requires_exact_schema_tables(tmp_path):
    scenario = {
        "format_version": 1, "stage": "claims_adjudication",
        "route": "deterministic_stage", "scoring_mode": "adjudication",
        "promised_outputs": ["audit/_run/claims_adjudication_verdicts.md"],
        "answer_sheet": "answers/sheet.json", "runs": 1,
    }
    scenario_path = tmp_path / "opaque-x1.json"
    scenario_path.write_text(json.dumps(scenario), encoding="utf-8")
    sheet = {
        "format_version": 1,
        "mechanism_schema_version": mech.MECHANISM_SCHEMA_VERSION,
        "disposition_complete": True, "false_positive_ceiling": 0,
        "output_contract": {
            "verdict_paths": ["audit/_run/claims_adjudication_verdicts.md"],
        },
        "expected_verdicts": [
            {"key": "accept", "obligation_id": "H-0001",
             "verdict": "capture_confirmed"},
        ],
    }
    sheet_path = tmp_path / "sheet.json"
    sheet_path.write_text(json.dumps(sheet), encoding="utf-8")
    run_dir = tmp_path / "run-001"
    out = run_dir / "sandbox/audit/_run"
    out.mkdir(parents=True)
    (run_dir / "replay-record.json").write_text(json.dumps({
        "format_version": 1, "scenario_id": "opaque-x1",
        "stage": "claims_adjudication", "route": "deterministic_stage",
        "run_index": 1, "timestamp": "2026-07-20T00:00:00+00:00",
        "identity": {
            "model_requested": "not-applicable", "model_reported": "not-applicable",
            "cli_version": "not-applicable", "code_commit": "a" * 40,
            "code_dirty": True, "requested_effort": "not-applicable",
            "observed_effort": "not-applicable",
            "mechanism_schema_version": mech.MECHANISM_SCHEMA_VERSION,
        },
        "promised_outputs_found": ["audit/_run/claims_adjudication_verdicts.md"],
    }, indent=2), encoding="utf-8")
    verdict_path = out / "claims_adjudication_verdicts.md"
    # An invalid two-column artifact must score red, not green.
    verdict_path.write_text(
        "| Obligation ID | Verdict |\n| --- | --- |\n"
        "| H-0001 | capture_confirmed |\n", encoding="utf-8")
    red = score_replay.score_run(scenario_path, scenario, sheet_path, sheet, run_dir)
    assert red["status"] == "red"
    assert any("exact" in problem for problem in red.get("format_problems", []))
    # The exact first-stage schema scores green.
    verdict_path.write_text(rb.md_table(adjudication.ADJUDICATION_VERDICT_COLS, [[
        "H-0001", "mapping", "capture_confirmed", "checked capture",
        *(["—"] * (len(adjudication.ADJUDICATION_VERDICT_COLS) - 4)),
    ]]), encoding="utf-8")
    green = score_replay.score_run(scenario_path, scenario, sheet_path, sheet, run_dir)
    assert green["status"] == "score"
